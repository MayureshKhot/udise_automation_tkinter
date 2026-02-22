#!/usr/bin/env python3
"""Tkinter app for interactive UDISE OCR tuning and batch processing.

This GUI lets you:
- Load Excel + model
- Select ROI once
- Preview per-digit preprocessing results
- Tune preprocessing settings live
- Run batch inference and write output Excel
"""

from __future__ import annotations

import json
import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import cv2
import numpy as np
import requests
import tkinter as tk
from PIL import Image, ImageTk
from openpyxl import load_workbook
from tkinter import filedialog, messagebox, ttk

from udise_ocr_pipeline import (
    decode_image_from_source,
    get_model_runner,
    load_model_config,
    parse_manual_roi,
    resolve_column,
    resolve_sheet,
    save_roi,
    sheet_headers,
    split_digits_equal,
    write_failure_log,
    crop_roi,
)

DEFAULT_EXCEL = "test_udise.xlsx"
DEFAULT_ROI_FILE = "udise_roi.json"


@dataclass
class PreprocessSettings:
    image_size: int = 28
    invert: bool = True
    blur_kernel: int = 3
    threshold_mode: str = "otsu"  # otsu | adaptive
    adaptive_block_size: int = 21
    adaptive_c: int = 10
    tight_crop: bool = True
    bbox_padding: int = 2


def ensure_odd(value: int, minimum: int) -> int:
    v = max(minimum, int(value))
    if v % 2 == 0:
        v += 1
    return v


def preprocess_digit_custom(digit_bgr: np.ndarray, s: PreprocessSettings) -> Tuple[np.ndarray, np.ndarray]:
    gray = cv2.cvtColor(digit_bgr, cv2.COLOR_BGR2GRAY)

    blur_k = ensure_odd(s.blur_kernel, 1)
    if blur_k > 1:
        gray = cv2.GaussianBlur(gray, (blur_k, blur_k), 0)

    if s.threshold_mode == "adaptive":
        block_size = ensure_odd(s.adaptive_block_size, 3)
        th = cv2.adaptiveThreshold(
            gray,
            255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY,
            block_size,
            int(s.adaptive_c),
        )
    else:
        _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    if s.invert:
        th = cv2.bitwise_not(th)

    proc = th
    if s.tight_crop:
        fg_mask = proc if s.invert else (255 - proc)
        nz = cv2.findNonZero(fg_mask)
        if nz is not None:
            x, y, w, h = cv2.boundingRect(nz)
            pad = max(0, int(s.bbox_padding))
            x0 = max(0, x - pad)
            y0 = max(0, y - pad)
            x1 = min(proc.shape[1], x + w + pad)
            y1 = min(proc.shape[0], y + h + pad)
            proc = proc[y0:y1, x0:x1]

    h, w = proc.shape[:2]
    image_size = max(8, int(s.image_size))
    scale = image_size / float(max(h, w))
    new_w = max(1, int(round(w * scale)))
    new_h = max(1, int(round(h * scale)))
    resized = cv2.resize(proc, (new_w, new_h), interpolation=cv2.INTER_AREA)

    bg = 0 if s.invert else 255
    canvas = np.full((image_size, image_size), bg, dtype=np.uint8)
    y0 = (image_size - new_h) // 2
    x0 = (image_size - new_w) // 2
    canvas[y0 : y0 + new_h, x0 : x0 + new_w] = resized

    arr = canvas.astype(np.float32) / 255.0
    arr = np.expand_dims(arr, axis=-1)
    return arr, canvas


class UDISEOCRApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("UDISE OCR Tuning App")
        self.geometry("1450x920")

        self.runner = None
        self.roi: Optional[Tuple[int, int, int, int]] = None
        self.current_image_bgr: Optional[np.ndarray] = None
        self.current_row: Optional[int] = None
        self.current_source: Optional[str] = None

        self.photo_refs: Dict[str, ImageTk.PhotoImage] = {}
        self.digit_photo_refs: List[ImageTk.PhotoImage] = []

        self._build_vars()
        self._build_ui()

    def _build_vars(self) -> None:
        self.excel_var = tk.StringVar(value=DEFAULT_EXCEL)
        self.output_excel_var = tk.StringVar(value="")
        self.sheet_var = tk.StringVar(value="")
        self.url_col_var = tk.StringVar(value="Original Image Path")
        self.out_col_var = tk.StringVar(value="UDISE Found")

        self.framework_var = tk.StringVar(value="keras")
        self.model_path_var = tk.StringVar(value="")
        self.model_config_var = tk.StringVar(value="")

        self.roi_file_var = tk.StringVar(value=DEFAULT_ROI_FILE)
        self.roi_text_var = tk.StringVar(value="")

        self.timeout_var = tk.IntVar(value=20)
        self.retries_var = tk.IntVar(value=2)
        self.backoff_var = tk.DoubleVar(value=1.5)
        self.start_row_var = tk.IntVar(value=2)
        self.max_rows_var = tk.IntVar(value=0)
        self.overwrite_var = tk.BooleanVar(value=False)
        self.save_every_var = tk.IntVar(value=50)

        self.sample_row_var = tk.IntVar(value=2)

        self.image_size_var = tk.IntVar(value=28)
        self.invert_var = tk.BooleanVar(value=True)
        self.blur_var = tk.IntVar(value=3)
        self.th_mode_var = tk.StringVar(value="otsu")
        self.ad_block_var = tk.IntVar(value=21)
        self.ad_c_var = tk.IntVar(value=10)
        self.tight_crop_var = tk.BooleanVar(value=True)
        self.pad_var = tk.IntVar(value=2)
        self.digit_count_var = tk.IntVar(value=11)

        self.status_var = tk.StringVar(value="Ready")

    def _build_ui(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        top = ttk.Frame(self, padding=8)
        top.grid(row=0, column=0, sticky="ew")
        top.columnconfigure(1, weight=1)

        self._row_file(top, 0, "Excel", self.excel_var, self.pick_excel)
        self._row_file(top, 1, "Output Excel", self.output_excel_var, self.pick_output_excel)
        self._row_file(top, 2, "Model", self.model_path_var, self.pick_model)
        self._row_file(top, 3, "Model Config", self.model_config_var, self.pick_model_config)

        meta = ttk.Frame(top)
        meta.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(6, 0))
        for i in range(8):
            meta.columnconfigure(i, weight=1)

        ttk.Label(meta, text="Sheet").grid(row=0, column=0, sticky="w")
        ttk.Entry(meta, textvariable=self.sheet_var, width=14).grid(row=0, column=1, sticky="ew")

        ttk.Label(meta, text="URL Col/Header").grid(row=0, column=2, sticky="w")
        ttk.Entry(meta, textvariable=self.url_col_var, width=20).grid(row=0, column=3, sticky="ew")

        ttk.Label(meta, text="Output Col/Header").grid(row=0, column=4, sticky="w")
        ttk.Entry(meta, textvariable=self.out_col_var, width=20).grid(row=0, column=5, sticky="ew")

        ttk.Label(meta, text="Framework").grid(row=0, column=6, sticky="w")
        ttk.Combobox(meta, textvariable=self.framework_var, state="readonly", values=["keras", "torch"], width=10).grid(row=0, column=7, sticky="w")

        main = ttk.Panedwindow(self, orient=tk.HORIZONTAL)
        main.grid(row=1, column=0, sticky="nsew")

        left = ttk.Frame(main, padding=8)
        right = ttk.Frame(main, padding=8)
        main.add(left, weight=2)
        main.add(right, weight=3)

        self._build_left_panel(left)
        self._build_right_panel(right)

        status_bar = ttk.Label(self, textvariable=self.status_var, anchor="w", padding=6)
        status_bar.grid(row=2, column=0, sticky="ew")

    def _row_file(self, parent, row, label, var, cmd):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(0, 8), pady=2)
        ttk.Entry(parent, textvariable=var).grid(row=row, column=1, sticky="ew", pady=2)
        ttk.Button(parent, text="Browse", command=cmd).grid(row=row, column=2, sticky="w", padx=(8, 0), pady=2)

    def _build_left_panel(self, parent):
        parent.columnconfigure(0, weight=1)

        roi_box = ttk.LabelFrame(parent, text="ROI & Sample", padding=8)
        roi_box.grid(row=0, column=0, sticky="ew")
        for i in range(6):
            roi_box.columnconfigure(i, weight=1)

        ttk.Label(roi_box, text="ROI File").grid(row=0, column=0, sticky="w")
        ttk.Entry(roi_box, textvariable=self.roi_file_var).grid(row=0, column=1, columnspan=4, sticky="ew")
        ttk.Button(roi_box, text="Load ROI", command=self.load_roi_from_file).grid(row=0, column=5, sticky="ew")

        ttk.Label(roi_box, text="ROI x,y,w,h").grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(roi_box, textvariable=self.roi_text_var).grid(row=1, column=1, columnspan=3, sticky="ew", pady=(6, 0))
        ttk.Button(roi_box, text="Apply ROI", command=self.apply_roi_text).grid(row=1, column=4, sticky="ew", pady=(6, 0))
        ttk.Button(roi_box, text="Pick ROI", command=self.pick_roi_from_sample).grid(row=1, column=5, sticky="ew", pady=(6, 0))

        ttk.Label(roi_box, text="Sample Row").grid(row=2, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(roi_box, textvariable=self.sample_row_var, width=8).grid(row=2, column=1, sticky="w", pady=(6, 0))
        ttk.Button(roi_box, text="Load Sample Row", command=self.load_sample_row).grid(row=2, column=2, columnspan=2, sticky="ew", pady=(6, 0))
        ttk.Button(roi_box, text="Find Next Pending", command=self.find_next_pending).grid(row=2, column=4, columnspan=2, sticky="ew", pady=(6, 0))

        pp_box = ttk.LabelFrame(parent, text="Preprocess Controls", padding=8)
        pp_box.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        for i in range(4):
            pp_box.columnconfigure(i, weight=1)

        ttk.Label(pp_box, text="Digit Count").grid(row=0, column=0, sticky="w")
        ttk.Spinbox(pp_box, from_=1, to=20, textvariable=self.digit_count_var, width=8).grid(row=0, column=1, sticky="w")

        ttk.Label(pp_box, text="Image Size").grid(row=0, column=2, sticky="w")
        ttk.Spinbox(pp_box, from_=8, to=96, textvariable=self.image_size_var, width=8).grid(row=0, column=3, sticky="w")

        ttk.Checkbutton(pp_box, text="Invert", variable=self.invert_var).grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Checkbutton(pp_box, text="Tight Crop", variable=self.tight_crop_var).grid(row=1, column=1, sticky="w", pady=(6, 0))

        ttk.Label(pp_box, text="Blur Kernel").grid(row=1, column=2, sticky="w", pady=(6, 0))
        ttk.Spinbox(pp_box, from_=1, to=11, increment=2, textvariable=self.blur_var, width=8).grid(row=1, column=3, sticky="w", pady=(6, 0))

        ttk.Label(pp_box, text="Threshold").grid(row=2, column=0, sticky="w", pady=(6, 0))
        ttk.Combobox(pp_box, textvariable=self.th_mode_var, state="readonly", values=["otsu", "adaptive"], width=12).grid(row=2, column=1, sticky="w", pady=(6, 0))

        ttk.Label(pp_box, text="Adaptive Block").grid(row=2, column=2, sticky="w", pady=(6, 0))
        ttk.Spinbox(pp_box, from_=3, to=51, increment=2, textvariable=self.ad_block_var, width=8).grid(row=2, column=3, sticky="w", pady=(6, 0))

        ttk.Label(pp_box, text="Adaptive C").grid(row=3, column=0, sticky="w", pady=(6, 0))
        ttk.Spinbox(pp_box, from_=-20, to=30, textvariable=self.ad_c_var, width=8).grid(row=3, column=1, sticky="w", pady=(6, 0))

        ttk.Label(pp_box, text="BBox Padding").grid(row=3, column=2, sticky="w", pady=(6, 0))
        ttk.Spinbox(pp_box, from_=0, to=20, textvariable=self.pad_var, width=8).grid(row=3, column=3, sticky="w", pady=(6, 0))

        actions = ttk.Frame(parent)
        actions.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        actions.columnconfigure(0, weight=1)
        actions.columnconfigure(1, weight=1)
        actions.columnconfigure(2, weight=1)

        ttk.Button(actions, text="Load Model", command=self.load_model).grid(row=0, column=0, sticky="ew", padx=2)
        ttk.Button(actions, text="Refresh Preview", command=self.refresh_preview).grid(row=0, column=1, sticky="ew", padx=2)
        ttk.Button(actions, text="Run Batch", command=self.run_batch_async).grid(row=0, column=2, sticky="ew", padx=2)

        run_cfg = ttk.LabelFrame(parent, text="Batch Settings", padding=8)
        run_cfg.grid(row=3, column=0, sticky="ew", pady=(8, 0))
        for i in range(4):
            run_cfg.columnconfigure(i, weight=1)

        ttk.Label(run_cfg, text="Start Row").grid(row=0, column=0, sticky="w")
        ttk.Entry(run_cfg, textvariable=self.start_row_var, width=8).grid(row=0, column=1, sticky="w")

        ttk.Label(run_cfg, text="Max Rows (0=all)").grid(row=0, column=2, sticky="w")
        ttk.Entry(run_cfg, textvariable=self.max_rows_var, width=8).grid(row=0, column=3, sticky="w")

        ttk.Label(run_cfg, text="Timeout").grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(run_cfg, textvariable=self.timeout_var, width=8).grid(row=1, column=1, sticky="w", pady=(6, 0))

        ttk.Label(run_cfg, text="Retries").grid(row=1, column=2, sticky="w", pady=(6, 0))
        ttk.Entry(run_cfg, textvariable=self.retries_var, width=8).grid(row=1, column=3, sticky="w", pady=(6, 0))

        ttk.Label(run_cfg, text="Backoff").grid(row=2, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(run_cfg, textvariable=self.backoff_var, width=8).grid(row=2, column=1, sticky="w", pady=(6, 0))

        ttk.Label(run_cfg, text="Save Every").grid(row=2, column=2, sticky="w", pady=(6, 0))
        ttk.Entry(run_cfg, textvariable=self.save_every_var, width=8).grid(row=2, column=3, sticky="w", pady=(6, 0))

        ttk.Checkbutton(run_cfg, text="Overwrite Existing", variable=self.overwrite_var).grid(row=3, column=0, columnspan=2, sticky="w", pady=(6, 0))

    def _build_right_panel(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(2, weight=1)

        preview = ttk.LabelFrame(parent, text="Image Preview", padding=8)
        preview.grid(row=0, column=0, sticky="nsew")
        preview.columnconfigure(0, weight=1)
        preview.columnconfigure(1, weight=1)

        self.full_img_label = ttk.Label(preview, text="Full image")
        self.full_img_label.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        self.roi_img_label = ttk.Label(preview, text="ROI crop")
        self.roi_img_label.grid(row=0, column=1, sticky="nsew")

        digits = ttk.LabelFrame(parent, text="Preprocessed Digits", padding=8)
        digits.grid(row=1, column=0, sticky="ew", pady=(8, 0))

        self.digit_frame = ttk.Frame(digits)
        self.digit_frame.grid(row=0, column=0, sticky="ew")

        self.pred_var = tk.StringVar(value="Prediction: -")
        ttk.Label(digits, textvariable=self.pred_var, font=("TkDefaultFont", 11, "bold")).grid(row=1, column=0, sticky="w", pady=(6, 0))

        logs = ttk.LabelFrame(parent, text="Logs", padding=8)
        logs.grid(row=2, column=0, sticky="nsew", pady=(8, 0))
        logs.rowconfigure(0, weight=1)
        logs.columnconfigure(0, weight=1)

        self.log_text = tk.Text(logs, wrap="word", height=14)
        self.log_text.grid(row=0, column=0, sticky="nsew")

        scroll = ttk.Scrollbar(logs, orient="vertical", command=self.log_text.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.log_text.configure(yscrollcommand=scroll.set)

    def log(self, msg: str) -> None:
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.status_var.set(msg)
        self.update_idletasks()

    def pick_excel(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if p:
            self.excel_var.set(p)
            if not self.output_excel_var.get().strip():
                out = Path(p).with_name(f"{Path(p).stem}_with_udise.xlsx")
                self.output_excel_var.set(str(out))

    def pick_output_excel(self):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if p:
            self.output_excel_var.set(p)

    def pick_model(self):
        p = filedialog.askopenfilename(filetypes=[("Model", "*.keras *.h5 *.pt *.pth"), ("All", "*.*")])
        if p:
            self.model_path_var.set(p)
            json_guess = Path(str(p) + ".json")
            if json_guess.exists() and not self.model_config_var.get().strip():
                self.model_config_var.set(str(json_guess))

    def pick_model_config(self):
        p = filedialog.askopenfilename(filetypes=[("JSON", "*.json")])
        if p:
            self.model_config_var.set(p)

    def load_roi_from_file(self):
        p = Path(self.roi_file_var.get().strip())
        if not p.exists():
            messagebox.showerror("Error", f"ROI file not found: {p}")
            return
        data = json.loads(p.read_text(encoding="utf-8"))
        self.roi = (int(data["x"]), int(data["y"]), int(data["w"]), int(data["h"]))
        self.roi_text_var.set(f"{self.roi[0]},{self.roi[1]},{self.roi[2]},{self.roi[3]}")
        self.log(f"ROI loaded from {p}")
        self.refresh_preview()

    def apply_roi_text(self):
        txt = self.roi_text_var.get().strip()
        try:
            r = parse_manual_roi(txt)
        except Exception as exc:
            messagebox.showerror("ROI Error", str(exc))
            return
        self.roi = (r.x, r.y, r.w, r.h)
        save_roi(self.roi_file_var.get().strip(), r)
        self.log(f"ROI set to {self.roi}")
        self.refresh_preview()

    def pick_roi_from_sample(self):
        if self.current_image_bgr is None:
            messagebox.showerror("No image", "Load a sample row first.")
            return
        x, y, w, h = cv2.selectROI(
            "Select UDISE ROI (Enter/Space confirm, c cancel)",
            self.current_image_bgr.copy(),
            showCrosshair=True,
            fromCenter=False,
        )
        cv2.destroyAllWindows()
        if w <= 0 or h <= 0:
            return
        self.roi = (int(x), int(y), int(w), int(h))
        self.roi_text_var.set(f"{x},{y},{w},{h}")
        save_roi(self.roi_file_var.get().strip(), parse_manual_roi(self.roi_text_var.get()))
        self.log(f"ROI selected: {self.roi}")
        self.refresh_preview()

    def _open_sheet(self):
        excel_path = Path(self.excel_var.get().strip())
        if not excel_path.exists():
            raise RuntimeError(f"Excel not found: {excel_path}")
        wb = load_workbook(excel_path)
        ws = resolve_sheet(wb, self.sheet_var.get().strip() or None)
        return wb, ws

    def _resolve_cols(self, ws):
        url_ref = self.url_col_var.get().strip() or None
        out_ref = self.out_col_var.get().strip() or None
        url_col = resolve_column(ws, url_ref, ["original image path", "image path on server", "image url", "url"], "H")
        out_col = resolve_column(ws, out_ref, ["udise found", "udise", "udise code"], "E", allow_create=True)
        return url_col, out_col

    def find_next_pending(self):
        try:
            wb, ws = self._open_sheet()
            url_col, out_col = self._resolve_cols(ws)
            start = max(2, int(self.sample_row_var.get()))
            for row in range(start, ws.max_row + 1):
                url = ws.cell(row=row, column=url_col).value
                out = ws.cell(row=row, column=out_col).value
                if url and (out is None or not str(out).strip()):
                    self.sample_row_var.set(row)
                    self.log(f"Next pending row: {row}")
                    break
            wb.close()
        except Exception as exc:
            messagebox.showerror("Error", str(exc))

    def load_model(self):
        try:
            model_path = self.model_path_var.get().strip()
            if not model_path:
                raise RuntimeError("Select a model path.")
            framework = self.framework_var.get().strip()
            config_path = self.model_config_var.get().strip() or None
            cfg = load_model_config(config_path, model_path)

            if cfg.get("framework") and str(cfg["framework"]) != framework:
                raise RuntimeError(
                    f"Framework mismatch. UI={framework}, config={cfg['framework']}"
                )

            if "image_size" in cfg:
                self.image_size_var.set(int(cfg["image_size"]))
            if "invert" in cfg:
                self.invert_var.set(bool(cfg["invert"]))
            if "digit_count" in cfg:
                self.digit_count_var.set(int(cfg["digit_count"]))

            self.runner = get_model_runner(framework, model_path)
            self.log(f"Model loaded: {model_path}")
            self.refresh_preview()
        except Exception as exc:
            messagebox.showerror("Model Error", str(exc))

    def load_sample_row(self):
        try:
            wb, ws = self._open_sheet()
            url_col, _ = self._resolve_cols(ws)
            row = int(self.sample_row_var.get())
            source = ws.cell(row=row, column=url_col).value
            wb.close()
            if not source or not str(source).strip():
                raise RuntimeError(f"Row {row} has empty source in URL column.")

            self.current_row = row
            self.current_source = str(source).strip()
            session = requests.Session()
            try:
                self.current_image_bgr = decode_image_from_source(
                    source=self.current_source,
                    timeout=int(self.timeout_var.get()),
                    retries=int(self.retries_var.get()),
                    retry_backoff=float(self.backoff_var.get()),
                    session=session,
                )
            finally:
                session.close()

            self.log(f"Loaded sample row {row}")
            self.refresh_preview()
        except Exception as exc:
            messagebox.showerror("Sample Error", str(exc))

    def _settings(self) -> PreprocessSettings:
        return PreprocessSettings(
            image_size=max(8, int(self.image_size_var.get())),
            invert=bool(self.invert_var.get()),
            blur_kernel=max(1, int(self.blur_var.get())),
            threshold_mode=self.th_mode_var.get().strip() or "otsu",
            adaptive_block_size=max(3, int(self.ad_block_var.get())),
            adaptive_c=int(self.ad_c_var.get()),
            tight_crop=bool(self.tight_crop_var.get()),
            bbox_padding=max(0, int(self.pad_var.get())),
        )

    def _to_photo(self, bgr_or_gray: np.ndarray, max_size=(540, 360)):
        if bgr_or_gray.ndim == 2:
            rgb = cv2.cvtColor(bgr_or_gray, cv2.COLOR_GRAY2RGB)
        else:
            rgb = cv2.cvtColor(bgr_or_gray, cv2.COLOR_BGR2RGB)
        h, w = rgb.shape[:2]
        scale = min(max_size[0] / w, max_size[1] / h, 1.0)
        if scale < 1.0:
            rgb = cv2.resize(rgb, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_AREA)
        return ImageTk.PhotoImage(Image.fromarray(rgb))

    def refresh_preview(self):
        if self.current_image_bgr is None:
            return
        if self.roi is None:
            full = self._to_photo(self.current_image_bgr)
            self.full_img_label.configure(image=full, text="")
            self.photo_refs["full"] = full
            self.roi_img_label.configure(text="Pick/apply ROI to preview", image="")
            return

        try:
            roi_obj = parse_manual_roi(
                f"{self.roi[0]},{self.roi[1]},{self.roi[2]},{self.roi[3]}"
            )
            roi_crop = crop_roi(self.current_image_bgr, roi_obj)
            digit_cells = split_digits_equal(roi_crop, int(self.digit_count_var.get()))
            s = self._settings()

            batch_list = []
            preview_list = []
            for cell in digit_cells:
                arr, preview = preprocess_digit_custom(cell, s)
                batch_list.append(arr)
                preview_list.append(preview)

            pred_text = "Prediction: model not loaded"
            if self.runner is not None:
                batch = np.stack(batch_list, axis=0)
                preds = self.runner.predict(batch)
                code = "".join(str(int(p)) for p in preds)
                pred_text = f"Prediction: {code}"
            self.pred_var.set(pred_text)

            full = self._to_photo(self.current_image_bgr)
            roi_p = self._to_photo(roi_crop)
            self.full_img_label.configure(image=full, text="")
            self.roi_img_label.configure(image=roi_p, text="")
            self.photo_refs["full"] = full
            self.photo_refs["roi"] = roi_p

            for child in self.digit_frame.winfo_children():
                child.destroy()
            self.digit_photo_refs.clear()
            for i, d in enumerate(preview_list):
                img = self._to_photo(d, max_size=(90, 90))
                self.digit_photo_refs.append(img)
                lbl = ttk.Label(self.digit_frame, image=img)
                lbl.grid(row=0, column=i, padx=2, pady=2)
                ttk.Label(self.digit_frame, text=str(i)).grid(row=1, column=i)

        except Exception as exc:
            self.log(f"Preview error: {exc}")

    def run_batch_async(self):
        t = threading.Thread(target=self.run_batch, daemon=True)
        t.start()

    def run_batch(self):
        try:
            if self.runner is None:
                raise RuntimeError("Load model first.")
            if self.roi is None:
                raise RuntimeError("Set ROI first.")

            excel_path = Path(self.excel_var.get().strip())
            if not excel_path.exists():
                raise RuntimeError(f"Excel not found: {excel_path}")

            out_path = self.output_excel_var.get().strip()
            output_excel = Path(out_path) if out_path else excel_path.with_name(f"{excel_path.stem}_with_udise.xlsx")

            wb = load_workbook(excel_path)
            ws = resolve_sheet(wb, self.sheet_var.get().strip() or None)
            url_col, out_col = self._resolve_cols(ws)

            first_row = max(2, int(self.start_row_var.get()))
            max_rows = int(self.max_rows_var.get())
            last_row = ws.max_row if max_rows <= 0 else min(ws.max_row, first_row + max_rows - 1)
            if first_row > last_row:
                raise RuntimeError("No rows to process with current row limits.")

            digit_count = int(self.digit_count_var.get())
            s = self._settings()
            overwrite = bool(self.overwrite_var.get())
            save_every = int(self.save_every_var.get())

            roi_obj = parse_manual_roi(
                f"{self.roi[0]},{self.roi[1]},{self.roi[2]},{self.roi[3]}"
            )

            session = requests.Session()
            processed = 0
            success = 0
            skipped = 0
            autosave_counter = 0
            failures: List[Tuple[int, str, str]] = []

            self.log(
                f"Batch started rows {first_row}..{last_row}, digit_count={digit_count}, image_size={s.image_size}, invert={s.invert}, th={s.threshold_mode}"
            )

            try:
                for row in range(first_row, last_row + 1):
                    existing = ws.cell(row=row, column=out_col).value
                    if (existing is not None) and str(existing).strip() and not overwrite:
                        skipped += 1
                        continue

                    source_val = ws.cell(row=row, column=url_col).value
                    if not source_val or not str(source_val).strip():
                        skipped += 1
                        continue

                    source = str(source_val).strip()
                    processed += 1
                    try:
                        image_bgr = decode_image_from_source(
                            source=source,
                            timeout=int(self.timeout_var.get()),
                            retries=int(self.retries_var.get()),
                            retry_backoff=float(self.backoff_var.get()),
                            session=session,
                        )

                        roi_crop = crop_roi(image_bgr, roi_obj)
                        digit_cells = split_digits_equal(roi_crop, digit_count)

                        batch = np.stack(
                            [preprocess_digit_custom(cell, s)[0] for cell in digit_cells],
                            axis=0,
                        )
                        preds = self.runner.predict(batch)
                        if len(preds) != digit_count:
                            raise RuntimeError(
                                f"Model returned {len(preds)} digits, expected {digit_count}."
                            )

                        code = "".join(str(int(p)) for p in preds)
                        ws.cell(row=row, column=out_col).value = code
                        success += 1
                        autosave_counter += 1
                        self.log(f"row {row}: {code}")

                        if save_every > 0 and autosave_counter >= save_every:
                            wb.save(output_excel)
                            autosave_counter = 0
                            self.log(f"[autosave] {output_excel}")
                    except Exception as exc:
                        failures.append((row, source, str(exc)))
                        self.log(f"row {row}: ERROR: {exc}")
            finally:
                session.close()

            wb.save(output_excel)
            write_failure_log("udise_failures_gui.csv", failures)
            self.log(
                f"Batch done. processed={processed}, skipped={skipped}, success={success}, failed={len(failures)}, output={output_excel}"
            )
            if failures:
                self.log("Failure log: udise_failures_gui.csv")
            messagebox.showinfo(
                "Done",
                f"Batch completed.\nSuccess: {success}\nFailed: {len(failures)}\nOutput: {output_excel}",
            )
        except Exception as exc:
            self.log(f"Batch error: {exc}")
            messagebox.showerror("Batch Error", str(exc))


def main():
    app = UDISEOCRApp()
    app.mainloop()


if __name__ == "__main__":
    main()
