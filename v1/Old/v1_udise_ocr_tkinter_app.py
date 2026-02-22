#!/usr/bin/env python3
"""Enhanced Tkinter app for interactive UDISE OCR tuning and production runs.

Key upgrades:
- Single ROI / Grid ROI modes (11 digit boxes)
- Fine-tune per-digit grid offsets
- Global shift D-pad controls
- Clickable digit previews that map back to ROI boxes
- Confidence-aware preview + optional manual review during batch
- Calibration / Production tabs to reduce clutter
"""

from __future__ import annotations

import json
import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import cv2
import numpy as np
import requests
import tkinter as tk
from PIL import Image, ImageTk
from openpyxl import load_workbook
from tkinter import filedialog, messagebox, simpledialog, ttk

from udise_ocr_pipeline import (
    KerasRunner,
    TorchRunner,
    crop_roi,
    decode_image_from_source,
    get_model_runner,
    load_model_config,
    parse_manual_roi,
    resolve_column,
    resolve_sheet,
    split_digits_equal,
    write_failure_log,
)

DEFAULT_EXCEL = "test_udise.xlsx"
DEFAULT_ROI_FILE = "udise_roi.json"
DEFAULT_DIGIT_COUNT = 11


@dataclass
class PreprocessSettings:
    image_size: int = 28
    invert: bool = True
    blur_kernel: int = 3
    threshold_mode: str = "otsu"  # otsu | adaptive
    adaptive_block_size: int = 21
    adaptive_c: int = 10
    tight_crop: bool = True
    bbox_padding: int = 2


@dataclass
class BatchReviewRequest:
    row: int
    predicted: str
    confidences: List[float]
    suspect_idxs: List[int]


def ensure_odd(value: int, minimum: int) -> int:
    v = max(minimum, int(value))
    if v % 2 == 0:
        v += 1
    return v


def preprocess_digit_custom(digit_bgr: np.ndarray, s: PreprocessSettings) -> Tuple[np.ndarray, np.ndarray]:
    gray = cv2.cvtColor(digit_bgr, cv2.COLOR_BGR2GRAY)

    blur_k = ensure_odd(s.blur_kernel, 1)
    if blur_k > 1:
        gray = cv2.GaussianBlur(gray, (blur_k, blur_k), 0)

    if s.threshold_mode == "adaptive":
        block_size = ensure_odd(s.adaptive_block_size, 3)
        th = cv2.adaptiveThreshold(
            gray,
            255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY,
            block_size,
            int(s.adaptive_c),
        )
    else:
        _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    if s.invert:
        th = cv2.bitwise_not(th)

    proc = th
    if s.tight_crop:
        fg_mask = proc if s.invert else (255 - proc)
        nz = cv2.findNonZero(fg_mask)
        if nz is not None:
            x, y, w, h = cv2.boundingRect(nz)
            pad = max(0, int(s.bbox_padding))
            x0 = max(0, x - pad)
            y0 = max(0, y - pad)
            x1 = min(proc.shape[1], x + w + pad)
            y1 = min(proc.shape[0], y + h + pad)
            proc = proc[y0:y1, x0:x1]

    h, w = proc.shape[:2]
    image_size = max(8, int(s.image_size))
    scale = image_size / float(max(h, w))
    new_w = max(1, int(round(w * scale)))
    new_h = max(1, int(round(h * scale)))
    resized = cv2.resize(proc, (new_w, new_h), interpolation=cv2.INTER_AREA)

    bg = 0 if s.invert else 255
    canvas = np.full((image_size, image_size), bg, dtype=np.uint8)
    y0 = (image_size - new_h) // 2
    x0 = (image_size - new_w) // 2
    canvas[y0 : y0 + new_h, x0 : x0 + new_w] = resized

    arr = canvas.astype(np.float32) / 255.0
    arr = np.expand_dims(arr, axis=-1)
    return arr, canvas


class UDISEOCRApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("UDISE OCR Tuning App")
        self.geometry("1540x960")

        self.runner = None
        self.current_image_bgr: Optional[np.ndarray] = None
        self.current_row: Optional[int] = None
        self.current_source: Optional[str] = None

        # ROI state
        self.single_roi: Optional[Tuple[int, int, int, int]] = None
        self.grid_base_boxes: List[Tuple[int, int, int, int]] = []
        self.grid_offsets: List[List[int]] = []  # per digit: [dx,dy,dw,dh]
        self.selected_digit_idx: Optional[int] = None

        # Preview/state
        self.last_digit_previews: List[np.ndarray] = []
        self.last_preds: List[int] = []
        self.last_confs: List[float] = []
        self.photo_refs: Dict[str, ImageTk.PhotoImage] = {}
        self.digit_photo_refs: List[ImageTk.PhotoImage] = []

        self._build_vars()
        self._build_styles()
        self._build_ui()

    def _build_vars(self) -> None:
        self.excel_var = tk.StringVar(value=DEFAULT_EXCEL)
        self.output_excel_var = tk.StringVar(value="")
        self.sheet_var = tk.StringVar(value="")
        self.url_col_var = tk.StringVar(value="Original Image Path")
        self.out_col_var = tk.StringVar(value="UDISE Found")

        self.framework_var = tk.StringVar(value="keras")
        self.model_path_var = tk.StringVar(value="")
        self.model_config_var = tk.StringVar(value="")
        self.model_ready_var = tk.StringVar(value="Model not loaded")

        self.roi_file_var = tk.StringVar(value=DEFAULT_ROI_FILE)
        self.roi_text_var = tk.StringVar(value="")
        self.roi_mode_var = tk.StringVar(value="single")  # single | grid
        self.grid_gap_var = tk.IntVar(value=0)

        self.timeout_var = tk.IntVar(value=20)
        self.retries_var = tk.IntVar(value=2)
        self.backoff_var = tk.DoubleVar(value=1.5)
        self.start_row_var = tk.IntVar(value=2)
        self.max_rows_var = tk.IntVar(value=0)
        self.overwrite_var = tk.BooleanVar(value=False)
        self.save_every_var = tk.IntVar(value=50)

        self.manual_review_var = tk.BooleanVar(value=False)
        self.conf_threshold_var = tk.DoubleVar(value=0.85)

        self.sample_row_var = tk.IntVar(value=2)

        self.image_size_var = tk.IntVar(value=28)
        self.invert_var = tk.BooleanVar(value=True)
        self.blur_var = tk.IntVar(value=3)
        self.th_mode_var = tk.StringVar(value="otsu")
        self.ad_block_var = tk.IntVar(value=21)
        self.ad_c_var = tk.IntVar(value=10)
        self.tight_crop_var = tk.BooleanVar(value=True)
        self.pad_var = tk.IntVar(value=2)
        self.digit_count_var = tk.IntVar(value=DEFAULT_DIGIT_COUNT)

        self.shift_step_var = tk.IntVar(value=1)

        self.pred_var = tk.StringVar(value="Prediction: model not loaded")
        self.status_var = tk.StringVar(value="Ready")

    def _build_styles(self) -> None:
        self.style = ttk.Style(self)
        try:
            self.style.configure("Ready.TButton", foreground="white", background="#2e7d32")
        except Exception:
            pass

    def _build_ui(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        top = ttk.Frame(self, padding=8)
        top.grid(row=0, column=0, sticky="ew")
        top.columnconfigure(1, weight=1)

        self._row_file(top, 0, "Excel", self.excel_var, self.pick_excel)
        self._row_file(top, 1, "Output Excel", self.output_excel_var, self.pick_output_excel)
        self._row_file(top, 2, "Model", self.model_path_var, self.pick_model)
        self._row_file(top, 3, "Model Config", self.model_config_var, self.pick_model_config)

        meta = ttk.Frame(top)
        meta.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(6, 0))
        for i in range(8):
            meta.columnconfigure(i, weight=1)

        ttk.Label(meta, text="Sheet").grid(row=0, column=0, sticky="w")
        ttk.Entry(meta, textvariable=self.sheet_var, width=14).grid(row=0, column=1, sticky="ew")

        ttk.Label(meta, text="URL Col/Header").grid(row=0, column=2, sticky="w")
        ttk.Entry(meta, textvariable=self.url_col_var, width=20).grid(row=0, column=3, sticky="ew")

        ttk.Label(meta, text="Output Col/Header").grid(row=0, column=4, sticky="w")
        ttk.Entry(meta, textvariable=self.out_col_var, width=20).grid(row=0, column=5, sticky="ew")

        ttk.Label(meta, text="Framework").grid(row=0, column=6, sticky="w")
        ttk.Combobox(meta, textvariable=self.framework_var, state="readonly", values=["keras", "torch"], width=10).grid(row=0, column=7, sticky="w")

        main = ttk.Panedwindow(self, orient=tk.HORIZONTAL)
        main.grid(row=1, column=0, sticky="nsew")

        left = ttk.Frame(main, padding=8)
        right = ttk.Frame(main, padding=8)
        main.add(left, weight=2)
        main.add(right, weight=3)

        self._build_left_panel(left)
        self._build_right_panel(right)

        status = ttk.Label(self, textvariable=self.status_var, anchor="w", padding=6)
        status.grid(row=2, column=0, sticky="ew")

    def _row_file(self, parent, row, label, var, cmd):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", padx=(0, 8), pady=2)
        ttk.Entry(parent, textvariable=var).grid(row=row, column=1, sticky="ew", pady=2)
        ttk.Button(parent, text="Browse", command=cmd).grid(row=row, column=2, sticky="w", padx=(8, 0), pady=2)

    def _build_left_panel(self, parent):
        parent.columnconfigure(0, weight=1)

        notebook = ttk.Notebook(parent)
        notebook.grid(row=0, column=0, sticky="nsew")

        calibration = ttk.Frame(notebook, padding=8)
        production = ttk.Frame(notebook, padding=8)
        notebook.add(calibration, text="Calibration")
        notebook.add(production, text="Production")

        self._build_calibration_tab(calibration)
        self._build_production_tab(production)

    def _build_calibration_tab(self, parent):
        parent.columnconfigure(0, weight=1)

        roi_box = ttk.LabelFrame(parent, text="ROI & Sample", padding=8)
        roi_box.grid(row=0, column=0, sticky="ew")
        for i in range(7):
            roi_box.columnconfigure(i, weight=1)

        ttk.Label(roi_box, text="ROI File").grid(row=0, column=0, sticky="w")
        ttk.Entry(roi_box, textvariable=self.roi_file_var).grid(row=0, column=1, columnspan=4, sticky="ew")
        ttk.Button(roi_box, text="Load ROI", command=self.load_roi_from_file).grid(row=0, column=5, sticky="ew")
        ttk.Button(roi_box, text="Save ROI", command=self.save_roi_to_file).grid(row=0, column=6, sticky="ew")

        ttk.Label(roi_box, text="ROI Mode").grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Radiobutton(roi_box, text="Single ROI", variable=self.roi_mode_var, value="single", command=self.refresh_preview).grid(row=1, column=1, sticky="w", pady=(6, 0))
        ttk.Radiobutton(roi_box, text="Grid ROI (11 boxes)", variable=self.roi_mode_var, value="grid", command=self.refresh_preview).grid(row=1, column=2, columnspan=2, sticky="w", pady=(6, 0))

        ttk.Label(roi_box, text="Grid Gap px").grid(row=1, column=4, sticky="e", pady=(6, 0))
        ttk.Entry(roi_box, textvariable=self.grid_gap_var, width=8).grid(row=1, column=5, sticky="w", pady=(6, 0))
        ttk.Button(roi_box, text="Fine-tune Grid", command=self.open_grid_tuner).grid(row=1, column=6, sticky="ew", pady=(6, 0))

        ttk.Label(roi_box, text="Single ROI x,y,w,h").grid(row=2, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(roi_box, textvariable=self.roi_text_var).grid(row=2, column=1, columnspan=3, sticky="ew", pady=(6, 0))
        ttk.Button(roi_box, text="Apply ROI", command=self.apply_roi_text).grid(row=2, column=4, sticky="ew", pady=(6, 0))
        ttk.Button(roi_box, text="Pick ROI", command=self.pick_roi_from_sample).grid(row=2, column=5, sticky="ew", pady=(6, 0))
        ttk.Button(roi_box, text="Pick Grid Start", command=self.pick_grid_start_box).grid(row=2, column=6, sticky="ew", pady=(6, 0))

        ttk.Label(roi_box, text="Sample Row").grid(row=3, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(roi_box, textvariable=self.sample_row_var, width=8).grid(row=3, column=1, sticky="w", pady=(6, 0))
        ttk.Button(roi_box, text="Load Sample Row", command=self.load_sample_row).grid(row=3, column=2, columnspan=2, sticky="ew", pady=(6, 0))
        ttk.Button(roi_box, text="Find Next Pending", command=self.find_next_pending).grid(row=3, column=4, columnspan=3, sticky="ew", pady=(6, 0))

        pp_box = ttk.LabelFrame(parent, text="Preprocess Controls", padding=8)
        pp_box.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        for i in range(4):
            pp_box.columnconfigure(i, weight=1)

        ttk.Label(pp_box, text="Digit Count").grid(row=0, column=0, sticky="w")
        ttk.Spinbox(pp_box, from_=1, to=20, textvariable=self.digit_count_var, width=8).grid(row=0, column=1, sticky="w")

        ttk.Label(pp_box, text="Image Size").grid(row=0, column=2, sticky="w")
        ttk.Spinbox(pp_box, from_=8, to=96, textvariable=self.image_size_var, width=8).grid(row=0, column=3, sticky="w")

        ttk.Checkbutton(pp_box, text="Invert", variable=self.invert_var).grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Checkbutton(pp_box, text="Tight Crop", variable=self.tight_crop_var).grid(row=1, column=1, sticky="w", pady=(6, 0))

        ttk.Label(pp_box, text="Blur Kernel").grid(row=1, column=2, sticky="w", pady=(6, 0))
        ttk.Spinbox(pp_box, from_=1, to=11, increment=2, textvariable=self.blur_var, width=8).grid(row=1, column=3, sticky="w", pady=(6, 0))

        ttk.Label(pp_box, text="Threshold").grid(row=2, column=0, sticky="w", pady=(6, 0))
        ttk.Combobox(pp_box, textvariable=self.th_mode_var, state="readonly", values=["otsu", "adaptive"], width=12).grid(row=2, column=1, sticky="w", pady=(6, 0))

        ttk.Label(pp_box, text="Adaptive Block").grid(row=2, column=2, sticky="w", pady=(6, 0))
        ttk.Spinbox(pp_box, from_=3, to=51, increment=2, textvariable=self.ad_block_var, width=8).grid(row=2, column=3, sticky="w", pady=(6, 0))

        ttk.Label(pp_box, text="Adaptive C").grid(row=3, column=0, sticky="w", pady=(6, 0))
        ttk.Spinbox(pp_box, from_=-20, to=30, textvariable=self.ad_c_var, width=8).grid(row=3, column=1, sticky="w", pady=(6, 0))

        ttk.Label(pp_box, text="BBox Padding").grid(row=3, column=2, sticky="w", pady=(6, 0))
        ttk.Spinbox(pp_box, from_=0, to=20, textvariable=self.pad_var, width=8).grid(row=3, column=3, sticky="w", pady=(6, 0))

        actions = ttk.Frame(parent)
        actions.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        actions.columnconfigure(0, weight=1)
        actions.columnconfigure(1, weight=1)
        actions.columnconfigure(2, weight=1)

        self.load_model_btn = ttk.Button(actions, text="Load Model", command=self.load_model)
        self.load_model_btn.grid(row=0, column=0, sticky="ew", padx=2)
        ttk.Button(actions, text="Refresh Preview", command=self.refresh_preview).grid(row=0, column=1, sticky="ew", padx=2)
        ttk.Button(actions, text="Run Batch", command=self.run_batch_async).grid(row=0, column=2, sticky="ew", padx=2)

        ttk.Label(parent, textvariable=self.model_ready_var).grid(row=3, column=0, sticky="w", pady=(8, 0))

    def _build_production_tab(self, parent):
        parent.columnconfigure(0, weight=1)

        run_cfg = ttk.LabelFrame(parent, text="Batch Settings", padding=8)
        run_cfg.grid(row=0, column=0, sticky="ew")
        for i in range(4):
            run_cfg.columnconfigure(i, weight=1)

        ttk.Label(run_cfg, text="Start Row").grid(row=0, column=0, sticky="w")
        ttk.Entry(run_cfg, textvariable=self.start_row_var, width=8).grid(row=0, column=1, sticky="w")

        ttk.Label(run_cfg, text="Max Rows (0=all)").grid(row=0, column=2, sticky="w")
        ttk.Entry(run_cfg, textvariable=self.max_rows_var, width=8).grid(row=0, column=3, sticky="w")

        ttk.Label(run_cfg, text="Timeout").grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(run_cfg, textvariable=self.timeout_var, width=8).grid(row=1, column=1, sticky="w", pady=(6, 0))

        ttk.Label(run_cfg, text="Retries").grid(row=1, column=2, sticky="w", pady=(6, 0))
        ttk.Entry(run_cfg, textvariable=self.retries_var, width=8).grid(row=1, column=3, sticky="w", pady=(6, 0))

        ttk.Label(run_cfg, text="Backoff").grid(row=2, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(run_cfg, textvariable=self.backoff_var, width=8).grid(row=2, column=1, sticky="w", pady=(6, 0))

        ttk.Label(run_cfg, text="Save Every").grid(row=2, column=2, sticky="w", pady=(6, 0))
        ttk.Entry(run_cfg, textvariable=self.save_every_var, width=8).grid(row=2, column=3, sticky="w", pady=(6, 0))

        ttk.Checkbutton(run_cfg, text="Overwrite Existing", variable=self.overwrite_var).grid(row=3, column=0, columnspan=2, sticky="w", pady=(6, 0))

        ttk.Checkbutton(run_cfg, text="Manual Review", variable=self.manual_review_var).grid(row=3, column=2, sticky="w", pady=(6, 0))
        ttk.Label(run_cfg, text="Threshold").grid(row=3, column=3, sticky="e", pady=(6, 0))
        ttk.Entry(run_cfg, textvariable=self.conf_threshold_var, width=6).grid(row=3, column=3, sticky="w", padx=(60, 0), pady=(6, 0))

        ttk.Label(
            parent,
            text="If Manual Review is enabled, batch pauses when any digit confidence is below threshold.",
            wraplength=430,
        ).grid(row=1, column=0, sticky="w", pady=(8, 0))

    def _build_right_panel(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(2, weight=1)

        preview = ttk.LabelFrame(parent, text="Image Preview", padding=8)
        preview.grid(row=0, column=0, sticky="nsew")
        preview.columnconfigure(0, weight=1)
        preview.columnconfigure(1, weight=1)

        self.full_img_label = ttk.Label(preview, text="Full image")
        self.full_img_label.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        self.roi_img_label = ttk.Label(preview, text="ROI preview")
        self.roi_img_label.grid(row=0, column=1, sticky="nsew")

        digits = ttk.LabelFrame(parent, text="Preprocessed Digits", padding=8)
        digits.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        digits.columnconfigure(0, weight=1)

        # Global shift controls (D-pad)
        dpad = ttk.Frame(digits)
        dpad.grid(row=0, column=0, sticky="w", pady=(0, 8))
        ttk.Label(dpad, text="Global Shift").grid(row=0, column=0, columnspan=3, sticky="w")
        ttk.Button(dpad, text="↑", width=3, command=lambda: self.shift_all(0, -self._shift_step())).grid(row=1, column=1, padx=2, pady=1)
        ttk.Button(dpad, text="←", width=3, command=lambda: self.shift_all(-self._shift_step(), 0)).grid(row=2, column=0, padx=2, pady=1)
        ttk.Button(dpad, text="→", width=3, command=lambda: self.shift_all(self._shift_step(), 0)).grid(row=2, column=2, padx=2, pady=1)
        ttk.Button(dpad, text="↓", width=3, command=lambda: self.shift_all(0, self._shift_step())).grid(row=3, column=1, padx=2, pady=1)
        ttk.Label(dpad, text="Step").grid(row=1, column=3, padx=(10, 2))
        ttk.Entry(dpad, textvariable=self.shift_step_var, width=5).grid(row=1, column=4)

        self.digit_frame = ttk.Frame(digits)
        self.digit_frame.grid(row=1, column=0, sticky="ew")

        ttk.Label(digits, textvariable=self.pred_var, font=("TkDefaultFont", 11, "bold")).grid(row=2, column=0, sticky="w", pady=(6, 0))

        logs = ttk.LabelFrame(parent, text="Logs", padding=8)
        logs.grid(row=2, column=0, sticky="nsew", pady=(8, 0))
        logs.rowconfigure(0, weight=1)
        logs.columnconfigure(0, weight=1)

        self.log_text = tk.Text(logs, wrap="word", height=14)
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scroll = ttk.Scrollbar(logs, orient="vertical", command=self.log_text.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.log_text.configure(yscrollcommand=scroll.set)

    # ---------- Utility ----------
    def log(self, msg: str) -> None:
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.status_var.set(msg)
        self.update_idletasks()

    def _shift_step(self) -> int:
        try:
            return max(1, int(self.shift_step_var.get()))
        except Exception:
            return 1

    def _settings(self) -> PreprocessSettings:
        return PreprocessSettings(
            image_size=max(8, int(self.image_size_var.get())),
            invert=bool(self.invert_var.get()),
            blur_kernel=max(1, int(self.blur_var.get())),
            threshold_mode=self.th_mode_var.get().strip() or "otsu",
            adaptive_block_size=max(3, int(self.ad_block_var.get())),
            adaptive_c=int(self.ad_c_var.get()),
            tight_crop=bool(self.tight_crop_var.get()),
            bbox_padding=max(0, int(self.pad_var.get())),
        )

    def _to_photo(self, bgr_or_gray: np.ndarray, max_size=(620, 420)):
        if bgr_or_gray.ndim == 2:
            rgb = cv2.cvtColor(bgr_or_gray, cv2.COLOR_GRAY2RGB)
        else:
            rgb = cv2.cvtColor(bgr_or_gray, cv2.COLOR_BGR2RGB)
        h, w = rgb.shape[:2]
        scale = min(max_size[0] / w, max_size[1] / h, 1.0)
        if scale < 1.0:
            rgb = cv2.resize(rgb, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_AREA)
        return ImageTk.PhotoImage(Image.fromarray(rgb))

    def _open_sheet(self):
        excel_path = Path(self.excel_var.get().strip())
        if not excel_path.exists():
            raise RuntimeError(f"Excel not found: {excel_path}")
        wb = load_workbook(excel_path)
        ws = resolve_sheet(wb, self.sheet_var.get().strip() or None)
        return wb, ws

    def _resolve_cols(self, ws):
        url_ref = self.url_col_var.get().strip() or None
        out_ref = self.out_col_var.get().strip() or None
        url_col = resolve_column(ws, url_ref, ["original image path", "image path on server", "image url", "url"], "H")
        out_col = resolve_column(ws, out_ref, ["udise found", "udise", "udise code"], "E", allow_create=True)
        return url_col, out_col

    def _effective_grid_boxes(self) -> List[Tuple[int, int, int, int]]:
        boxes: List[Tuple[int, int, int, int]] = []
        if not self.grid_base_boxes:
            return boxes
        for i, base in enumerate(self.grid_base_boxes):
            dx, dy, dw, dh = self.grid_offsets[i] if i < len(self.grid_offsets) else [0, 0, 0, 0]
            x, y, w, h = base
            w2 = max(1, w + int(dw))
            h2 = max(1, h + int(dh))
            boxes.append((int(x + dx), int(y + dy), int(w2), int(h2)))
        return boxes

    def _digit_boxes_for_image(self, image_bgr: np.ndarray) -> List[Tuple[int, int, int, int]]:
        if self.roi_mode_var.get() == "grid":
            boxes = self._effective_grid_boxes()
            valid: List[Tuple[int, int, int, int]] = []
            ih, iw = image_bgr.shape[:2]
            for x, y, w, h in boxes:
                x0 = max(0, x)
                y0 = max(0, y)
                x1 = min(iw, x + w)
                y1 = min(ih, y + h)
                if x1 > x0 and y1 > y0:
                    valid.append((x0, y0, x1 - x0, y1 - y0))
            return valid

        if self.single_roi is None:
            return []
        x, y, w, h = self.single_roi
        roi_obj = parse_manual_roi(f"{x},{y},{w},{h}")
        roi_crop = crop_roi(image_bgr, roi_obj)
        cells = split_digits_equal(roi_crop, int(self.digit_count_var.get()))

        # Recreate cell absolute boxes so UI can highlight per digit in full image.
        cell_boxes: List[Tuple[int, int, int, int]] = []
        roi_x, roi_y, roi_w, roi_h = self.single_roi
        cell_w = roi_w / float(int(self.digit_count_var.get()))
        for i in range(int(self.digit_count_var.get())):
            c0 = int(round(i * cell_w))
            c1 = int(round((i + 1) * cell_w))
            c1 = max(c1, c0 + 1)
            cell_boxes.append((roi_x + c0, roi_y, c1 - c0, roi_h))
        return cell_boxes

    def _crop_from_boxes(self, image_bgr: np.ndarray, boxes: Sequence[Tuple[int, int, int, int]]) -> List[np.ndarray]:
        crops: List[np.ndarray] = []
        ih, iw = image_bgr.shape[:2]
        for x, y, w, h in boxes:
            x0 = max(0, x)
            y0 = max(0, y)
            x1 = min(iw, x + w)
            y1 = min(ih, y + h)
            if x1 <= x0 or y1 <= y0:
                continue
            crops.append(image_bgr[y0:y1, x0:x1])
        return crops

    def _predict_with_confidence(self, batch: np.ndarray) -> Tuple[List[int], List[float]]:
        if self.runner is None:
            return [], []

        if isinstance(self.runner, KerasRunner):
            x = self.runner._prepare_input(batch)  # Reuse runner's input adaptation.
            logits = np.asarray(self.runner.model.predict(x, verbose=0))
            preds = logits.argmax(axis=1).astype(int).tolist()
            confs = logits.max(axis=1).astype(float).tolist()
            return preds, confs

        if isinstance(self.runner, TorchRunner):
            x = np.transpose(batch, (0, 3, 1, 2)).astype(np.float32)
            tensor = self.runner.torch.from_numpy(x)
            with self.runner.torch.no_grad():
                out = self.runner.model(tensor)
            logits = out.detach().cpu().numpy()
            exp = np.exp(logits - logits.max(axis=1, keepdims=True))
            probs = exp / exp.sum(axis=1, keepdims=True)
            preds = probs.argmax(axis=1).astype(int).tolist()
            confs = probs.max(axis=1).astype(float).tolist()
            return preds, confs

        preds = self.runner.predict(batch)
        confs = [1.0 for _ in preds]
        return preds, confs

    # ---------- File pickers ----------
    def pick_excel(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if p:
            self.excel_var.set(p)
            if not self.output_excel_var.get().strip():
                out = Path(p).with_name(f"{Path(p).stem}_with_udise.xlsx")
                self.output_excel_var.set(str(out))

    def pick_output_excel(self):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if p:
            self.output_excel_var.set(p)

    def pick_model(self):
        p = filedialog.askopenfilename(filetypes=[("Model", "*.keras *.h5 *.pt *.pth"), ("All", "*.*")])
        if p:
            self.model_path_var.set(p)
            json_guess = Path(str(p) + ".json")
            if json_guess.exists() and not self.model_config_var.get().strip():
                self.model_config_var.set(str(json_guess))

    def pick_model_config(self):
        p = filedialog.askopenfilename(filetypes=[("JSON", "*.json")])
        if p:
            self.model_config_var.set(p)

    # ---------- ROI / Grid ----------
    def load_roi_from_file(self):
        p = Path(self.roi_file_var.get().strip())
        if not p.exists():
            messagebox.showerror("Error", f"ROI file not found: {p}")
            return
        data = json.loads(p.read_text(encoding="utf-8"))

        mode = str(data.get("mode", "single"))
        self.roi_mode_var.set(mode)

        if mode == "grid":
            self.grid_base_boxes = [tuple(map(int, b)) for b in data.get("grid_base_boxes", [])]
            offsets = data.get("grid_offsets", [])
            self.grid_offsets = [list(map(int, o)) for o in offsets]
            if len(self.grid_offsets) < len(self.grid_base_boxes):
                for _ in range(len(self.grid_base_boxes) - len(self.grid_offsets)):
                    self.grid_offsets.append([0, 0, 0, 0])
            self.log(f"Grid ROI loaded from {p}")
        else:
            x = int(data["x"])
            y = int(data["y"])
            w = int(data["w"])
            h = int(data["h"])
            self.single_roi = (x, y, w, h)
            self.roi_text_var.set(f"{x},{y},{w},{h}")
            self.log(f"Single ROI loaded from {p}")

        self.refresh_preview()

    def save_roi_to_file(self):
        p = Path(self.roi_file_var.get().strip())
        p.parent.mkdir(parents=True, exist_ok=True)
        if self.roi_mode_var.get() == "grid":
            payload = {
                "mode": "grid",
                "grid_base_boxes": self.grid_base_boxes,
                "grid_offsets": self.grid_offsets,
            }
        else:
            if self.single_roi is None:
                messagebox.showerror("Error", "No single ROI to save.")
                return
            x, y, w, h = self.single_roi
            payload = {"mode": "single", "x": x, "y": y, "w": w, "h": h}

        p.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        self.log(f"ROI saved to {p}")

    def apply_roi_text(self):
        txt = self.roi_text_var.get().strip()
        try:
            r = parse_manual_roi(txt)
        except Exception as exc:
            messagebox.showerror("ROI Error", str(exc))
            return
        self.single_roi = (r.x, r.y, r.w, r.h)
        self.roi_mode_var.set("single")
        self.log(f"Single ROI set: {self.single_roi}")
        self.refresh_preview()

    def pick_roi_from_sample(self):
        if self.current_image_bgr is None:
            messagebox.showerror("No image", "Load a sample row first.")
            return
        x, y, w, h = cv2.selectROI(
            "Select ROI (Enter/Space confirm, c cancel)",
            self.current_image_bgr.copy(),
            showCrosshair=True,
            fromCenter=False,
        )
        cv2.destroyAllWindows()
        if w <= 0 or h <= 0:
            return
        self.single_roi = (int(x), int(y), int(w), int(h))
        self.roi_text_var.set(f"{x},{y},{w},{h}")
        self.roi_mode_var.set("single")
        self.log(f"Single ROI selected: {self.single_roi}")
        self.refresh_preview()

    def pick_grid_start_box(self):
        if self.current_image_bgr is None:
            messagebox.showerror("No image", "Load a sample row first.")
            return
        digit_count = int(self.digit_count_var.get())
        gap = int(self.grid_gap_var.get())

        x, y, w, h = cv2.selectROI(
            "Pick FIRST digit box (grid will auto-generate)",
            self.current_image_bgr.copy(),
            showCrosshair=True,
            fromCenter=False,
        )
        cv2.destroyAllWindows()
        if w <= 0 or h <= 0:
            return

        self.grid_base_boxes = []
        for i in range(digit_count):
            xi = int(x + i * (w + gap))
            self.grid_base_boxes.append((xi, int(y), int(w), int(h)))
        self.grid_offsets = [[0, 0, 0, 0] for _ in range(digit_count)]
        self.roi_mode_var.set("grid")
        self.log(f"Grid generated with {digit_count} boxes from first box.")
        self.refresh_preview()

    def open_grid_tuner(self):
        if self.roi_mode_var.get() != "grid":
            messagebox.showinfo("Grid", "Switch ROI mode to Grid first.")
            return
        if not self.grid_base_boxes:
            messagebox.showerror("Grid", "No grid available. Use 'Pick Grid Start' first.")
            return

        top = tk.Toplevel(self)
        top.title("Fine-tune Grid Offsets")
        top.geometry("760x520")

        container = ttk.Frame(top, padding=8)
        container.pack(fill="both", expand=True)

        header = ttk.Frame(container)
        header.pack(fill="x")
        ttk.Label(header, text="Digit", width=8).grid(row=0, column=0)
        ttk.Label(header, text="dx", width=10).grid(row=0, column=1)
        ttk.Label(header, text="dy", width=10).grid(row=0, column=2)
        ttk.Label(header, text="dw", width=10).grid(row=0, column=3)
        ttk.Label(header, text="dh", width=10).grid(row=0, column=4)

        canvas = tk.Canvas(container)
        scroll = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        body = ttk.Frame(canvas)
        body.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        vars_list: List[Tuple[tk.IntVar, tk.IntVar, tk.IntVar, tk.IntVar]] = []
        for i in range(len(self.grid_base_boxes)):
            if i >= len(self.grid_offsets):
                self.grid_offsets.append([0, 0, 0, 0])
            dx, dy, dw, dh = self.grid_offsets[i]
            vdx, vdy, vdw, vdh = tk.IntVar(value=dx), tk.IntVar(value=dy), tk.IntVar(value=dw), tk.IntVar(value=dh)
            vars_list.append((vdx, vdy, vdw, vdh))

            ttk.Label(body, text=str(i), width=8).grid(row=i, column=0, padx=2, pady=2)
            ttk.Spinbox(body, from_=-200, to=200, textvariable=vdx, width=8).grid(row=i, column=1, padx=2, pady=2)
            ttk.Spinbox(body, from_=-200, to=200, textvariable=vdy, width=8).grid(row=i, column=2, padx=2, pady=2)
            ttk.Spinbox(body, from_=-200, to=200, textvariable=vdw, width=8).grid(row=i, column=3, padx=2, pady=2)
            ttk.Spinbox(body, from_=-200, to=200, textvariable=vdh, width=8).grid(row=i, column=4, padx=2, pady=2)

        btns = ttk.Frame(top, padding=8)
        btns.pack(fill="x")

        def apply_offsets():
            for i, (vdx, vdy, vdw, vdh) in enumerate(vars_list):
                self.grid_offsets[i] = [int(vdx.get()), int(vdy.get()), int(vdw.get()), int(vdh.get())]
            self.log("Grid offsets updated.")
            self.refresh_preview()

        ttk.Button(btns, text="Apply", command=apply_offsets).pack(side="left", padx=4)
        ttk.Button(btns, text="Apply & Close", command=lambda: (apply_offsets(), top.destroy())).pack(side="left", padx=4)
        ttk.Button(btns, text="Close", command=top.destroy).pack(side="right", padx=4)

    def shift_all(self, dx: int, dy: int):
        if self.roi_mode_var.get() == "grid":
            if not self.grid_base_boxes:
                return
            moved = []
            for x, y, w, h in self.grid_base_boxes:
                moved.append((x + dx, y + dy, w, h))
            self.grid_base_boxes = moved
            self.log(f"Shifted grid by ({dx}, {dy})")
        else:
            if self.single_roi is None:
                return
            x, y, w, h = self.single_roi
            self.single_roi = (x + dx, y + dy, w, h)
            self.roi_text_var.set(f"{self.single_roi[0]},{self.single_roi[1]},{w},{h}")
            self.log(f"Shifted ROI by ({dx}, {dy})")
        self.refresh_preview()

    # ---------- Data load ----------
    def find_next_pending(self):
        try:
            wb, ws = self._open_sheet()
            url_col, out_col = self._resolve_cols(ws)
            start = max(2, int(self.sample_row_var.get()))
            for row in range(start, ws.max_row + 1):
                url = ws.cell(row=row, column=url_col).value
                out = ws.cell(row=row, column=out_col).value
                if url and (out is None or not str(out).strip()):
                    self.sample_row_var.set(row)
                    self.log(f"Next pending row: {row}")
                    break
            wb.close()
        except Exception as exc:
            messagebox.showerror("Error", str(exc))

    def load_model(self):
        try:
            model_path = self.model_path_var.get().strip()
            if not model_path:
                raise RuntimeError("Select a model path.")
            framework = self.framework_var.get().strip()
            config_path = self.model_config_var.get().strip() or None
            cfg = load_model_config(config_path, model_path)

            if cfg.get("framework") and str(cfg["framework"]) != framework:
                raise RuntimeError(f"Framework mismatch. UI={framework}, config={cfg['framework']}")

            if "image_size" in cfg:
                self.image_size_var.set(int(cfg["image_size"]))
            if "invert" in cfg:
                self.invert_var.set(bool(cfg["invert"]))
            if "digit_count" in cfg:
                self.digit_count_var.set(int(cfg["digit_count"]))

            self.runner = get_model_runner(framework, model_path)
            self.model_ready_var.set("Model Ready")
            self.pred_var.set("Model Ready")
            self.load_model_btn.configure(style="Ready.TButton")
            self.log(f"Model loaded: {model_path}")
            self.refresh_preview()
        except Exception as exc:
            messagebox.showerror("Model Error", str(exc))

    def load_sample_row(self):
        try:
            wb, ws = self._open_sheet()
            url_col, _ = self._resolve_cols(ws)
            row = int(self.sample_row_var.get())
            source = ws.cell(row=row, column=url_col).value
            wb.close()
            if not source or not str(source).strip():
                raise RuntimeError(f"Row {row} has empty source in URL column.")

            self.current_row = row
            self.current_source = str(source).strip()
            session = requests.Session()
            try:
                self.current_image_bgr = decode_image_from_source(
                    source=self.current_source,
                    timeout=int(self.timeout_var.get()),
                    retries=int(self.retries_var.get()),
                    retry_backoff=float(self.backoff_var.get()),
                    session=session,
                )
            finally:
                session.close()

            self.log(f"Loaded sample row {row}")
            self.refresh_preview()
        except Exception as exc:
            messagebox.showerror("Sample Error", str(exc))

    # ---------- Preview ----------
    def _draw_boxes_overlay(
        self,
        image_bgr: np.ndarray,
        boxes: Sequence[Tuple[int, int, int, int]],
        low_conf_idxs: Optional[Sequence[int]] = None,
        selected_idx: Optional[int] = None,
    ) -> np.ndarray:
        out = image_bgr.copy()
        low_set = set(low_conf_idxs or [])
        for i, (x, y, w, h) in enumerate(boxes):
            color = (0, 255, 0)
            if i in low_set:
                color = (0, 0, 255)
            if selected_idx is not None and i == selected_idx:
                color = (255, 255, 0)
            cv2.rectangle(out, (x, y), (x + w, y + h), color, 2)
            cv2.putText(out, str(i), (x + 1, max(10, y - 4)), cv2.FONT_HERSHEY_SIMPLEX, 0.45, color, 1, cv2.LINE_AA)
        return out

    def on_digit_click(self, idx: int):
        self.selected_digit_idx = idx
        self.refresh_preview()

    def refresh_preview(self):
        if self.current_image_bgr is None:
            return

        try:
            boxes = self._digit_boxes_for_image(self.current_image_bgr)
            if not boxes:
                full = self._to_photo(self.current_image_bgr)
                self.full_img_label.configure(image=full, text="")
                self.photo_refs["full"] = full
                self.roi_img_label.configure(text="Pick/apply ROI or grid first", image="")
                self.pred_var.set("Prediction: model not loaded" if self.runner is None else "Model Ready")
                return

            digit_cells = self._crop_from_boxes(self.current_image_bgr, boxes)
            s = self._settings()

            batch_list: List[np.ndarray] = []
            preview_list: List[np.ndarray] = []
            for cell in digit_cells:
                arr, preview = preprocess_digit_custom(cell, s)
                batch_list.append(arr)
                preview_list.append(preview)

            preds: List[int] = []
            confs: List[float] = []
            pred_text = "Prediction: model not loaded"
            if self.runner is not None and batch_list:
                batch = np.stack(batch_list, axis=0)
                preds, confs = self._predict_with_confidence(batch)
                if preds:
                    code = "".join(str(int(p)) for p in preds)
                    pred_text = f"Model Ready | Prediction: {code}"
                else:
                    pred_text = "Model Ready"
            else:
                if self.runner is not None:
                    pred_text = "Model Ready"

            self.last_digit_previews = preview_list
            self.last_preds = preds
            self.last_confs = confs
            self.pred_var.set(pred_text)

            threshold = float(self.conf_threshold_var.get())
            low_idxs = [i for i, c in enumerate(confs) if c < threshold]

            overlay = self._draw_boxes_overlay(
                self.current_image_bgr,
                boxes,
                low_conf_idxs=low_idxs,
                selected_idx=self.selected_digit_idx,
            )

            roi_preview = self.current_image_bgr
            if self.roi_mode_var.get() == "single" and self.single_roi is not None:
                x, y, w, h = self.single_roi
                roi_obj = parse_manual_roi(f"{x},{y},{w},{h}")
                roi_preview = crop_roi(self.current_image_bgr, roi_obj)
            elif boxes:
                x0 = min(x for x, _, _, _ in boxes)
                y0 = min(y for _, y, _, _ in boxes)
                x1 = max(x + w for x, _, w, _ in boxes)
                y1 = max(y + h for _, y, _, h in boxes)
                roi_preview = self.current_image_bgr[y0:y1, x0:x1]

            full = self._to_photo(overlay)
            roi_p = self._to_photo(roi_preview)
            self.full_img_label.configure(image=full, text="")
            self.roi_img_label.configure(image=roi_p, text="")
            self.photo_refs["full"] = full
            self.photo_refs["roi"] = roi_p

            for child in self.digit_frame.winfo_children():
                child.destroy()
            self.digit_photo_refs.clear()

            for i, d in enumerate(preview_list):
                img = self._to_photo(d, max_size=(90, 90))
                self.digit_photo_refs.append(img)

                border_color = "#b71c1c" if i < len(confs) and confs[i] < threshold else "#444444"
                frame = tk.Frame(self.digit_frame, highlightthickness=2, highlightbackground=border_color)
                frame.grid(row=0, column=i, padx=2, pady=2)
                lbl = tk.Label(frame, image=img, cursor="hand2")
                lbl.pack()
                lbl.bind("<Button-1>", lambda _e, idx=i: self.on_digit_click(idx))

                caption = str(i)
                if i < len(preds):
                    caption += f" | {preds[i]}"
                if i < len(confs):
                    caption += f" ({confs[i]:.2f})"
                ttk.Label(self.digit_frame, text=caption).grid(row=1, column=i)

        except Exception as exc:
            self.log(f"Preview error: {exc}")

    # ---------- Batch ----------
    def run_batch_async(self):
        t = threading.Thread(target=self.run_batch, daemon=True)
        t.start()

    def _request_manual_review_dialog(
        self,
        req: BatchReviewRequest,
        previews: Sequence[np.ndarray],
        done_event: threading.Event,
        result_holder: Dict[str, str],
    ):
        top = tk.Toplevel(self)
        top.title(f"Manual Review - Row {req.row}")
        top.geometry("1200x360")
        top.transient(self)
        top.grab_set()

        wrapper = ttk.Frame(top, padding=8)
        wrapper.pack(fill="both", expand=True)

        suspect_txt = ", ".join(f"{i}:{req.confidences[i]:.2f}" for i in req.suspect_idxs)
        ttk.Label(
            wrapper,
            text=f"Low-confidence digits detected at [{suspect_txt}]. Edit code and continue.",
        ).pack(anchor="w")

        row = ttk.Frame(wrapper)
        row.pack(fill="x", pady=(8, 6))

        photos: List[ImageTk.PhotoImage] = []
        thr = float(self.conf_threshold_var.get())
        for i, d in enumerate(previews):
            img = self._to_photo(d, max_size=(70, 70))
            photos.append(img)
            border = "#b71c1c" if i in req.suspect_idxs else "#444444"
            f = tk.Frame(row, highlightthickness=2, highlightbackground=border)
            f.pack(side="left", padx=2)
            tk.Label(f, image=img).pack()
            txt = f"{i}:{req.predicted[i]}\n{req.confidences[i]:.2f}" if i < len(req.confidences) else f"{i}:{req.predicted[i]}"
            tk.Label(f, text=txt).pack()

        entry_var = tk.StringVar(value=req.predicted)
        form = ttk.Frame(wrapper)
        form.pack(fill="x", pady=(8, 6))
        ttk.Label(form, text="Correct code:").pack(side="left")
        entry = ttk.Entry(form, textvariable=entry_var, width=max(20, len(req.predicted) + 4))
        entry.pack(side="left", padx=(6, 0))
        entry.focus_set()

        btns = ttk.Frame(wrapper)
        btns.pack(fill="x", pady=(8, 0))

        def finish(ok: bool):
            if ok:
                code = entry_var.get().strip()
                if len(code) != len(req.predicted) or not code.isdigit():
                    messagebox.showerror("Invalid", f"Enter exactly {len(req.predicted)} digits.")
                    return
                result_holder["code"] = code
            else:
                result_holder["code"] = req.predicted
            top.destroy()
            done_event.set()

        ttk.Button(btns, text="Use Edited", command=lambda: finish(True)).pack(side="left", padx=4)
        ttk.Button(btns, text="Use Predicted", command=lambda: finish(False)).pack(side="left", padx=4)

        def on_close():
            result_holder["code"] = req.predicted
            top.destroy()
            done_event.set()

        top.protocol("WM_DELETE_WINDOW", on_close)

    def _manual_review_if_needed(
        self,
        row: int,
        code: str,
        confs: List[float],
        previews: Sequence[np.ndarray],
    ) -> str:
        if not self.manual_review_var.get():
            return code

        thr = float(self.conf_threshold_var.get())
        suspect = [i for i, c in enumerate(confs) if c < thr]
        if not suspect:
            return code

        req = BatchReviewRequest(row=row, predicted=code, confidences=confs, suspect_idxs=suspect)
        done = threading.Event()
        result: Dict[str, str] = {}
        self.after(0, lambda: self._request_manual_review_dialog(req, previews, done, result))
        done.wait()
        return result.get("code", code)

    def run_batch(self):
        try:
            if self.runner is None:
                raise RuntimeError("Load model first.")

            excel_path = Path(self.excel_var.get().strip())
            if not excel_path.exists():
                raise RuntimeError(f"Excel not found: {excel_path}")

            out_path = self.output_excel_var.get().strip()
            output_excel = Path(out_path) if out_path else excel_path.with_name(f"{excel_path.stem}_with_udise.xlsx")

            wb = load_workbook(excel_path)
            ws = resolve_sheet(wb, self.sheet_var.get().strip() or None)
            url_col, out_col = self._resolve_cols(ws)

            first_row = max(2, int(self.start_row_var.get()))
            max_rows = int(self.max_rows_var.get())
            last_row = ws.max_row if max_rows <= 0 else min(ws.max_row, first_row + max_rows - 1)
            if first_row > last_row:
                raise RuntimeError("No rows to process with current row limits.")

            digit_count = int(self.digit_count_var.get())
            s = self._settings()
            overwrite = bool(self.overwrite_var.get())
            save_every = int(self.save_every_var.get())

            session = requests.Session()
            processed = 0
            success = 0
            skipped = 0
            autosave_counter = 0
            failures: List[Tuple[int, str, str]] = []

            self.log(
                f"Batch started rows {first_row}..{last_row}, mode={self.roi_mode_var.get()}, digit_count={digit_count}, manual_review={self.manual_review_var.get()}"
            )

            try:
                for row in range(first_row, last_row + 1):
                    existing = ws.cell(row=row, column=out_col).value
                    if (existing is not None) and str(existing).strip() and not overwrite:
                        skipped += 1
                        continue

                    source_val = ws.cell(row=row, column=url_col).value
                    if not source_val or not str(source_val).strip():
                        skipped += 1
                        continue

                    source = str(source_val).strip()
                    processed += 1
                    try:
                        image_bgr = decode_image_from_source(
                            source=source,
                            timeout=int(self.timeout_var.get()),
                            retries=int(self.retries_var.get()),
                            retry_backoff=float(self.backoff_var.get()),
                            session=session,
                        )

                        boxes = self._digit_boxes_for_image(image_bgr)
                        if len(boxes) != digit_count:
                            raise RuntimeError(f"Digit boxes count={len(boxes)} expected={digit_count}.")

                        digit_cells = self._crop_from_boxes(image_bgr, boxes)
                        batch_list = [preprocess_digit_custom(cell, s)[0] for cell in digit_cells]
                        preview_list = [preprocess_digit_custom(cell, s)[1] for cell in digit_cells]
                        batch = np.stack(batch_list, axis=0)

                        preds, confs = self._predict_with_confidence(batch)
                        if len(preds) != digit_count:
                            raise RuntimeError(f"Model returned {len(preds)} digits, expected {digit_count}.")

                        code = "".join(str(int(p)) for p in preds)
                        code = self._manual_review_if_needed(row, code, confs, preview_list)

                        ws.cell(row=row, column=out_col).value = code
                        success += 1
                        autosave_counter += 1
                        self.log(f"row {row}: {code}")

                        if save_every > 0 and autosave_counter >= save_every:
                            wb.save(output_excel)
                            autosave_counter = 0
                            self.log(f"[autosave] {output_excel}")
                    except Exception as exc:
                        failures.append((row, source, str(exc)))
                        self.log(f"row {row}: ERROR: {exc}")
            finally:
                session.close()

            wb.save(output_excel)
            write_failure_log("udise_failures_gui.csv", failures)
            self.log(
                f"Batch done. processed={processed}, skipped={skipped}, success={success}, failed={len(failures)}, output={output_excel}"
            )
            if failures:
                self.log("Failure log: udise_failures_gui.csv")
            messagebox.showinfo(
                "Done",
                f"Batch completed.\nSuccess: {success}\nFailed: {len(failures)}\nOutput: {output_excel}",
            )
        except Exception as exc:
            self.log(f"Batch error: {exc}")
            messagebox.showerror("Batch Error", str(exc))


def main():
    app = UDISEOCRApp()
    app.mainloop()


if __name__ == "__main__":
    main()
