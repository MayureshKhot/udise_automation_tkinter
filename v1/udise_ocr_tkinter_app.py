#!/usr/bin/env python3
"""Enhanced Tkinter app for interactive UDISE OCR tuning and production runs.

Key upgrades:
- Single ROI / Grid ROI modes (11 digit boxes)
- Fine-tune per-digit grid offsets
- Global shift D-pad controls
- Clickable digit previews that map back to ROI boxes
- Confidence-aware preview + optional manual review during batch
- Calibration / Production tabs to reduce clutter
"""

from __future__ import annotations

import json
import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import cv2
import numpy as np
import requests
import tkinter as tk
from PIL import Image, ImageTk
from openpyxl import load_workbook
from tkinter import filedialog, messagebox, ttk
from preprocessing.morphology_grid_cleaner import remove_grid_lines

from udise_ocr_pipeline import (
    DEFAULT_GROQ_ENDPOINT,
    KerasRunner,
    TorchRunner,
    crop_roi,
    decode_image_from_source,
    get_model_runner,
    load_model_config,
    parse_manual_roi,
    resolve_column,
    resolve_sheet,
    split_digits_equal,
    write_failure_log,
)

DEFAULT_EXCEL = "test_udise.xlsx"
DEFAULT_ROI_FILE = "udise_roi.json"
DEFAULT_DIGIT_COUNT = 11


@dataclass
class PreprocessSettings:
    image_size: int = 28
    invert: bool = True
    blur_kernel: int = 3
    threshold_mode: str = "otsu"  # otsu | adaptive
    adaptive_block_size: int = 21
    adaptive_c: int = 10
    tight_crop: bool = True
    bbox_padding: int = 2
    remove_vertical_lines: bool = False
    line_min_height_ratio: float = 0.75
    line_max_width: int = 3
    line_edge_margin: int = 4
    use_morphology_cleaning: bool = False


@dataclass
class BatchReviewRequest:
    row: int
    predicted: str
    confidences: List[float]
    suspect_idxs: List[int]


def ensure_odd(value: int, minimum: int) -> int:
    v = max(minimum, int(value))
    if v % 2 == 0:
        v += 1
    return v


def preprocess_digit_custom(digit_bgr: np.ndarray, s: PreprocessSettings) -> Tuple[np.ndarray, np.ndarray]:
    gray = cv2.cvtColor(digit_bgr, cv2.COLOR_BGR2GRAY)

    blur_k = ensure_odd(s.blur_kernel, 1)
    if blur_k > 1:
        gray = cv2.GaussianBlur(gray, (blur_k, blur_k), 0)

    if s.threshold_mode == "adaptive":
        block_size = ensure_odd(s.adaptive_block_size, 3)
        th = cv2.adaptiveThreshold(
            gray,
            255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY,
            block_size,
            int(s.adaptive_c),
        )
    else:
        _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    if s.invert:
        th = cv2.bitwise_not(th)

    proc = th
    if s.use_morphology_cleaning:
        clean_src = proc if s.invert else cv2.bitwise_not(proc)
        clean_out = remove_grid_lines(clean_src)
        proc = clean_out if s.invert else cv2.bitwise_not(clean_out)

    if s.remove_vertical_lines:
        proc = remove_vertical_grid_lines(proc, s.invert, s.line_min_height_ratio, s.line_max_width, s.line_edge_margin)

    if s.tight_crop:
        fg_mask = proc if s.invert else (255 - proc)
        nz = cv2.findNonZero(fg_mask)
        if nz is not None:
            x, y, w, h = cv2.boundingRect(nz)
            pad = max(0, int(s.bbox_padding))
            x0 = max(0, x - pad)
            y0 = max(0, y - pad)
            x1 = min(proc.shape[1], x + w + pad)
            y1 = min(proc.shape[0], y + h + pad)
            proc = proc[y0:y1, x0:x1]

    h, w = proc.shape[:2]
    image_size = max(8, int(s.image_size))
    scale = image_size / float(max(h, w))
    new_w = max(1, int(round(w * scale)))
    new_h = max(1, int(round(h * scale)))
    resized = cv2.resize(proc, (new_w, new_h), interpolation=cv2.INTER_AREA)

    bg = 0 if s.invert else 255
    canvas = np.full((image_size, image_size), bg, dtype=np.uint8)
    y0 = (image_size - new_h) // 2
    x0 = (image_size - new_w) // 2
    canvas[y0 : y0 + new_h, x0 : x0 + new_w] = resized

    arr = canvas.astype(np.float32) / 255.0
    arr = np.expand_dims(arr, axis=-1)
    return arr, canvas


def remove_vertical_grid_lines(
    binary_img: np.ndarray,
    invert: bool,
    min_height_ratio: float,
    max_width: int,
    edge_margin: int,
) -> np.ndarray:
    """Remove likely vertical separator lines while preserving digits.

    Strategy:
    1) Build foreground mask (digits + lines = white).
    2) Morphologically extract vertical structures.
    3) Remove only tall, thin components near left/right edges of digit cell.
    """
    h, w = binary_img.shape[:2]
    if h < 3 or w < 3:
        return binary_img

    fg = binary_img.copy() if invert else (255 - binary_img)

    k_h = max(3, int(round(h * max(0.3, min(1.0, min_height_ratio)))))
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, k_h))
    vertical_mask = cv2.morphologyEx(fg, cv2.MORPH_OPEN, kernel)

    n_labels, labels, stats, _ = cv2.connectedComponentsWithStats(vertical_mask, connectivity=8)
    remove_mask = np.zeros_like(fg)
    min_h = int(round(h * max(0.3, min(1.0, min_height_ratio))))
    edge = max(0, int(edge_margin))
    max_w = max(1, int(max_width))

    for i in range(1, n_labels):
        x = stats[i, cv2.CC_STAT_LEFT]
        y = stats[i, cv2.CC_STAT_TOP]
        cw = stats[i, cv2.CC_STAT_WIDTH]
        ch = stats[i, cv2.CC_STAT_HEIGHT]
        if ch < min_h or cw > max_w:
            continue

        near_left = x <= edge
        near_right = (x + cw) >= (w - edge)
        if near_left or near_right:
            remove_mask[labels == i] = 255

    fg_clean = cv2.bitwise_and(fg, cv2.bitwise_not(remove_mask))
    return fg_clean if invert else (255 - fg_clean)


class UDISEOCRApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("UDISE OCR Tuning App")
        self.geometry("1660x980")
        self.minsize(1320, 820)

        self.runner = None
        self.current_image_bgr: Optional[np.ndarray] = None
        self.current_row: Optional[int] = None
        self.current_source: Optional[str] = None

        # ROI state
        self.single_roi: Optional[Tuple[int, int, int, int]] = None
        self.grid_base_boxes: List[Tuple[int, int, int, int]] = []
        self.grid_offsets: List[List[int]] = []  # per digit: [dx,dy,dw,dh]
        self.selected_digit_idx: Optional[int] = None

        # Preview/state
        self.last_digit_previews: List[np.ndarray] = []
        self.last_preds: List[int] = []
        self.last_confs: List[float] = []
        self.photo_refs: Dict[str, ImageTk.PhotoImage] = {}
        self.digit_photo_refs: List[ImageTk.PhotoImage] = []
        self.model_expected_hw: Optional[Tuple[int, int]] = None

        self._build_vars()
        self._build_styles()
        self._build_ui()
        self._on_framework_change()

    def report_callback_exception(self, exc, val, tb):  # type: ignore[override]
        err = self._short_error(val if isinstance(val, Exception) else Exception(str(val)))
        self.log(f"UI callback error: {err}")
        messagebox.showerror("UI Error", err)

    def _build_vars(self) -> None:
        self.excel_var = tk.StringVar(value=DEFAULT_EXCEL)
        self.output_excel_var = tk.StringVar(value="")
        self.sheet_var = tk.StringVar(value="")
        self.url_col_var = tk.StringVar(value="Original Image Path")
        self.out_col_var = tk.StringVar(value="UDISE Found")

        self.framework_var = tk.StringVar(value="keras")
        self.model_path_var = tk.StringVar(value="")
        self.model_config_var = tk.StringVar(value="")
        self.api_key_var = tk.StringVar(value="")
        self.api_model_var = tk.StringVar(value="meta-llama/llama-4-scout-17b-16e-instruct")
        self.api_endpoint_var = tk.StringVar(value=DEFAULT_GROQ_ENDPOINT)
        self.model_ready_var = tk.StringVar(value="Model not loaded")

        self.roi_file_var = tk.StringVar(value=DEFAULT_ROI_FILE)
        self.roi_text_var = tk.StringVar(value="")
        self.roi_mode_var = tk.StringVar(value="single")  # single | grid
        self.grid_gap_var = tk.IntVar(value=0)

        self.timeout_var = tk.IntVar(value=20)
        self.retries_var = tk.IntVar(value=2)
        self.backoff_var = tk.DoubleVar(value=1.5)
        self.start_row_var = tk.IntVar(value=2)
        self.max_rows_var = tk.IntVar(value=0)
        self.overwrite_var = tk.BooleanVar(value=False)
        self.save_every_var = tk.IntVar(value=50)

        self.manual_review_var = tk.BooleanVar(value=False)
        self.conf_threshold_var = tk.DoubleVar(value=0.85)

        self.sample_row_var = tk.IntVar(value=2)

        self.image_size_var = tk.IntVar(value=28)
        self.invert_var = tk.BooleanVar(value=True)
        self.blur_var = tk.IntVar(value=3)
        self.th_mode_var = tk.StringVar(value="otsu")
        self.ad_block_var = tk.IntVar(value=21)
        self.ad_c_var = tk.IntVar(value=10)
        self.tight_crop_var = tk.BooleanVar(value=True)
        self.pad_var = tk.IntVar(value=2)
        self.remove_vlines_var = tk.BooleanVar(value=False)
        self.vline_min_h_ratio_var = tk.DoubleVar(value=0.75)
        self.vline_max_w_var = tk.IntVar(value=3)
        self.vline_edge_margin_var = tk.IntVar(value=4)
        self.morph_clean_var = tk.BooleanVar(value=False)
        self.digit_count_var = tk.IntVar(value=DEFAULT_DIGIT_COUNT)

        self.shift_step_var = tk.IntVar(value=1)
        self.shift_scope_var = tk.StringVar(value="global")  # global | selected
        self.selected_box_var = tk.IntVar(value=0)
        self.box_resize_step_var = tk.IntVar(value=1)
        self.grid_divisor_var = tk.IntVar(value=11)

        self.pred_var = tk.StringVar(value="Prediction: model not loaded")
        self.status_var = tk.StringVar(value="Ready")

    def _build_styles(self) -> None:
        self.style = ttk.Style(self)
        try:
            self.style.configure("Ready.TButton", foreground="white", background="#2e7d32")
        except Exception:
            pass

    def _build_ui(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)

        top = ttk.Frame(self, padding=4)
        top.grid(row=0, column=0, sticky="ew")
        top.columnconfigure(0, weight=1)

        self.top_excel_row = self._row_file(top, 0, "Excel", self.excel_var, self.pick_excel)
        self.top_output_row = self._row_file(top, 1, "Output Excel", self.output_excel_var, self.pick_output_excel)
        self.top_model_row = self._row_file(top, 2, "Model", self.model_path_var, self.pick_model)
        self.top_model_cfg_row = self._row_file(top, 3, "Model Config", self.model_config_var, self.pick_model_config)
        self.top_api_model_row = self._row_text(top, 4, "API Model", self.api_model_var)
        self.top_api_key_row = self._row_text(top, 5, "API Key", self.api_key_var, show="*")
        self.top_api_endpoint_row = self._row_text(top, 6, "API Endpoint", self.api_endpoint_var)

        meta = ttk.Frame(top)
        meta.grid(row=7, column=0, sticky="ew", pady=(4, 0))
        for i in range(8):
            meta.columnconfigure(i, weight=1)

        ttk.Label(meta, text="Sheet").grid(row=0, column=0, sticky="w")
        ttk.Entry(meta, textvariable=self.sheet_var, width=14).grid(row=0, column=1, sticky="ew")

        ttk.Label(meta, text="URL Col/Header").grid(row=0, column=2, sticky="w")
        ttk.Entry(meta, textvariable=self.url_col_var, width=20).grid(row=0, column=3, sticky="ew")

        ttk.Label(meta, text="Output Col/Header").grid(row=0, column=4, sticky="w")
        ttk.Entry(meta, textvariable=self.out_col_var, width=20).grid(row=0, column=5, sticky="ew")

        ttk.Label(meta, text="Framework").grid(row=0, column=6, sticky="w")
        fw_combo = ttk.Combobox(
            meta,
            textvariable=self.framework_var,
            state="readonly",
            values=["keras", "torch", "groq"],
            width=10,
        )
        fw_combo.grid(row=0, column=7, sticky="w")
        fw_combo.bind("<<ComboboxSelected>>", lambda _e: self._on_framework_change())

        quick_actions = ttk.Frame(top)
        quick_actions.grid(row=8, column=0, sticky="ew", pady=(4, 0))
        for i in range(4):
            quick_actions.columnconfigure(i, weight=1)
        self.quick_load_model_btn = ttk.Button(quick_actions, text="Load Model", command=self.load_model)
        self.quick_load_model_btn.grid(row=0, column=0, sticky="ew", padx=2)
        ttk.Button(quick_actions, text="Refresh Preview", command=self.refresh_preview).grid(row=0, column=1, sticky="ew", padx=2)
        ttk.Button(quick_actions, text="Preview Morphology", command=self.preview_morphology_preprocess).grid(
            row=0, column=2, sticky="ew", padx=2
        )
        ttk.Button(quick_actions, text="Run Batch", command=self.run_batch_async).grid(row=0, column=3, sticky="ew", padx=2)

        main = ttk.Panedwindow(self, orient=tk.HORIZONTAL)
        main.grid(row=1, column=0, sticky="nsew")

        left = ttk.Frame(main, padding=8)
        right = ttk.Frame(main, padding=8)
        main.add(left, weight=2)
        main.add(right, weight=3)

        self._build_left_panel(left)
        self._build_right_panel(right)

        status = ttk.Label(self, textvariable=self.status_var, anchor="w", padding=6)
        status.grid(row=2, column=0, sticky="ew")

    def _row_file(self, parent, row, label, var, cmd):
        frame = ttk.Frame(parent)
        frame.grid(row=row, column=0, sticky="ew", pady=1)
        frame.columnconfigure(1, weight=1)
        ttk.Label(frame, text=label, width=11, anchor="e").grid(row=0, column=0, sticky="e", padx=(0, 6))
        ttk.Entry(frame, textvariable=var).grid(row=0, column=1, sticky="ew")
        ttk.Button(frame, text="Browse", command=cmd, width=8).grid(row=0, column=2, sticky="e", padx=(6, 0))
        return frame

    def _row_text(self, parent, row, label, var, show: Optional[str] = None):
        frame = ttk.Frame(parent)
        frame.grid(row=row, column=0, sticky="ew", pady=1)
        frame.columnconfigure(1, weight=1)
        ttk.Label(frame, text=label, width=11, anchor="e").grid(row=0, column=0, sticky="e", padx=(0, 6))
        entry_kwargs = {"textvariable": var}
        if show is not None:
            entry_kwargs["show"] = show
        ttk.Entry(frame, **entry_kwargs).grid(row=0, column=1, sticky="ew")
        return frame

    def _build_left_panel(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)

        notebook = ttk.Notebook(parent)
        notebook.grid(row=0, column=0, sticky="nsew")

        calibration = ttk.Frame(notebook, padding=8)
        production = ttk.Frame(notebook, padding=8)
        logs_tab = ttk.Frame(notebook, padding=8)
        notebook.add(calibration, text="Calibration")
        notebook.add(production, text="Production")
        notebook.add(logs_tab, text="Logs")

        self._build_calibration_tab(calibration)
        self._build_production_tab(production)
        self._build_logs_tab(logs_tab)

    def _build_logs_tab(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)
        self.log_text = tk.Text(parent, wrap="word")
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scroll = ttk.Scrollbar(parent, orient="vertical", command=self.log_text.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.log_text.configure(yscrollcommand=scroll.set)

    def _build_calibration_tab(self, parent):
        parent.columnconfigure(0, weight=1)

        roi_box = ttk.LabelFrame(parent, text="ROI & Sample", padding=8)
        roi_box.grid(row=0, column=0, sticky="ew")
        for i in range(7):
            roi_box.columnconfigure(i, weight=1)

        ttk.Label(roi_box, text="ROI File").grid(row=0, column=0, sticky="w")
        ttk.Entry(roi_box, textvariable=self.roi_file_var).grid(row=0, column=1, columnspan=4, sticky="ew")
        ttk.Button(roi_box, text="Load ROI", command=self.load_roi_from_file).grid(row=0, column=5, sticky="ew")
        ttk.Button(roi_box, text="Save ROI", command=self.save_roi_to_file).grid(row=0, column=6, sticky="ew")

        ttk.Label(roi_box, text="ROI Mode").grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Radiobutton(roi_box, text="Single ROI", variable=self.roi_mode_var, value="single", command=self.refresh_preview).grid(row=1, column=1, sticky="w", pady=(6, 0))
        ttk.Radiobutton(roi_box, text="Grid ROI (11 boxes)", variable=self.roi_mode_var, value="grid", command=self.refresh_preview).grid(row=1, column=2, columnspan=2, sticky="w", pady=(6, 0))

        ttk.Label(roi_box, text="Grid Gap px").grid(row=1, column=4, sticky="e", pady=(6, 0))
        ttk.Entry(roi_box, textvariable=self.grid_gap_var, width=8).grid(row=1, column=5, sticky="w", pady=(6, 0))
        ttk.Button(roi_box, text="Fine-tune Grid", command=self.open_grid_tuner).grid(row=1, column=6, sticky="ew", pady=(6, 0))

        ttk.Label(roi_box, text="Single ROI x,y,w,h").grid(row=2, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(roi_box, textvariable=self.roi_text_var).grid(row=2, column=1, columnspan=3, sticky="ew", pady=(6, 0))
        ttk.Button(roi_box, text="Apply ROI", command=self.apply_roi_text).grid(row=2, column=4, sticky="ew", pady=(6, 0))
        ttk.Button(roi_box, text="Pick ROI", command=self.pick_roi_from_sample).grid(row=2, column=5, sticky="ew", pady=(6, 0))
        ttk.Button(roi_box, text="Pick Grid Start", command=self.pick_grid_start_box).grid(row=2, column=6, sticky="ew", pady=(6, 0))

        ttk.Label(roi_box, text="Sample Row").grid(row=3, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(roi_box, textvariable=self.sample_row_var, width=8).grid(row=3, column=1, sticky="w", pady=(6, 0))
        ttk.Button(roi_box, text="Load Sample Row", command=self.load_sample_row).grid(row=3, column=2, columnspan=2, sticky="ew", pady=(6, 0))
        ttk.Button(roi_box, text="Find Next Pending", command=self.find_next_pending).grid(row=3, column=4, columnspan=3, sticky="ew", pady=(6, 0))

        pp_box = ttk.LabelFrame(parent, text="Preprocess Controls", padding=8)
        pp_box.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        for i in range(4):
            pp_box.columnconfigure(i, weight=1)

        ttk.Label(pp_box, text="Digit Count").grid(row=0, column=0, sticky="w")
        ttk.Spinbox(pp_box, from_=1, to=20, textvariable=self.digit_count_var, width=8).grid(row=0, column=1, sticky="w")

        ttk.Label(pp_box, text="Image Size").grid(row=0, column=2, sticky="w")
        ttk.Spinbox(pp_box, from_=8, to=96, textvariable=self.image_size_var, width=8).grid(row=0, column=3, sticky="w")

        ttk.Checkbutton(pp_box, text="Invert", variable=self.invert_var).grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Checkbutton(pp_box, text="Tight Crop", variable=self.tight_crop_var).grid(row=1, column=1, sticky="w", pady=(6, 0))

        ttk.Label(pp_box, text="Blur Kernel").grid(row=1, column=2, sticky="w", pady=(6, 0))
        ttk.Spinbox(pp_box, from_=1, to=11, increment=2, textvariable=self.blur_var, width=8).grid(row=1, column=3, sticky="w", pady=(6, 0))

        ttk.Label(pp_box, text="Threshold").grid(row=2, column=0, sticky="w", pady=(6, 0))
        ttk.Combobox(pp_box, textvariable=self.th_mode_var, state="readonly", values=["otsu", "adaptive"], width=12).grid(row=2, column=1, sticky="w", pady=(6, 0))

        ttk.Label(pp_box, text="Adaptive Block").grid(row=2, column=2, sticky="w", pady=(6, 0))
        ttk.Spinbox(pp_box, from_=3, to=51, increment=2, textvariable=self.ad_block_var, width=8).grid(row=2, column=3, sticky="w", pady=(6, 0))

        ttk.Label(pp_box, text="Adaptive C").grid(row=3, column=0, sticky="w", pady=(6, 0))
        ttk.Spinbox(pp_box, from_=-20, to=30, textvariable=self.ad_c_var, width=8).grid(row=3, column=1, sticky="w", pady=(6, 0))

        ttk.Label(pp_box, text="BBox Padding").grid(row=3, column=2, sticky="w", pady=(6, 0))
        ttk.Spinbox(pp_box, from_=0, to=20, textvariable=self.pad_var, width=8).grid(row=3, column=3, sticky="w", pady=(6, 0))

        ttk.Checkbutton(pp_box, text="Remove Vertical Grid Lines", variable=self.remove_vlines_var).grid(
            row=4, column=0, columnspan=2, sticky="w", pady=(6, 0)
        )
        ttk.Checkbutton(pp_box, text="Morphology Grid Cleaner", variable=self.morph_clean_var).grid(
            row=6, column=0, columnspan=2, sticky="w", pady=(6, 0)
        )
        ttk.Label(pp_box, text="Line Min H Ratio").grid(row=4, column=2, sticky="w", pady=(6, 0))
        ttk.Spinbox(pp_box, from_=0.3, to=1.0, increment=0.05, textvariable=self.vline_min_h_ratio_var, width=8).grid(
            row=4, column=3, sticky="w", pady=(6, 0)
        )

        ttk.Label(pp_box, text="Line Max Width").grid(row=5, column=0, sticky="w", pady=(6, 0))
        ttk.Spinbox(pp_box, from_=1, to=10, textvariable=self.vline_max_w_var, width=8).grid(
            row=5, column=1, sticky="w", pady=(6, 0)
        )
        ttk.Label(pp_box, text="Edge Margin").grid(row=5, column=2, sticky="w", pady=(6, 0))
        ttk.Spinbox(pp_box, from_=0, to=20, textvariable=self.vline_edge_margin_var, width=8).grid(
            row=5, column=3, sticky="w", pady=(6, 0)
        )

        actions = ttk.Frame(parent)
        actions.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        actions.columnconfigure(0, weight=1)
        actions.columnconfigure(1, weight=1)
        actions.columnconfigure(2, weight=1)
        actions.columnconfigure(3, weight=1)

        self.load_model_btn = ttk.Button(actions, text="Load Model", command=self.load_model)
        self.load_model_btn.grid(row=0, column=0, sticky="ew", padx=2)
        ttk.Button(actions, text="Refresh Preview", command=self.refresh_preview).grid(row=0, column=1, sticky="ew", padx=2)
        ttk.Button(actions, text="Preview Morphology Preprocess", command=self.preview_morphology_preprocess).grid(
            row=0, column=2, sticky="ew", padx=2
        )
        ttk.Button(actions, text="Run Batch", command=self.run_batch_async).grid(row=0, column=3, sticky="ew", padx=2)

        ttk.Label(parent, textvariable=self.model_ready_var).grid(row=3, column=0, sticky="w", pady=(8, 0))

    def _build_production_tab(self, parent):
        parent.columnconfigure(0, weight=1)

        run_cfg = ttk.LabelFrame(parent, text="Batch Settings", padding=8)
        run_cfg.grid(row=0, column=0, sticky="ew")
        for i in range(4):
            run_cfg.columnconfigure(i, weight=1)

        ttk.Label(run_cfg, text="Start Row").grid(row=0, column=0, sticky="w")
        ttk.Entry(run_cfg, textvariable=self.start_row_var, width=8).grid(row=0, column=1, sticky="w")

        ttk.Label(run_cfg, text="Max Rows (0=all)").grid(row=0, column=2, sticky="w")
        ttk.Entry(run_cfg, textvariable=self.max_rows_var, width=8).grid(row=0, column=3, sticky="w")

        ttk.Label(run_cfg, text="Timeout").grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(run_cfg, textvariable=self.timeout_var, width=8).grid(row=1, column=1, sticky="w", pady=(6, 0))

        ttk.Label(run_cfg, text="Retries").grid(row=1, column=2, sticky="w", pady=(6, 0))
        ttk.Entry(run_cfg, textvariable=self.retries_var, width=8).grid(row=1, column=3, sticky="w", pady=(6, 0))

        ttk.Label(run_cfg, text="Backoff").grid(row=2, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(run_cfg, textvariable=self.backoff_var, width=8).grid(row=2, column=1, sticky="w", pady=(6, 0))

        ttk.Label(run_cfg, text="Save Every").grid(row=2, column=2, sticky="w", pady=(6, 0))
        ttk.Entry(run_cfg, textvariable=self.save_every_var, width=8).grid(row=2, column=3, sticky="w", pady=(6, 0))

        ttk.Checkbutton(run_cfg, text="Overwrite Existing", variable=self.overwrite_var).grid(row=3, column=0, columnspan=2, sticky="w", pady=(6, 0))

        ttk.Checkbutton(run_cfg, text="Manual Review", variable=self.manual_review_var).grid(row=3, column=2, sticky="w", pady=(6, 0))
        ttk.Label(run_cfg, text="Threshold").grid(row=3, column=3, sticky="e", pady=(6, 0))
        ttk.Entry(run_cfg, textvariable=self.conf_threshold_var, width=6).grid(row=3, column=3, sticky="w", padx=(60, 0), pady=(6, 0))

        ttk.Label(
            parent,
            text="If Manual Review is enabled, batch pauses when any digit confidence is below threshold.",
            wraplength=430,
        ).grid(row=1, column=0, sticky="w", pady=(8, 0))

    def _build_right_panel(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=3)
        parent.rowconfigure(1, weight=2)

        preview = ttk.LabelFrame(parent, text="Image Preview", padding=8)
        preview.grid(row=0, column=0, sticky="nsew")
        preview.columnconfigure(0, weight=1)
        preview.columnconfigure(1, weight=1)

        self.full_img_label = ttk.Label(preview, text="Full image")
        self.full_img_label.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        self.roi_img_label = ttk.Label(preview, text="ROI preview")
        self.roi_img_label.grid(row=0, column=1, sticky="nsew")

        digits = ttk.LabelFrame(parent, text="Preprocessed Digits", padding=8)
        digits.grid(row=1, column=0, sticky="nsew", pady=(8, 0))
        digits.columnconfigure(0, weight=0)
        digits.columnconfigure(1, weight=1)
        digits.rowconfigure(0, weight=1)

        # Shift/box controls at left, digit strip at right.
        dpad = ttk.LabelFrame(digits, text="Box Controls", padding=6)
        dpad.grid(row=0, column=0, sticky="nsw", padx=(0, 8))
        ttk.Label(dpad, text="Move").grid(row=0, column=0, columnspan=3, sticky="w")
        ttk.Button(dpad, text="↑", width=3, command=lambda: self.shift_all(0, -self._shift_step())).grid(row=1, column=1, padx=2, pady=1)
        ttk.Button(dpad, text="←", width=3, command=lambda: self.shift_all(-self._shift_step(), 0)).grid(row=2, column=0, padx=2, pady=1)
        ttk.Button(dpad, text="→", width=3, command=lambda: self.shift_all(self._shift_step(), 0)).grid(row=2, column=2, padx=2, pady=1)
        ttk.Button(dpad, text="↓", width=3, command=lambda: self.shift_all(0, self._shift_step())).grid(row=3, column=1, padx=2, pady=1)
        ttk.Label(dpad, text="Step").grid(row=1, column=3, padx=(10, 2), sticky="e")
        ttk.Entry(dpad, textvariable=self.shift_step_var, width=5).grid(row=1, column=4, sticky="w")
        ttk.Label(dpad, text="Scope").grid(row=2, column=3, padx=(10, 2), sticky="e")
        ttk.Combobox(
            dpad,
            textvariable=self.shift_scope_var,
            state="readonly",
            values=["global", "selected"],
            width=9,
        ).grid(row=2, column=4, sticky="w")
        ttk.Label(dpad, text="Box").grid(row=3, column=3, padx=(10, 2), sticky="e")
        ttk.Spinbox(dpad, from_=0, to=30, textvariable=self.selected_box_var, width=5).grid(row=3, column=4, sticky="w")

        ttk.Label(dpad, text="Box Width").grid(row=4, column=0, sticky="w", pady=(6, 0))
        ttk.Button(dpad, text="-", width=3, command=lambda: self.resize_selected_box(-self._resize_step())).grid(row=4, column=1, pady=(6, 0))
        ttk.Button(dpad, text="+", width=3, command=lambda: self.resize_selected_box(self._resize_step())).grid(row=4, column=2, pady=(6, 0))
        ttk.Label(dpad, text="Step").grid(row=4, column=3, padx=(10, 2), pady=(6, 0), sticky="e")
        ttk.Entry(dpad, textvariable=self.box_resize_step_var, width=5).grid(row=4, column=4, pady=(6, 0), sticky="w")

        ttk.Label(dpad, text="Grid = ROI / X").grid(row=5, column=0, columnspan=2, sticky="w", pady=(6, 0))
        ttk.Entry(dpad, textvariable=self.grid_divisor_var, width=6).grid(row=5, column=2, pady=(6, 0), sticky="w")
        ttk.Button(dpad, text="Apply", command=self.rebuild_grid_by_divisor).grid(row=5, column=3, columnspan=2, pady=(6, 0), sticky="ew")

        self.digit_canvas = tk.Canvas(digits, height=220, highlightthickness=0)
        self.digit_canvas.grid(row=0, column=1, sticky="nsew")
        self.digit_scroll_x = ttk.Scrollbar(digits, orient="horizontal", command=self.digit_canvas.xview)
        self.digit_scroll_x.grid(row=1, column=1, sticky="ew")
        self.digit_scroll_y = ttk.Scrollbar(digits, orient="vertical", command=self.digit_canvas.yview)
        self.digit_scroll_y.grid(row=0, column=2, sticky="ns")
        self.digit_canvas.configure(xscrollcommand=self.digit_scroll_x.set, yscrollcommand=self.digit_scroll_y.set)
        self.digit_frame = ttk.Frame(self.digit_canvas)
        self.digit_canvas_window = self.digit_canvas.create_window((0, 0), window=self.digit_frame, anchor="nw")
        self.digit_frame.bind("<Configure>", lambda _e: self.digit_canvas.configure(scrollregion=self.digit_canvas.bbox("all")))

        ttk.Label(digits, textvariable=self.pred_var, font=("TkDefaultFont", 11, "bold")).grid(
            row=2, column=0, columnspan=3, sticky="w", pady=(6, 0)
        )

    # ---------- Utility ----------
    def log(self, msg: str) -> None:
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.status_var.set(msg)

    def _shift_step(self) -> int:
        try:
            return max(1, int(self.shift_step_var.get()))
        except Exception:
            return 1

    def _resize_step(self) -> int:
        try:
            return max(1, int(self.box_resize_step_var.get()))
        except Exception:
            return 1

    def _selected_box_idx(self) -> int:
        idx = int(self.selected_box_var.get())
        max_idx = max(0, int(self.digit_count_var.get()) - 1)
        idx = max(0, min(max_idx, idx))
        self.selected_box_var.set(idx)
        self.selected_digit_idx = idx
        return idx

    def _on_framework_change(self) -> None:
        fw = (self.framework_var.get() or "").strip()
        if hasattr(self, "load_model_btn"):
            if fw == "groq":
                self.load_model_btn.configure(text="Load Groq API")
            else:
                self.load_model_btn.configure(text="Load Local Model")
        if hasattr(self, "quick_load_model_btn"):
            if fw == "groq":
                self.quick_load_model_btn.configure(text="Load Groq API")
            else:
                self.quick_load_model_btn.configure(text="Load Local Model")
        if hasattr(self, "top_model_row"):
            if fw == "groq":
                self.top_model_row.grid_remove()
                self.top_model_cfg_row.grid_remove()
                self.top_api_model_row.grid()
                self.top_api_key_row.grid()
                self.top_api_endpoint_row.grid()
            else:
                self.top_model_row.grid()
                self.top_model_cfg_row.grid()
                self.top_api_model_row.grid_remove()
                self.top_api_key_row.grid_remove()
                self.top_api_endpoint_row.grid_remove()

    def _has_groq_credentials(self) -> bool:
        return bool(self.api_key_var.get().strip()) and bool(self.api_model_var.get().strip())

    def resize_selected_box(self, dw: int) -> None:
        if self.roi_mode_var.get() != "grid":
            self.log("Per-box width adjustment is available in Grid ROI mode.")
            return
        if not self.grid_base_boxes:
            return
        idx = self._selected_box_idx()
        while len(self.grid_offsets) < len(self.grid_base_boxes):
            self.grid_offsets.append([0, 0, 0, 0])
        self.grid_offsets[idx][2] += int(dw)
        self.log(f"Adjusted box {idx} width by {dw}px")
        self.refresh_preview()

    def rebuild_grid_by_divisor(self) -> None:
        try:
            div = max(1, int(self.grid_divisor_var.get()))
        except Exception:
            messagebox.showerror("Grid Divisor", "Enter a valid integer divisor.")
            return

        base_roi: Optional[Tuple[int, int, int, int]] = None
        if self.single_roi is not None:
            base_roi = self.single_roi
        elif self.grid_base_boxes:
            x0 = min(x for x, _, _, _ in self.grid_base_boxes)
            y0 = min(y for _, y, _, _ in self.grid_base_boxes)
            x1 = max(x + w for x, _, w, _ in self.grid_base_boxes)
            y1 = max(y + h for _, y, _, h in self.grid_base_boxes)
            base_roi = (x0, y0, max(1, x1 - x0), max(1, y1 - y0))
        elif self.current_image_bgr is not None:
            ih, iw = self.current_image_bgr.shape[:2]
            base_roi = (0, 0, iw, ih)

        if base_roi is None:
            messagebox.showerror("Grid Divisor", "Load sample and set ROI/grid first.")
            return

        x, y, w, h = base_roi
        cell_w = max(1, int(round(w / float(div))))
        self.grid_base_boxes = []
        for i in range(div):
            xi = x + int(round(i * (w / float(div))))
            if i == div - 1:
                wi = max(1, (x + w) - xi)
            else:
                wi = cell_w
            self.grid_base_boxes.append((xi, y, wi, h))
        self.grid_offsets = [[0, 0, 0, 0] for _ in range(div)]
        self.digit_count_var.set(div)
        self.roi_mode_var.set("grid")
        self.selected_box_var.set(0)
        self.selected_digit_idx = 0
        self.log(f"Rebuilt grid using divisor {div}: ROI split into {div} boxes")
        self.refresh_preview()

    def _settings(self) -> PreprocessSettings:
        return PreprocessSettings(
            image_size=max(8, int(self.image_size_var.get())),
            invert=bool(self.invert_var.get()),
            blur_kernel=max(1, int(self.blur_var.get())),
            threshold_mode=self.th_mode_var.get().strip() or "otsu",
            adaptive_block_size=max(3, int(self.ad_block_var.get())),
            adaptive_c=int(self.ad_c_var.get()),
            tight_crop=bool(self.tight_crop_var.get()),
            bbox_padding=max(0, int(self.pad_var.get())),
            remove_vertical_lines=bool(self.remove_vlines_var.get()),
            line_min_height_ratio=max(0.3, min(1.0, float(self.vline_min_h_ratio_var.get()))),
            line_max_width=max(1, int(self.vline_max_w_var.get())),
            line_edge_margin=max(0, int(self.vline_edge_margin_var.get())),
            use_morphology_cleaning=bool(self.morph_clean_var.get()),
        )

    def _to_photo(self, bgr_or_gray: np.ndarray, max_size=(620, 420)):
        if bgr_or_gray.ndim == 2:
            rgb = cv2.cvtColor(bgr_or_gray, cv2.COLOR_GRAY2RGB)
        else:
            rgb = cv2.cvtColor(bgr_or_gray, cv2.COLOR_BGR2RGB)
        h, w = rgb.shape[:2]
        scale = min(max_size[0] / w, max_size[1] / h, 1.0)
        if scale < 1.0:
            rgb = cv2.resize(rgb, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_AREA)
        return ImageTk.PhotoImage(Image.fromarray(rgb))

    def _photo_for_widget(self, arr: np.ndarray, widget, fallback=(620, 420), pad=(16, 16)):
        try:
            ww = max(1, int(widget.winfo_width()) - int(pad[0]))
            wh = max(1, int(widget.winfo_height()) - int(pad[1]))
            # If widget is not yet laid out, fall back to a sensible default.
            if ww <= 10 or wh <= 10:
                return self._to_photo(arr, max_size=fallback)
            return self._to_photo(arr, max_size=(ww, wh))
        except Exception:
            return self._to_photo(arr, max_size=fallback)

    def _build_digit_strip(self, digits: Sequence[np.ndarray], tile_size: int = 64, gap: int = 6) -> np.ndarray:
        if not digits:
            return np.full((tile_size, tile_size, 3), 220, dtype=np.uint8)
        out_w = (tile_size * len(digits)) + gap * max(0, len(digits) - 1)
        canvas = np.full((tile_size, out_w, 3), 235, dtype=np.uint8)
        x = 0
        for d in digits:
            if d.ndim == 2:
                tile = d
            else:
                tile = cv2.cvtColor(d, cv2.COLOR_BGR2GRAY)
            tile = cv2.resize(tile, (tile_size, tile_size), interpolation=cv2.INTER_NEAREST)
            tile_rgb = cv2.cvtColor(tile, cv2.COLOR_GRAY2BGR)
            canvas[:, x : x + tile_size] = tile_rgb
            x += tile_size + gap
        return canvas

    def _short_error(self, exc: Exception) -> str:
        msg = str(exc).strip()
        if not msg:
            return exc.__class__.__name__

        # Keep UI readable by showing only the first useful line.
        first = msg.splitlines()[0].strip()

        if "Matrix size-incompatible" in msg or "shape" in msg.lower():
            if self.model_expected_hw:
                eh, ew = self.model_expected_hw
                return (
                    f"Model input mismatch. Expected {eh}x{ew}, current Image Size is "
                    f"{int(self.image_size_var.get())}. Set Image Size to {eh} and retry."
                )
            return "Model input mismatch. Check Image Size and model input shape."
        return first

    def _expected_hw_from_runner(self) -> Optional[Tuple[int, int]]:
        if isinstance(self.runner, KerasRunner):
            shape = tuple(self.runner.input_shape)
            if len(shape) == 4:
                # NHWC
                if shape[-1] in (1, 3) and shape[1] and shape[2]:
                    return int(shape[1]), int(shape[2])
                # NCHW
                if shape[1] in (1, 3) and shape[2] and shape[3]:
                    return int(shape[2]), int(shape[3])
        return None

    def _auto_match_image_size_to_model(self) -> None:
        hw = self._expected_hw_from_runner()
        self.model_expected_hw = hw
        if hw and hw[0] == hw[1]:
            expected = int(hw[0])
            if int(self.image_size_var.get()) != expected:
                self.image_size_var.set(expected)
                self.log(f"Adjusted Image Size to model expected {expected}x{expected}")

    def _select_roi_with_zoom(self, window_title: str, image: np.ndarray) -> Optional[Tuple[int, int, int, int]]:
        src = image.copy()
        h, w = src.shape[:2]
        rgb = cv2.cvtColor(src, cv2.COLOR_BGR2RGB)

        top = tk.Toplevel(self)
        top.title(window_title)
        top.geometry("1180x820")
        top.transient(self)
        top.grab_set()

        header = ttk.Frame(top, padding=(8, 8, 8, 4))
        header.pack(fill="x")
        ttk.Label(
            header,
            text=(
                "Left drag: draw ROI  |  Mouse wheel: zoom  |  Right drag: pan  |  "
                "Confirm to apply ROI"
            ),
        ).pack(side="left")

        status_var = tk.StringVar(value="ROI: -")
        ttk.Label(header, textvariable=status_var).pack(side="right")

        body = ttk.Frame(top, padding=(8, 0, 8, 0))
        body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        canvas = tk.Canvas(body, background="#1e1e1e", highlightthickness=0, cursor="crosshair")
        hbar = ttk.Scrollbar(body, orient="horizontal", command=canvas.xview)
        vbar = ttk.Scrollbar(body, orient="vertical", command=canvas.yview)
        canvas.configure(xscrollcommand=hbar.set, yscrollcommand=vbar.set)

        canvas.grid(row=0, column=0, sticky="nsew")
        vbar.grid(row=0, column=1, sticky="ns")
        hbar.grid(row=1, column=0, sticky="ew")

        footer = ttk.Frame(top, padding=8)
        footer.pack(fill="x")

        state = {
            "scale": 1.0,
            "photo": None,
            "start": None,
            "rect_canvas": None,
            "roi_src": None,
        }
        img_tag = "IMG"
        roi_tag = "ROI_RECT"

        def _clamp_roi(x0: float, y0: float, x1: float, y1: float) -> Optional[Tuple[int, int, int, int]]:
            ax0 = int(round(max(0, min(x0, x1))))
            ay0 = int(round(max(0, min(y0, y1))))
            ax1 = int(round(min(w, max(x0, x1))))
            ay1 = int(round(min(h, max(y0, y1))))
            rw = ax1 - ax0
            rh = ay1 - ay0
            if rw <= 0 or rh <= 0:
                return None
            return ax0, ay0, rw, rh

        def _draw_roi() -> None:
            canvas.delete(roi_tag)
            roi = state["roi_src"]
            if not roi:
                return
            x, y, rw, rh = roi
            s = float(state["scale"])
            canvas.create_rectangle(
                int(round(x * s)),
                int(round(y * s)),
                int(round((x + rw) * s)),
                int(round((y + rh) * s)),
                outline="#00ff7f",
                width=2,
                tags=roi_tag,
            )

        def _render() -> None:
            s = float(state["scale"])
            if s <= 1.0:
                disp = cv2.resize(rgb, (max(1, int(w * s)), max(1, int(h * s))), interpolation=cv2.INTER_AREA)
            else:
                disp = cv2.resize(rgb, (max(1, int(w * s)), max(1, int(h * s))), interpolation=cv2.INTER_LINEAR)
            photo = ImageTk.PhotoImage(Image.fromarray(disp))
            state["photo"] = photo
            canvas.delete(img_tag)
            canvas.create_image(0, 0, image=photo, anchor="nw", tags=img_tag)
            canvas.configure(scrollregion=(0, 0, disp.shape[1], disp.shape[0]))
            _draw_roi()

        def _canvas_to_src(cx: float, cy: float) -> Tuple[float, float]:
            s = float(state["scale"])
            return cx / s, cy / s

        def on_left_down(event):
            cx = canvas.canvasx(event.x)
            cy = canvas.canvasy(event.y)
            sx, sy = _canvas_to_src(cx, cy)
            state["start"] = (sx, sy)
            state["roi_src"] = None
            canvas.delete(roi_tag)

        def on_left_drag(event):
            if state["start"] is None:
                return
            cx = canvas.canvasx(event.x)
            cy = canvas.canvasy(event.y)
            sx, sy = state["start"]
            ex, ey = _canvas_to_src(cx, cy)
            roi = _clamp_roi(sx, sy, ex, ey)
            state["roi_src"] = roi
            if roi:
                x, y, rw, rh = roi
                status_var.set(f"ROI: {x},{y},{rw},{rh}")
            else:
                status_var.set("ROI: -")
            _draw_roi()

        def on_left_up(_event):
            state["start"] = None

        def on_pan_start(event):
            canvas.scan_mark(event.x, event.y)

        def on_pan_drag(event):
            canvas.scan_dragto(event.x, event.y, gain=1)

        def _zoom_by(factor: float, event=None):
            old_s = float(state["scale"])
            new_s = max(0.2, min(6.0, old_s * factor))
            if abs(new_s - old_s) < 1e-9:
                return

            if event is not None:
                cx = canvas.canvasx(event.x)
                cy = canvas.canvasy(event.y)
            else:
                cx = canvas.canvasx(canvas.winfo_width() * 0.5)
                cy = canvas.canvasy(canvas.winfo_height() * 0.5)

            rel_x = cx / max(1e-6, old_s)
            rel_y = cy / max(1e-6, old_s)
            state["scale"] = new_s
            _render()

            new_cx = rel_x * new_s
            new_cy = rel_y * new_s
            sr = canvas.bbox(img_tag)
            if not sr:
                return
            total_w = max(1, sr[2] - sr[0])
            total_h = max(1, sr[3] - sr[1])
            nx = (new_cx - event.x) / total_w if event is not None else 0.0
            ny = (new_cy - event.y) / total_h if event is not None else 0.0
            canvas.xview_moveto(max(0.0, min(1.0, nx)))
            canvas.yview_moveto(max(0.0, min(1.0, ny)))

        def on_mousewheel(event):
            delta = getattr(event, "delta", 0)
            if delta > 0:
                _zoom_by(1.15, event)
            elif delta < 0:
                _zoom_by(1 / 1.15, event)

        def on_mousewheel_linux(event):
            if getattr(event, "num", 0) == 4:
                _zoom_by(1.15, event)
            elif getattr(event, "num", 0) == 5:
                _zoom_by(1 / 1.15, event)

        result_holder: Dict[str, Optional[Tuple[int, int, int, int]]] = {"roi": None}

        def confirm():
            roi = state["roi_src"]
            if roi is None:
                messagebox.showerror("ROI", "Draw ROI first.")
                return
            result_holder["roi"] = roi
            top.destroy()

        def cancel():
            result_holder["roi"] = None
            top.destroy()

        def zoom_in():
            _zoom_by(1.15)

        def zoom_out():
            _zoom_by(1 / 1.15)

        def zoom_reset():
            state["scale"] = 1.0
            _render()
            canvas.xview_moveto(0.0)
            canvas.yview_moveto(0.0)

        ttk.Button(footer, text="Zoom -", command=zoom_out).pack(side="left", padx=2)
        ttk.Button(footer, text="Zoom +", command=zoom_in).pack(side="left", padx=2)
        ttk.Button(footer, text="100%", command=zoom_reset).pack(side="left", padx=2)
        ttk.Button(footer, text="Cancel", command=cancel).pack(side="right", padx=2)
        ttk.Button(footer, text="Confirm ROI", command=confirm).pack(side="right", padx=2)

        canvas.bind("<ButtonPress-1>", on_left_down)
        canvas.bind("<B1-Motion>", on_left_drag)
        canvas.bind("<ButtonRelease-1>", on_left_up)
        canvas.bind("<ButtonPress-3>", on_pan_start)
        canvas.bind("<B3-Motion>", on_pan_drag)
        canvas.bind("<MouseWheel>", on_mousewheel)
        canvas.bind("<Button-4>", on_mousewheel_linux)
        canvas.bind("<Button-5>", on_mousewheel_linux)

        _render()
        top.protocol("WM_DELETE_WINDOW", cancel)
        self.wait_window(top)
        return result_holder["roi"]

    def _safe_select_roi(self, window_title: str, image: np.ndarray) -> Optional[Tuple[int, int, int, int]]:
        try:
            return self._select_roi_with_zoom(window_title, image)
        except Exception as exc:
            self.log(f"Zoom ROI picker error, falling back to OpenCV: {self._short_error(exc)}")
            try:
                x, y, w, h = cv2.selectROI(
                    window_title,
                    image.copy(),
                    showCrosshair=True,
                    fromCenter=False,
                )
                cv2.destroyAllWindows()
                if w <= 0 or h <= 0:
                    return None
                return int(x), int(y), int(w), int(h)
            except Exception as cv_exc:
                messagebox.showerror(
                    "ROI Error",
                    (
                        f"{self._short_error(cv_exc)}\n\n"
                        "Both zoom ROI picker and OpenCV ROI picker failed."
                    ),
                )
                return None

    def _open_sheet(self):
        excel_path = Path(self.excel_var.get().strip())
        if not excel_path.exists():
            raise RuntimeError(f"Excel not found: {excel_path}")
        wb = load_workbook(excel_path)
        ws = resolve_sheet(wb, self.sheet_var.get().strip() or None)
        return wb, ws

    def _resolve_cols(self, ws):
        url_ref = self.url_col_var.get().strip() or None
        out_ref = self.out_col_var.get().strip() or None
        url_col = resolve_column(ws, url_ref, ["original image path", "image path on server", "image url", "url"], "H")
        out_col = resolve_column(ws, out_ref, ["udise found", "udise", "udise code"], "E", allow_create=True)
        return url_col, out_col

    def _effective_grid_boxes(self) -> List[Tuple[int, int, int, int]]:
        boxes: List[Tuple[int, int, int, int]] = []
        if not self.grid_base_boxes:
            return boxes
        for i, base in enumerate(self.grid_base_boxes):
            dx, dy, dw, dh = self.grid_offsets[i] if i < len(self.grid_offsets) else [0, 0, 0, 0]
            x, y, w, h = base
            w2 = max(1, w + int(dw))
            h2 = max(1, h + int(dh))
            boxes.append((int(x + dx), int(y + dy), int(w2), int(h2)))
        return boxes

    def _digit_boxes_for_image(self, image_bgr: np.ndarray) -> List[Tuple[int, int, int, int]]:
        if self.roi_mode_var.get() == "grid":
            boxes = self._effective_grid_boxes()
            clamped: List[Tuple[int, int, int, int]] = []
            ih, iw = image_bgr.shape[:2]
            for x, y, w, h in boxes:
                x0 = max(0, x)
                y0 = max(0, y)
                x1 = min(iw, x + w)
                y1 = min(ih, y + h)
                # Keep index alignment even if partially/fully outside image.
                cw = max(1, x1 - x0)
                ch = max(1, y1 - y0)
                clamped.append((x0, y0, cw, ch))
            return clamped

        if self.single_roi is None:
            return []
        x, y, w, h = self.single_roi
        roi_obj = parse_manual_roi(f"{x},{y},{w},{h}")
        roi_crop = crop_roi(image_bgr, roi_obj)
        cells = split_digits_equal(roi_crop, int(self.digit_count_var.get()))

        # Recreate cell absolute boxes so UI can highlight per digit in full image.
        cell_boxes: List[Tuple[int, int, int, int]] = []
        roi_x, roi_y, roi_w, roi_h = self.single_roi
        cell_w = roi_w / float(int(self.digit_count_var.get()))
        for i in range(int(self.digit_count_var.get())):
            c0 = int(round(i * cell_w))
            c1 = int(round((i + 1) * cell_w))
            c1 = max(c1, c0 + 1)
            cell_boxes.append((roi_x + c0, roi_y, c1 - c0, roi_h))
        return cell_boxes

    def _crop_from_boxes(self, image_bgr: np.ndarray, boxes: Sequence[Tuple[int, int, int, int]]) -> List[np.ndarray]:
        crops: List[np.ndarray] = []
        ih, iw = image_bgr.shape[:2]
        for x, y, w, h in boxes:
            x0 = max(0, x)
            y0 = max(0, y)
            x1 = min(iw, x + w)
            y1 = min(ih, y + h)
            if x1 <= x0 or y1 <= y0:
                # Keep output count stable with a white fallback patch.
                hh = max(8, int(h))
                ww = max(8, int(w))
                crops.append(np.full((hh, ww, 3), 255, dtype=np.uint8))
            else:
                crops.append(image_bgr[y0:y1, x0:x1])
        return crops

    def _normalize_boxes_count(
        self,
        boxes: Sequence[Tuple[int, int, int, int]],
        digit_count: int,
        image_shape: Tuple[int, int, int],
    ) -> List[Tuple[int, int, int, int]]:
        out = list(boxes)
        if len(out) == digit_count:
            return out
        if len(out) > digit_count:
            return out[:digit_count]

        ih, iw = image_shape[:2]
        if out:
            x, y, w, h = out[-1]
        else:
            h = max(12, ih // 20)
            w = max(12, iw // max(1, digit_count))
            y = max(0, (ih - h) // 2)
            x = max(0, (iw - (digit_count * w)) // 2)

        while len(out) < digit_count:
            nx = min(iw - 1, x + w)
            out.append((nx, y, w, h))
            x = nx
        return out

    def _coerce_digit_code(self, preds: Sequence[int], digit_count: int, fill_digit: str = "0") -> str:
        digits = []
        for p in preds:
            try:
                d = int(p)
                if 0 <= d <= 9:
                    digits.append(str(d))
                else:
                    digits.append(fill_digit)
            except Exception:
                digits.append(fill_digit)
        if len(digits) < digit_count:
            digits.extend([fill_digit] * (digit_count - len(digits)))
        return "".join(digits[:digit_count])

    def _predict_with_confidence(self, batch: np.ndarray) -> Tuple[List[int], List[float]]:
        if self.runner is None:
            return [], []

        if hasattr(self.runner, "predict_with_confidence"):
            try:
                preds, confs = self.runner.predict_with_confidence(batch)
                return list(preds), list(confs)
            except Exception as exc:
                raise RuntimeError(self._short_error(exc)) from None

        try:
            if isinstance(self.runner, KerasRunner):
                expected = self._expected_hw_from_runner()
                if expected is not None:
                    got_h, got_w = int(batch.shape[1]), int(batch.shape[2])
                    exp_h, exp_w = expected
                    if (got_h, got_w) != (exp_h, exp_w):
                        raise RuntimeError(
                            f"Model input mismatch. Expected {exp_h}x{exp_w}, got {got_h}x{got_w}."
                        )
                x = self.runner._prepare_input(batch)  # Reuse runner's input adaptation.
                logits = np.asarray(self.runner.model.predict(x, verbose=0))
                preds = logits.argmax(axis=1).astype(int).tolist()
                confs = logits.max(axis=1).astype(float).tolist()
                return preds, confs

            if isinstance(self.runner, TorchRunner):
                x = np.transpose(batch, (0, 3, 1, 2)).astype(np.float32)
                tensor = self.runner.torch.from_numpy(x)
                with self.runner.torch.no_grad():
                    out = self.runner.model(tensor)
                logits = out.detach().cpu().numpy()
                exp = np.exp(logits - logits.max(axis=1, keepdims=True))
                probs = exp / exp.sum(axis=1, keepdims=True)
                preds = probs.argmax(axis=1).astype(int).tolist()
                confs = probs.max(axis=1).astype(float).tolist()
                return preds, confs
        except Exception as exc:
            raise RuntimeError(self._short_error(exc)) from None

        preds = self.runner.predict(batch)
        confs = [1.0 for _ in preds]
        return preds, confs

    # ---------- File pickers ----------
    def pick_excel(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if p:
            self.excel_var.set(p)
            if not self.output_excel_var.get().strip():
                out = Path(p).with_name(f"{Path(p).stem}_with_udise.xlsx")
                self.output_excel_var.set(str(out))

    def pick_output_excel(self):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if p:
            self.output_excel_var.set(p)

    def pick_model(self):
        p = filedialog.askopenfilename(filetypes=[("Model", "*.keras *.h5 *.pt *.pth"), ("All", "*.*")])
        if p:
            self.model_path_var.set(p)
            json_guess = Path(str(p) + ".json")
            if json_guess.exists() and not self.model_config_var.get().strip():
                self.model_config_var.set(str(json_guess))

    def pick_model_config(self):
        p = filedialog.askopenfilename(filetypes=[("JSON", "*.json")])
        if p:
            self.model_config_var.set(p)

    # ---------- ROI / Grid ----------
    def load_roi_from_file(self):
        p = Path(self.roi_file_var.get().strip())
        if not p.exists():
            messagebox.showerror("Error", f"ROI file not found: {p}")
            return
        data = json.loads(p.read_text(encoding="utf-8"))

        mode = str(data.get("mode", "single"))
        self.roi_mode_var.set(mode)

        if mode == "grid":
            self.grid_base_boxes = [tuple(map(int, b)) for b in data.get("grid_base_boxes", [])]
            offsets = data.get("grid_offsets", [])
            self.grid_offsets = [list(map(int, o)) for o in offsets]
            if len(self.grid_offsets) < len(self.grid_base_boxes):
                for _ in range(len(self.grid_base_boxes) - len(self.grid_offsets)):
                    self.grid_offsets.append([0, 0, 0, 0])
            self.log(f"Grid ROI loaded from {p}")
        else:
            x = int(data["x"])
            y = int(data["y"])
            w = int(data["w"])
            h = int(data["h"])
            self.single_roi = (x, y, w, h)
            self.roi_text_var.set(f"{x},{y},{w},{h}")
            self.log(f"Single ROI loaded from {p}")

        self.refresh_preview()

    def save_roi_to_file(self):
        p = Path(self.roi_file_var.get().strip())
        p.parent.mkdir(parents=True, exist_ok=True)
        if self.roi_mode_var.get() == "grid":
            payload = {
                "mode": "grid",
                "grid_base_boxes": self.grid_base_boxes,
                "grid_offsets": self.grid_offsets,
            }
        else:
            if self.single_roi is None:
                messagebox.showerror("Error", "No single ROI to save.")
                return
            x, y, w, h = self.single_roi
            payload = {"mode": "single", "x": x, "y": y, "w": w, "h": h}

        p.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        self.log(f"ROI saved to {p}")

    def apply_roi_text(self):
        txt = self.roi_text_var.get().strip()
        try:
            r = parse_manual_roi(txt)
        except Exception as exc:
            messagebox.showerror("ROI Error", str(exc))
            return
        self.single_roi = (r.x, r.y, r.w, r.h)
        self.roi_mode_var.set("single")
        self.log(f"Single ROI set: {self.single_roi}")
        self.refresh_preview()

    def pick_roi_from_sample(self):
        if self.current_image_bgr is None:
            messagebox.showerror("No image", "Load a sample row first.")
            return
        roi = self._safe_select_roi(
            "Select ROI (Enter/Space confirm, c cancel)",
            self.current_image_bgr,
        )
        if roi is None:
            return
        x, y, w, h = roi
        self.single_roi = (int(x), int(y), int(w), int(h))
        self.roi_text_var.set(f"{x},{y},{w},{h}")
        self.roi_mode_var.set("single")
        self.log(f"Single ROI selected: {self.single_roi}")
        self.refresh_preview()

    def pick_grid_start_box(self):
        if self.current_image_bgr is None:
            messagebox.showerror("No image", "Load a sample row first.")
            return
        digit_count = int(self.digit_count_var.get())
        gap = int(self.grid_gap_var.get())

        roi = self._safe_select_roi(
            "Pick FIRST digit box (grid will auto-generate)",
            self.current_image_bgr,
        )
        if roi is None:
            return
        x, y, w, h = roi

        self.grid_base_boxes = []
        for i in range(digit_count):
            xi = int(x + i * (w + gap))
            self.grid_base_boxes.append((xi, int(y), int(w), int(h)))
        self.grid_offsets = [[0, 0, 0, 0] for _ in range(digit_count)]
        self.roi_mode_var.set("grid")
        self.log(f"Grid generated with {digit_count} boxes from first box.")
        self.refresh_preview()

    def open_grid_tuner(self):
        if self.roi_mode_var.get() != "grid":
            messagebox.showinfo("Grid", "Switch ROI mode to Grid first.")
            return
        if not self.grid_base_boxes:
            messagebox.showerror("Grid", "No grid available. Use 'Pick Grid Start' first.")
            return

        top = tk.Toplevel(self)
        top.title("Fine-tune Grid Offsets")
        top.geometry("760x520")

        container = ttk.Frame(top, padding=8)
        container.pack(fill="both", expand=True)

        header = ttk.Frame(container)
        header.pack(fill="x")
        ttk.Label(header, text="Digit", width=8).grid(row=0, column=0)
        ttk.Label(header, text="dx", width=10).grid(row=0, column=1)
        ttk.Label(header, text="dy", width=10).grid(row=0, column=2)
        ttk.Label(header, text="dw", width=10).grid(row=0, column=3)
        ttk.Label(header, text="dh", width=10).grid(row=0, column=4)

        canvas = tk.Canvas(container)
        scroll = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        body = ttk.Frame(canvas)
        body.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

        vars_list: List[Tuple[tk.IntVar, tk.IntVar, tk.IntVar, tk.IntVar]] = []
        for i in range(len(self.grid_base_boxes)):
            if i >= len(self.grid_offsets):
                self.grid_offsets.append([0, 0, 0, 0])
            dx, dy, dw, dh = self.grid_offsets[i]
            vdx, vdy, vdw, vdh = tk.IntVar(value=dx), tk.IntVar(value=dy), tk.IntVar(value=dw), tk.IntVar(value=dh)
            vars_list.append((vdx, vdy, vdw, vdh))

            ttk.Label(body, text=str(i), width=8).grid(row=i, column=0, padx=2, pady=2)
            ttk.Spinbox(body, from_=-200, to=200, textvariable=vdx, width=8).grid(row=i, column=1, padx=2, pady=2)
            ttk.Spinbox(body, from_=-200, to=200, textvariable=vdy, width=8).grid(row=i, column=2, padx=2, pady=2)
            ttk.Spinbox(body, from_=-200, to=200, textvariable=vdw, width=8).grid(row=i, column=3, padx=2, pady=2)
            ttk.Spinbox(body, from_=-200, to=200, textvariable=vdh, width=8).grid(row=i, column=4, padx=2, pady=2)

        btns = ttk.Frame(top, padding=8)
        btns.pack(fill="x")

        def apply_offsets():
            for i, (vdx, vdy, vdw, vdh) in enumerate(vars_list):
                self.grid_offsets[i] = [int(vdx.get()), int(vdy.get()), int(vdw.get()), int(vdh.get())]
            self.log("Grid offsets updated.")
            self.refresh_preview()

        ttk.Button(btns, text="Apply", command=apply_offsets).pack(side="left", padx=4)
        ttk.Button(btns, text="Apply & Close", command=lambda: (apply_offsets(), top.destroy())).pack(side="left", padx=4)
        ttk.Button(btns, text="Close", command=top.destroy).pack(side="right", padx=4)

    def shift_all(self, dx: int, dy: int):
        if self.roi_mode_var.get() == "grid":
            if not self.grid_base_boxes:
                return
            scope = (self.shift_scope_var.get() or "global").strip().lower()
            if scope == "selected":
                idx = self._selected_box_idx()
                while len(self.grid_offsets) < len(self.grid_base_boxes):
                    self.grid_offsets.append([0, 0, 0, 0])
                self.grid_offsets[idx][0] += int(dx)
                self.grid_offsets[idx][1] += int(dy)
                self.log(f"Shifted box {idx} by ({dx}, {dy})")
            else:
                moved = []
                for x, y, w, h in self.grid_base_boxes:
                    moved.append((x + dx, y + dy, w, h))
                self.grid_base_boxes = moved
                self.log(f"Shifted grid by ({dx}, {dy})")
        else:
            if self.single_roi is None:
                return
            x, y, w, h = self.single_roi
            self.single_roi = (x + dx, y + dy, w, h)
            self.roi_text_var.set(f"{self.single_roi[0]},{self.single_roi[1]},{w},{h}")
            self.log(f"Shifted ROI by ({dx}, {dy})")
        self.refresh_preview()

    # ---------- Data load ----------
    def find_next_pending(self):
        try:
            wb, ws = self._open_sheet()
            url_col, out_col = self._resolve_cols(ws)
            start = max(2, int(self.sample_row_var.get()))
            for row in range(start, ws.max_row + 1):
                url = ws.cell(row=row, column=url_col).value
                out = ws.cell(row=row, column=out_col).value
                if url and (out is None or not str(out).strip()):
                    self.sample_row_var.set(row)
                    self.log(f"Next pending row: {row}")
                    break
            wb.close()
        except Exception as exc:
            messagebox.showerror("Error", str(exc))

    def load_model(self):
        try:
            framework = self.framework_var.get().strip()
            if framework in ("keras", "torch"):
                model_path = self.model_path_var.get().strip()
                # UX fallback: if user filled Groq fields but forgot to switch framework.
                if not model_path and self._has_groq_credentials():
                    framework = "groq"
                    self.framework_var.set("groq")
                    self._on_framework_change()
                    self.log("Framework auto-switched to groq (local model path is empty).")

            if framework in ("keras", "torch"):
                model_path = self.model_path_var.get().strip()
                if not model_path:
                    raise RuntimeError("Select a model path.")
                config_path = self.model_config_var.get().strip() or None
                cfg = load_model_config(config_path, model_path)

                if cfg.get("framework") and str(cfg["framework"]) != framework:
                    raise RuntimeError(f"Framework mismatch. UI={framework}, config={cfg['framework']}")

                if "image_size" in cfg:
                    self.image_size_var.set(int(cfg["image_size"]))
                if "invert" in cfg:
                    self.invert_var.set(bool(cfg["invert"]))
                if "digit_count" in cfg:
                    self.digit_count_var.set(int(cfg["digit_count"]))
                if "use_morphology_cleaning" in cfg:
                    self.morph_clean_var.set(bool(cfg["use_morphology_cleaning"]))

                self.runner = get_model_runner(framework, model_path)
                self._auto_match_image_size_to_model()
                loaded_label = model_path
            elif framework == "groq":
                api_key = self.api_key_var.get().strip()
                api_model = self.api_model_var.get().strip()
                api_endpoint = self.api_endpoint_var.get().strip() or DEFAULT_GROQ_ENDPOINT
                if not api_key:
                    raise RuntimeError("Enter API key for Groq.")
                if not api_model:
                    raise RuntimeError("Enter API model for Groq.")
                self.runner = get_model_runner(
                    framework="groq",
                    api_key=api_key,
                    api_model=api_model,
                    api_endpoint=api_endpoint,
                    timeout=int(self.timeout_var.get()),
                )
                self.model_expected_hw = None
                loaded_label = f"groq:{api_model}"
            else:
                raise RuntimeError(f"Unsupported framework: {framework}")

            self.model_ready_var.set("Model Ready")
            self.pred_var.set("Model Ready")
            self.load_model_btn.configure(style="Ready.TButton")
            self.log(f"Model loaded: {loaded_label}")
            self.refresh_preview()
        except Exception as exc:
            self.model_ready_var.set("Model not loaded")
            self.pred_var.set("Prediction: model not loaded")
            messagebox.showerror("Model Error", self._short_error(exc))

    def load_sample_row(self):
        try:
            wb, ws = self._open_sheet()
            url_col, _ = self._resolve_cols(ws)
            row = int(self.sample_row_var.get())
            source = ws.cell(row=row, column=url_col).value
            wb.close()
            if not source or not str(source).strip():
                raise RuntimeError(f"Row {row} has empty source in URL column.")

            self.current_row = row
            self.current_source = str(source).strip()
            session = requests.Session()
            try:
                self.current_image_bgr = decode_image_from_source(
                    source=self.current_source,
                    timeout=int(self.timeout_var.get()),
                    retries=int(self.retries_var.get()),
                    retry_backoff=float(self.backoff_var.get()),
                    session=session,
                )
            finally:
                session.close()

            self.log(f"Loaded sample row {row}")
            self.refresh_preview()
        except Exception as exc:
            messagebox.showerror("Sample Error", str(exc))

    # ---------- Preview ----------
    def _draw_boxes_overlay(
        self,
        image_bgr: np.ndarray,
        boxes: Sequence[Tuple[int, int, int, int]],
        low_conf_idxs: Optional[Sequence[int]] = None,
        selected_idx: Optional[int] = None,
    ) -> np.ndarray:
        out = image_bgr.copy()
        low_set = set(low_conf_idxs or [])
        for i, (x, y, w, h) in enumerate(boxes):
            color = (0, 255, 0)
            if i in low_set:
                color = (0, 0, 255)
            if selected_idx is not None and i == selected_idx:
                color = (255, 255, 0)
            cv2.rectangle(out, (x, y), (x + w, y + h), color, 2)
            cv2.putText(out, str(i), (x + 1, max(10, y - 4)), cv2.FONT_HERSHEY_SIMPLEX, 0.45, color, 1, cv2.LINE_AA)
        return out

    def on_digit_click(self, idx: int):
        self.selected_digit_idx = idx
        self.selected_box_var.set(idx)
        self.refresh_preview()

    def refresh_preview(self):
        if self.current_image_bgr is None:
            return

        try:
            boxes = self._digit_boxes_for_image(self.current_image_bgr)
            if not boxes:
                full = self._photo_for_widget(self.current_image_bgr, self.full_img_label, fallback=(760, 520))
                self.full_img_label.configure(image=full, text="")
                self.photo_refs["full"] = full
                self.roi_img_label.configure(text="Pick/apply ROI or grid first", image="")
                self.pred_var.set("Prediction: model not loaded" if self.runner is None else "Model Ready")
                return

            digit_cells = self._crop_from_boxes(self.current_image_bgr, boxes)
            s = self._settings()

            batch_list: List[np.ndarray] = []
            preview_list: List[np.ndarray] = []
            for cell in digit_cells:
                arr, preview = preprocess_digit_custom(cell, s)
                batch_list.append(arr)
                preview_list.append(preview)

            preds: List[int] = []
            confs: List[float] = []
            pred_text = "Prediction: model not loaded"
            if self.runner is not None and batch_list:
                batch = np.stack(batch_list, axis=0)
                preds, confs = self._predict_with_confidence(batch)
                if preds:
                    code = "".join(str(int(p)) for p in preds)
                    pred_text = f"Model Ready | Prediction: {code}"
                else:
                    pred_text = "Model Ready"
            else:
                if self.runner is not None:
                    pred_text = "Model Ready"

            self.last_digit_previews = preview_list
            self.last_preds = preds
            self.last_confs = confs
            self.pred_var.set(pred_text)

            threshold = float(self.conf_threshold_var.get())
            low_idxs = [i for i, c in enumerate(confs) if c < threshold]

            overlay = self._draw_boxes_overlay(
                self.current_image_bgr,
                boxes,
                low_conf_idxs=low_idxs,
                selected_idx=self.selected_digit_idx,
            )

            roi_preview = self.current_image_bgr
            if self.roi_mode_var.get() == "single" and self.single_roi is not None:
                x, y, w, h = self.single_roi
                roi_obj = parse_manual_roi(f"{x},{y},{w},{h}")
                roi_preview = crop_roi(self.current_image_bgr, roi_obj)
            elif boxes:
                x0 = min(x for x, _, _, _ in boxes)
                y0 = min(y for _, y, _, _ in boxes)
                x1 = max(x + w for x, _, w, _ in boxes)
                y1 = max(y + h for _, y, _, h in boxes)
                roi_preview = self.current_image_bgr[y0:y1, x0:x1]

            full = self._photo_for_widget(overlay, self.full_img_label, fallback=(760, 520))
            roi_p = self._photo_for_widget(roi_preview, self.roi_img_label, fallback=(760, 520))
            self.full_img_label.configure(image=full, text="")
            self.roi_img_label.configure(image=roi_p, text="")
            self.photo_refs["full"] = full
            self.photo_refs["roi"] = roi_p

            for child in self.digit_frame.winfo_children():
                child.destroy()
            self.digit_photo_refs.clear()

            for i, d in enumerate(preview_list):
                img = self._to_photo(d, max_size=(90, 90))
                self.digit_photo_refs.append(img)

                border_color = "#b71c1c" if i < len(confs) and confs[i] < threshold else "#444444"
                frame = tk.Frame(self.digit_frame, highlightthickness=2, highlightbackground=border_color)
                frame.grid(row=0, column=i, padx=2, pady=2)
                lbl = tk.Label(frame, image=img, cursor="hand2")
                lbl.pack()
                lbl.bind("<Button-1>", lambda _e, idx=i: self.on_digit_click(idx))

                caption = str(i)
                if i < len(preds):
                    caption += f" | {preds[i]}"
                if i < len(confs):
                    caption += f" ({confs[i]:.2f})"
                ttk.Label(self.digit_frame, text=caption).grid(row=1, column=i)

        except Exception as exc:
            self.pred_var.set("Preview failed (check settings)")
            self.log(f"Preview error: {self._short_error(exc)}")

    def preview_morphology_preprocess(self):
        if self.current_image_bgr is None:
            messagebox.showerror("No image", "Load a sample row first.")
            return
        try:
            boxes = self._digit_boxes_for_image(self.current_image_bgr)
            if not boxes:
                raise RuntimeError("Pick/apply ROI or grid first.")
            digit_cells = self._crop_from_boxes(self.current_image_bgr, boxes)
            s = self._settings()
            # Explicitly force morphology cleaner for this visualization action only.
            s.use_morphology_cleaning = True
            previews: List[np.ndarray] = []
            for cell in digit_cells:
                _, preview = preprocess_digit_custom(cell, s)
                previews.append(preview)
            strip = self._build_digit_strip(previews, tile_size=64, gap=6)
            roi_p = self._photo_for_widget(strip, self.roi_img_label, fallback=(760, 520))
            self.roi_img_label.configure(image=roi_p, text="")
            self.photo_refs["roi"] = roi_p
            self.log("Displayed morphology preprocessing preview (no inference).")
        except Exception as exc:
            messagebox.showerror("Morphology Preview Error", self._short_error(exc))

    # ---------- Batch ----------
    def run_batch_async(self):
        t = threading.Thread(target=self.run_batch, daemon=True)
        t.start()

    def _request_manual_review_dialog(
        self,
        req: BatchReviewRequest,
        previews: Sequence[np.ndarray],
        done_event: threading.Event,
        result_holder: Dict[str, str],
    ):
        top = tk.Toplevel(self)
        top.title(f"Manual Review - Row {req.row}")
        top.geometry("1200x360")
        top.transient(self)
        top.grab_set()

        wrapper = ttk.Frame(top, padding=8)
        wrapper.pack(fill="both", expand=True)

        suspect_txt = ", ".join(f"{i}:{req.confidences[i]:.2f}" for i in req.suspect_idxs)
        ttk.Label(
            wrapper,
            text=f"Low-confidence digits detected at [{suspect_txt}]. Edit code and continue.",
        ).pack(anchor="w")

        row = ttk.Frame(wrapper)
        row.pack(fill="x", pady=(8, 6))

        photos: List[ImageTk.PhotoImage] = []
        thr = float(self.conf_threshold_var.get())
        for i, d in enumerate(previews):
            img = self._to_photo(d, max_size=(70, 70))
            photos.append(img)
            border = "#b71c1c" if i in req.suspect_idxs else "#444444"
            f = tk.Frame(row, highlightthickness=2, highlightbackground=border)
            f.pack(side="left", padx=2)
            tk.Label(f, image=img).pack()
            txt = f"{i}:{req.predicted[i]}\n{req.confidences[i]:.2f}" if i < len(req.confidences) else f"{i}:{req.predicted[i]}"
            tk.Label(f, text=txt).pack()

        entry_var = tk.StringVar(value=req.predicted)
        form = ttk.Frame(wrapper)
        form.pack(fill="x", pady=(8, 6))
        ttk.Label(form, text="Correct code:").pack(side="left")
        entry = ttk.Entry(form, textvariable=entry_var, width=max(20, len(req.predicted) + 4))
        entry.pack(side="left", padx=(6, 0))
        entry.focus_set()

        btns = ttk.Frame(wrapper)
        btns.pack(fill="x", pady=(8, 0))

        def finish(ok: bool):
            if ok:
                code = entry_var.get().strip()
                if len(code) != len(req.predicted) or not code.isdigit():
                    messagebox.showerror("Invalid", f"Enter exactly {len(req.predicted)} digits.")
                    return
                result_holder["code"] = code
            else:
                result_holder["code"] = req.predicted
            top.destroy()
            done_event.set()

        ttk.Button(btns, text="Use Edited", command=lambda: finish(True)).pack(side="left", padx=4)
        ttk.Button(btns, text="Use Predicted", command=lambda: finish(False)).pack(side="left", padx=4)

        def on_close():
            result_holder["code"] = req.predicted
            top.destroy()
            done_event.set()

        top.protocol("WM_DELETE_WINDOW", on_close)

    def _manual_review_if_needed(
        self,
        row: int,
        code: str,
        confs: List[float],
        previews: Sequence[np.ndarray],
    ) -> str:
        if not self.manual_review_var.get():
            return code

        thr = float(self.conf_threshold_var.get())
        suspect = [i for i, c in enumerate(confs) if c < thr]
        if not suspect:
            return code

        req = BatchReviewRequest(row=row, predicted=code, confidences=confs, suspect_idxs=suspect)
        done = threading.Event()
        result: Dict[str, str] = {}
        self.after(0, lambda: self._request_manual_review_dialog(req, previews, done, result))
        done.wait()
        return result.get("code", code)

    def run_batch(self):
        try:
            if self.runner is None:
                raise RuntimeError("Load model first.")

            excel_path = Path(self.excel_var.get().strip())
            if not excel_path.exists():
                raise RuntimeError(f"Excel not found: {excel_path}")

            out_path = self.output_excel_var.get().strip()
            output_excel = Path(out_path) if out_path else excel_path.with_name(f"{excel_path.stem}_with_udise.xlsx")

            wb = load_workbook(excel_path)
            ws = resolve_sheet(wb, self.sheet_var.get().strip() or None)
            url_col, out_col = self._resolve_cols(ws)

            first_row = max(2, int(self.start_row_var.get()))
            max_rows = int(self.max_rows_var.get())
            last_row = ws.max_row if max_rows <= 0 else min(ws.max_row, first_row + max_rows - 1)
            if first_row > last_row:
                raise RuntimeError("No rows to process with current row limits.")

            digit_count = int(self.digit_count_var.get())
            s = self._settings()
            overwrite = bool(self.overwrite_var.get())
            save_every = int(self.save_every_var.get())

            session = requests.Session()
            processed = 0
            success = 0
            skipped = 0
            autosave_counter = 0
            failures: List[Tuple[int, str, str]] = []

            self.log(
                f"Batch started rows {first_row}..{last_row}, mode={self.roi_mode_var.get()}, digit_count={digit_count}, "
                f"manual_review={self.manual_review_var.get()}, remove_vlines={s.remove_vertical_lines}"
            )

            try:
                for row in range(first_row, last_row + 1):
                    existing = ws.cell(row=row, column=out_col).value
                    if (existing is not None) and str(existing).strip() and not overwrite:
                        skipped += 1
                        continue

                    source_val = ws.cell(row=row, column=url_col).value
                    if not source_val or not str(source_val).strip():
                        skipped += 1
                        continue

                    source = str(source_val).strip()
                    processed += 1
                    try:
                        image_bgr = decode_image_from_source(
                            source=source,
                            timeout=int(self.timeout_var.get()),
                            retries=int(self.retries_var.get()),
                            retry_backoff=float(self.backoff_var.get()),
                            session=session,
                        )

                        boxes = self._digit_boxes_for_image(image_bgr)
                        if len(boxes) != digit_count:
                            self.log(
                                f"row {row}: box count {len(boxes)} != {digit_count}, auto-normalizing to {digit_count}"
                            )
                        boxes = self._normalize_boxes_count(boxes, digit_count, image_bgr.shape)

                        digit_cells = self._crop_from_boxes(image_bgr, boxes)
                        batch_list = [preprocess_digit_custom(cell, s)[0] for cell in digit_cells]
                        preview_list = [preprocess_digit_custom(cell, s)[1] for cell in digit_cells]
                        batch = np.stack(batch_list, axis=0)

                        preds, confs = self._predict_with_confidence(batch)
                        code = self._coerce_digit_code(preds, digit_count)
                        if len(preds) != digit_count:
                            self.log(
                                f"row {row}: model returned {len(preds)} digits, coerced to {digit_count}"
                            )
                        code = self._manual_review_if_needed(row, code, confs, preview_list)

                        ws.cell(row=row, column=out_col).value = code
                        success += 1
                        autosave_counter += 1
                        self.log(f"row {row}: {code}")

                        if save_every > 0 and autosave_counter >= save_every:
                            wb.save(output_excel)
                            autosave_counter = 0
                            self.log(f"[autosave] {output_excel}")
                    except Exception as exc:
                        err = self._short_error(exc)
                        failures.append((row, source, err))
                        # Hard requirement: always write a fixed 11-digit code in output column.
                        fallback_code = "0" * digit_count
                        ws.cell(row=row, column=out_col).value = fallback_code
                        self.log(f"row {row}: fallback written due to error -> {fallback_code}")
                        self.log(f"row {row}: ERROR: {err}")
            finally:
                session.close()

            wb.save(output_excel)
            write_failure_log("udise_failures_gui.csv", failures)
            self.log(
                f"Batch done. processed={processed}, skipped={skipped}, success={success}, failed={len(failures)}, output={output_excel}"
            )
            if failures:
                self.log("Failure log: udise_failures_gui.csv")
            messagebox.showinfo(
                "Done",
                f"Batch completed.\nSuccess: {success}\nFailed: {len(failures)}\nOutput: {output_excel}",
            )
        except Exception as exc:
            err = self._short_error(exc)
            self.log(f"Batch error: {err}")
            messagebox.showerror("Batch Error", err)


def main():
    app = UDISEOCRApp()
    app.mainloop()


if __name__ == "__main__":
    main()
