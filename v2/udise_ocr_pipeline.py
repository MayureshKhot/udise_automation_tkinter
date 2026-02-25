#!/usr/bin/env python3
"""Extract 11-digit UDISE codes from image sources listed in an Excel sheet.

Pipeline:
1. Read image source (URL or local path) from Excel.
2. Load image with retries.
3. Use one-time interactive ROI selection (or saved/manual ROI).
4. Crop UDISE region and split into digit cells.
5. Preprocess each digit image.
6. Run CNN digit model (Keras or PyTorch).
7. Write result back to Excel.
"""

from __future__ import annotations

import argparse
import base64
import csv
import json
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple
from urllib.parse import unquote, urlparse

import cv2
import numpy as np
import requests
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
from preprocessing.morphology_grid_cleaner import remove_grid_lines

DEFAULT_URL_COLUMN = "H"
DEFAULT_OUTPUT_COLUMN = "E"
DEFAULT_ROI_FILE = "udise_roi.json"
DEFAULT_FAILURE_LOG = "udise_failures.csv"
EXPECTED_DIGITS = 11
DEFAULT_GROQ_ENDPOINT = "https://api.groq.com/openai/v1/chat/completions"

URL_HEADER_CANDIDATES = [
    "original image path",
    "image path on server",
    "image url",
    "url",
]
OUTPUT_HEADER_CANDIDATES = [
    "udise found",
    "udise",
    "udise code",
]
MODEL_CONFIG_KEYS = {
    "framework",
    "image_size",
    "invert",
    "digit_count",
    "use_morphology_cleaning",
    "architecture",
    "dataset",
    "test_accuracy",
    "test_loss",
    "epochs_ran",
}


@dataclass
class ROI:
    x: int
    y: int
    w: int
    h: int

    def as_tuple(self) -> Tuple[int, int, int, int]:
        return (self.x, self.y, self.w, self.h)


class ModelRunner:
    def predict(self, digit_batch: np.ndarray) -> List[int]:
        raise NotImplementedError

    def predict_with_confidence(self, digit_batch: np.ndarray) -> Tuple[List[int], List[float]]:
        preds = self.predict(digit_batch)
        return preds, [1.0 for _ in preds]


class KerasRunner(ModelRunner):
    def __init__(self, model_path: str):
        try:
            import tensorflow as tf
        except Exception as exc:  # pragma: no cover
            raise RuntimeError(
                "TensorFlow is not installed. Install `tensorflow` to use --framework keras."
            ) from exc

        self.model = tf.keras.models.load_model(model_path, compile=False)
        shape = self.model.input_shape
        if isinstance(shape, list):
            shape = shape[0]
        self.input_shape = tuple(shape)

    def _prepare_input(self, digit_batch: np.ndarray) -> np.ndarray:
        shape = self.input_shape
        # Common supported inputs:
        # [N,H,W], [N,H,W,1], [N,H,W,3], [N,1,H,W]
        if len(shape) == 4:
            if shape[-1] == 1:
                return digit_batch
            if shape[-1] == 3:
                return np.repeat(digit_batch, 3, axis=-1)
            if shape[1] == 1:
                return np.transpose(digit_batch, (0, 3, 1, 2))
        if len(shape) == 3:
            return np.squeeze(digit_batch, axis=-1)
        raise RuntimeError(
            f"Unsupported Keras input shape {shape}. Expected [N,H,W], [N,H,W,1], [N,H,W,3] or [N,1,H,W]."
        )

    def predict(self, digit_batch: np.ndarray) -> List[int]:
        x = self._prepare_input(digit_batch)
        logits = self.model.predict(x, verbose=0)
        logits = np.asarray(logits)
        if logits.ndim != 2 or logits.shape[1] < 10:
            raise RuntimeError(
                f"Unexpected Keras output shape {logits.shape}; expected [N, >=10]."
            )
        return logits.argmax(axis=1).astype(int).tolist()

    def predict_with_confidence(self, digit_batch: np.ndarray) -> Tuple[List[int], List[float]]:
        x = self._prepare_input(digit_batch)
        logits = np.asarray(self.model.predict(x, verbose=0))
        if logits.ndim != 2 or logits.shape[1] < 10:
            raise RuntimeError(
                f"Unexpected Keras output shape {logits.shape}; expected [N, >=10]."
            )
        preds = logits.argmax(axis=1).astype(int).tolist()
        confs = logits.max(axis=1).astype(float).tolist()
        return preds, confs


class TorchRunner(ModelRunner):
    def __init__(self, model_path: str):
        try:
            import torch
        except Exception as exc:  # pragma: no cover
            raise RuntimeError(
                "PyTorch is not installed. Install `torch` to use --framework torch."
            ) from exc

        self.torch = torch

        model = None
        try:
            model = torch.jit.load(model_path, map_location="cpu")
        except Exception:
            pass

        if model is None:
            obj = torch.load(model_path, map_location="cpu")
            if hasattr(obj, "eval"):
                model = obj
            else:
                raise RuntimeError(
                    "Unable to load torch model. Provide a TorchScript model or a full nn.Module object."
                )

        self.model = model.eval()

    def predict(self, digit_batch: np.ndarray) -> List[int]:
        # NHWC -> NCHW for torch
        x = np.transpose(digit_batch, (0, 3, 1, 2)).astype(np.float32)
        tensor = self.torch.from_numpy(x)
        with self.torch.no_grad():
            out = self.model(tensor)
        logits = out.detach().cpu().numpy()
        if logits.ndim != 2 or logits.shape[1] < 10:
            raise RuntimeError(
                f"Unexpected Torch output shape {logits.shape}; expected [N, >=10]."
            )
        return logits.argmax(axis=1).astype(int).tolist()

    def predict_with_confidence(self, digit_batch: np.ndarray) -> Tuple[List[int], List[float]]:
        x = np.transpose(digit_batch, (0, 3, 1, 2)).astype(np.float32)
        tensor = self.torch.from_numpy(x)
        with self.torch.no_grad():
            out = self.model(tensor)
        logits = out.detach().cpu().numpy()
        if logits.ndim != 2 or logits.shape[1] < 10:
            raise RuntimeError(
                f"Unexpected Torch output shape {logits.shape}; expected [N, >=10]."
            )
        exp = np.exp(logits - logits.max(axis=1, keepdims=True))
        probs = exp / exp.sum(axis=1, keepdims=True)
        preds = probs.argmax(axis=1).astype(int).tolist()
        confs = probs.max(axis=1).astype(float).tolist()
        return preds, confs


class GroqVisionRunner(ModelRunner):
    def __init__(
        self,
        api_key: str,
        model: str,
        endpoint: str = DEFAULT_GROQ_ENDPOINT,
        timeout: int = 20,
    ):
        if not api_key.strip():
            raise RuntimeError("Groq API key is required.")
        if not model.strip():
            raise RuntimeError("Groq model name is required.")
        self.api_key = api_key.strip()
        self.model = model.strip()
        self.endpoint = endpoint.strip() or DEFAULT_GROQ_ENDPOINT
        self.timeout = max(1, int(timeout))
        self.session = requests.Session()
        self.system_prompt = (
            "You are an OCR helper. Return exactly one character digit from 0 to 9. "
            "No words, no punctuation, no explanations."
        )
        self.user_prompt = (
            "Identify the single handwritten/printed digit in this image tile and return only one digit 0-9."
        )
        self._validate_model_name_for_vision()

    def _validate_model_name_for_vision(self) -> None:
        # We send image_url payloads, so a vision-capable model is required.
        model_l = self.model.lower()
        vision_hints = (
            "vision",
            "llava",
            "gpt-4o",
            "qwen2.5-vl",
            "gemma-vision",
            "llama-4-scout",
            "llama-4-maverick",
        )
        if not any(h in model_l for h in vision_hints):
            raise RuntimeError(
                f"Groq model '{self.model}' appears non-vision. "
                "Choose a vision-capable model (for example: "
                "meta-llama/llama-4-scout-17b-16e-instruct)."
            )

    def _encode_digit_image(self, digit_arr: np.ndarray) -> str:
        arr = digit_arr
        if arr.ndim == 3 and arr.shape[-1] == 1:
            arr = np.squeeze(arr, axis=-1)
        if arr.ndim != 2:
            raise RuntimeError(f"Expected digit image shape [H,W] or [H,W,1], got {arr.shape}")

        arr_u8 = np.clip(arr * 255.0, 0, 255).astype(np.uint8)
        upscale = cv2.resize(arr_u8, (112, 112), interpolation=cv2.INTER_NEAREST)
        rgb = cv2.cvtColor(upscale, cv2.COLOR_GRAY2RGB)
        ok, encoded = cv2.imencode(".png", rgb)
        if not ok:
            raise RuntimeError("Failed to encode digit image for Groq request.")
        return base64.b64encode(encoded.tobytes()).decode("ascii")

    def _parse_digit(self, text: str) -> Tuple[int, float]:
        content = (text or "").strip()
        if re.fullmatch(r"[0-9]", content):
            return int(content), 1.0
        m = re.search(r"\b([0-9])\b", content)
        if m:
            return int(m.group(1)), 0.7
        return 0, 0.0

    def _extract_content_text(self, payload: Dict[str, object]) -> str:
        choices = payload.get("choices")
        if not isinstance(choices, list) or not choices:
            return ""
        first = choices[0]
        if not isinstance(first, dict):
            return ""
        message = first.get("message")
        if not isinstance(message, dict):
            return ""
        content = message.get("content")
        if isinstance(content, str):
            return content
        if isinstance(content, list):
            text_parts: List[str] = []
            for item in content:
                if isinstance(item, dict) and item.get("type") == "text":
                    text_val = item.get("text")
                    if isinstance(text_val, str):
                        text_parts.append(text_val)
            return " ".join(text_parts).strip()
        return ""

    def predict_with_confidence(self, digit_batch: np.ndarray) -> Tuple[List[int], List[float]]:
        preds: List[int] = []
        confs: List[float] = []

        for i in range(int(digit_batch.shape[0])):
            b64_image = self._encode_digit_image(digit_batch[i])
            payload = {
                "model": self.model,
                "temperature": 0,
                "messages": [
                    {"role": "system", "content": self.system_prompt},
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": self.user_prompt},
                            {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64_image}"}},
                        ],
                    },
                ],
            }
            headers = {
                "Authorization": f"Bearer {self.api_key}",
                "Content-Type": "application/json",
            }
            try:
                resp = self.session.post(
                    self.endpoint,
                    headers=headers,
                    json=payload,
                    timeout=self.timeout,
                )
                if resp.status_code >= 400:
                    body = resp.text.strip()
                    if len(body) > 500:
                        body = body[:500] + "..."
                    raise RuntimeError(
                        f"HTTP {resp.status_code} from Groq. "
                        f"Model='{self.model}'. Response: {body}"
                    )
                content = self._extract_content_text(resp.json())
                digit, conf = self._parse_digit(content)
            except Exception as exc:
                raise RuntimeError(f"Groq request failed at digit index {i}: {exc}") from exc

            preds.append(digit)
            confs.append(conf)
        return preds, confs

    def predict(self, digit_batch: np.ndarray) -> List[int]:
        preds, _ = self.predict_with_confidence(digit_batch)
        return preds


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read UDISE from image sources in Excel using a CNN digit model."
    )
    parser.add_argument("--excel", required=True, help="Input Excel file path (.xlsx)")
    parser.add_argument(
        "--output-excel",
        default=None,
        help="Output Excel path. Default: <input_stem>_with_udise.xlsx",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Sheet name or 1-based index. Default: first sheet.",
    )

    parser.add_argument(
        "--url-column",
        default=None,
        help=(
            "URL/source column letter or header. If omitted, auto-detects headers and falls back to H."
        ),
    )
    parser.add_argument(
        "--output-column",
        default=None,
        help=(
            "Output column letter or header. If omitted, auto-detects headers and falls back to E."
        ),
    )

    parser.add_argument(
        "--framework",
        choices=["keras", "torch", "groq"],
        required=True,
        help="Model framework",
    )
    parser.add_argument("--model-path", default="", help="Path to local CNN model file")
    parser.add_argument(
        "--model-config",
        default=None,
        help=(
            "Optional JSON config path generated during training. "
            "If omitted, tries <model_path>.json if present."
        ),
    )
    parser.add_argument("--api-key", default="", help="External API key (used for --framework groq)")
    parser.add_argument("--api-model", default="", help="External model name (used for --framework groq)")
    parser.add_argument(
        "--api-endpoint",
        default=DEFAULT_GROQ_ENDPOINT,
        help=f"External chat-completions endpoint (default: {DEFAULT_GROQ_ENDPOINT})",
    )

    parser.add_argument(
        "--roi-file",
        default=DEFAULT_ROI_FILE,
        help=f"JSON path to save/load ROI (default: {DEFAULT_ROI_FILE})",
    )
    parser.add_argument(
        "--pick-roi",
        action="store_true",
        help="Force interactive ROI selection on first valid image",
    )
    parser.add_argument(
        "--roi",
        default=None,
        help="Manual ROI as 'x,y,w,h'. If given, picker is skipped.",
    )

    parser.add_argument(
        "--digit-count",
        type=int,
        default=EXPECTED_DIGITS,
        help=f"Number of digits in code (default: {EXPECTED_DIGITS})",
    )
    parser.add_argument(
        "--invert",
        action=argparse.BooleanOptionalAction,
        default=None,
        help="Invert thresholded digit image before model input. Default comes from model config if available.",
    )
    parser.add_argument(
        "--image-size",
        type=int,
        default=None,
        help="Digit image size for model input. Default comes from model config or 28.",
    )
    parser.add_argument(
        "--use-morphology-cleaning",
        action=argparse.BooleanOptionalAction,
        default=None,
        help=(
            "Apply morphology-based vertical gridline cleaning before model input. "
            "Default comes from model config if available."
        ),
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=20,
        help="HTTP timeout in seconds (default: 20)",
    )
    parser.add_argument(
        "--retries",
        type=int,
        default=2,
        help="Download retries per image (default: 2)",
    )
    parser.add_argument(
        "--retry-backoff",
        type=float,
        default=1.5,
        help="Retry backoff multiplier in seconds (default: 1.5)",
    )

    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="First data row to process (default: 2)",
    )
    parser.add_argument(
        "--max-rows",
        type=int,
        default=0,
        help="Process at most N rows from --start-row (0 = all)",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite already-filled output cells",
    )
    parser.add_argument(
        "--save-every",
        type=int,
        default=50,
        help="Autosave workbook every N successful rows (default: 50)",
    )

    parser.add_argument(
        "--debug-dir",
        default=None,
        help="Optional folder to dump debug crops and digits for first processed row",
    )
    parser.add_argument(
        "--failure-log",
        default=DEFAULT_FAILURE_LOG,
        help=f"CSV path for failed rows (default: {DEFAULT_FAILURE_LOG})",
    )
    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Reduce per-row logging",
    )
    return parser.parse_args()


def get_model_runner(
    framework: str,
    model_path: str = "",
    api_key: str = "",
    api_model: str = "",
    api_endpoint: str = DEFAULT_GROQ_ENDPOINT,
    timeout: int = 20,
) -> ModelRunner:
    if framework == "keras":
        if not model_path:
            raise RuntimeError("--model-path is required for keras.")
        return KerasRunner(model_path)
    if framework == "torch":
        if not model_path:
            raise RuntimeError("--model-path is required for torch.")
        return TorchRunner(model_path)
    if framework == "groq":
        return GroqVisionRunner(
            api_key=api_key,
            model=api_model,
            endpoint=api_endpoint,
            timeout=timeout,
        )
    raise ValueError(f"Unsupported framework: {framework}")


def guess_model_config_path(model_path: str) -> Optional[Path]:
    p = Path(model_path)
    candidate = p.with_suffix(p.suffix + ".json")
    if candidate.exists():
        return candidate
    return None


def load_model_config(path: Optional[str], model_path: str = "") -> Dict[str, object]:
    config_path: Optional[Path] = None
    if path:
        config_path = Path(path)
        if not config_path.exists():
            raise FileNotFoundError(f"Model config file not found: {config_path}")
    elif model_path:
        config_path = guess_model_config_path(model_path)

    if not config_path:
        return {}

    data = json.loads(config_path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise RuntimeError("Model config must be a JSON object.")

    unknown = set(data.keys()) - MODEL_CONFIG_KEYS
    if unknown:
        # Ignore unknown keys, but keep this transparent.
        print(f"[model-config] ignoring unknown keys: {sorted(unknown)}")
    print(f"[model-config] loaded: {config_path}")
    return data


def load_roi(path: str) -> Optional[ROI]:
    p = Path(path)
    if not p.exists():
        return None
    data = json.loads(p.read_text(encoding="utf-8"))
    return ROI(int(data["x"]), int(data["y"]), int(data["w"]), int(data["h"]))


def save_roi(path: str, roi: ROI) -> None:
    data = {"x": roi.x, "y": roi.y, "w": roi.w, "h": roi.h}
    Path(path).write_text(json.dumps(data, indent=2), encoding="utf-8")


def parse_manual_roi(text: str) -> ROI:
    parts = [p.strip() for p in text.split(",")]
    if len(parts) != 4:
        raise RuntimeError("--roi must be in format x,y,w,h")
    try:
        x, y, w, h = [int(p) for p in parts]
    except ValueError as exc:
        raise RuntimeError("--roi contains non-integer values") from exc
    if w <= 0 or h <= 0:
        raise RuntimeError("--roi width and height must be positive")
    return ROI(x=x, y=y, w=w, h=h)


def choose_roi(image_bgr: np.ndarray) -> ROI:
    title = "Select UDISE ROI (Enter/Space confirm, c cancel)"
    x, y, w, h = cv2.selectROI(title, image_bgr.copy(), showCrosshair=True, fromCenter=False)
    cv2.destroyWindow(title)
    if w <= 0 or h <= 0:
        raise RuntimeError("ROI selection canceled or invalid.")
    return ROI(x=int(x), y=int(y), w=int(w), h=int(h))


def crop_roi(image_bgr: np.ndarray, roi: ROI) -> np.ndarray:
    h, w = image_bgr.shape[:2]
    x0 = max(0, roi.x)
    y0 = max(0, roi.y)
    x1 = min(w, roi.x + roi.w)
    y1 = min(h, roi.y + roi.h)
    if x1 <= x0 or y1 <= y0:
        raise RuntimeError("ROI is outside image bounds.")
    return image_bgr[y0:y1, x0:x1]


def split_digits_equal(udise_roi_bgr: np.ndarray, digit_count: int) -> List[np.ndarray]:
    _, w = udise_roi_bgr.shape[:2]
    if w < digit_count:
        raise RuntimeError(
            f"ROI width {w} too small for {digit_count} digits. Reselect ROI."
        )

    digit_images: List[np.ndarray] = []
    cell_width = w / float(digit_count)
    for i in range(digit_count):
        x0 = int(round(i * cell_width))
        x1 = int(round((i + 1) * cell_width))
        x1 = max(x1, x0 + 1)
        digit_images.append(udise_roi_bgr[:, x0:x1])
    return digit_images


def preprocess_digit(
    digit_bgr: np.ndarray,
    image_size: int = 28,
    invert: bool = False,
    use_morphology_cleaning: bool = False,
) -> np.ndarray:
    gray = cv2.cvtColor(digit_bgr, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (3, 3), 0)

    _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    if invert:
        th = cv2.bitwise_not(th)

    if use_morphology_cleaning:
        # Cleaner expects bright foreground structures.
        clean_src = th if invert else cv2.bitwise_not(th)
        clean_out = remove_grid_lines(clean_src)
        th = clean_out if invert else cv2.bitwise_not(clean_out)

    # Crop around foreground when present.
    nz = cv2.findNonZero(255 - th if not invert else th)
    if nz is not None:
        x, y, w, h = cv2.boundingRect(nz)
        th = th[y : y + h, x : x + w]

    h, w = th.shape[:2]
    scale = image_size / float(max(h, w))
    new_w = max(1, int(round(w * scale)))
    new_h = max(1, int(round(h * scale)))
    resized = cv2.resize(th, (new_w, new_h), interpolation=cv2.INTER_AREA)

    canvas = np.full((image_size, image_size), 255, dtype=np.uint8)
    y0 = (image_size - new_h) // 2
    x0 = (image_size - new_w) // 2
    canvas[y0 : y0 + new_h, x0 : x0 + new_w] = resized

    arr = canvas.astype(np.float32) / 255.0
    return np.expand_dims(arr, axis=-1)


def decode_image_from_source(
    source: str,
    timeout: int,
    retries: int,
    retry_backoff: float,
    session: requests.Session,
) -> np.ndarray:
    parsed = urlparse(source)
    scheme = parsed.scheme.lower()

    if scheme in ("http", "https"):
        attempts = max(1, retries + 1)
        last_error: Optional[Exception] = None
        for attempt in range(1, attempts + 1):
            try:
                resp = session.get(source, timeout=timeout)
                resp.raise_for_status()
                data = np.frombuffer(resp.content, dtype=np.uint8)
                image = cv2.imdecode(data, cv2.IMREAD_COLOR)
                if image is None:
                    raise RuntimeError("Failed to decode image bytes.")
                return image
            except Exception as exc:
                last_error = exc
                if attempt < attempts:
                    time.sleep(retry_backoff * attempt)
        raise RuntimeError(f"Download failed after {attempts} attempts: {last_error}")

    local_path = source
    if scheme == "file":
        local_path = unquote(parsed.path)

    p = Path(local_path).expanduser()
    if not p.exists():
        raise FileNotFoundError(f"Local image not found: {p}")

    image = cv2.imread(str(p), cv2.IMREAD_COLOR)
    if image is None:
        raise RuntimeError(f"Failed to read local image: {p}")
    return image


def normalize_header(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()


def sheet_headers(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val is None:
            continue
        headers[normalize_header(str(val))] = col
    return headers


def resolve_column(
    ws,
    column_ref: Optional[str],
    auto_candidates: Sequence[str],
    fallback_excel_column: str,
    allow_create: bool = False,
) -> int:
    if column_ref is None:
        headers = sheet_headers(ws)
        for candidate in auto_candidates:
            key = normalize_header(candidate)
            if key in headers:
                return headers[key]
        return column_index_from_string(fallback_excel_column)

    ref = str(column_ref).strip()
    if ref.isalpha() and 1 <= len(ref) <= 3:
        return column_index_from_string(ref.upper())

    headers = sheet_headers(ws)
    key = normalize_header(ref)
    if key in headers:
        return headers[key]

    if allow_create:
        col = ws.max_column + 1
        ws.cell(1, col).value = ref
        return col

    raise RuntimeError(f"Could not resolve column '{column_ref}'.")


def resolve_sheet(wb, sheet_ref: Optional[str]):
    if sheet_ref is None:
        return wb[wb.sheetnames[0]]
    s = str(sheet_ref).strip()
    if s.isdigit():
        idx = int(s)
        if idx < 1 or idx > len(wb.sheetnames):
            raise RuntimeError(f"Sheet index out of range: {idx}")
        return wb[wb.sheetnames[idx - 1]]
    if s not in wb.sheetnames:
        raise RuntimeError(f"Sheet not found: {s}")
    return wb[s]


def dump_debug(debug_dir: str, row_num: int, roi_img: np.ndarray, digit_cells: Sequence[np.ndarray]) -> None:
    d = Path(debug_dir)
    d.mkdir(parents=True, exist_ok=True)
    cv2.imwrite(str(d / f"row_{row_num}_roi.png"), roi_img)
    for i, cell in enumerate(digit_cells):
        cv2.imwrite(str(d / f"row_{row_num}_digit_cell_{i}.png"), cell)


def write_failure_log(path: str, failures: Sequence[Tuple[int, str, str]]) -> None:
    if not path:
        return
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["row", "source", "error"])
        for row, source, err in failures:
            w.writerow([row, source, err])


def main() -> None:
    args = parse_args()
    if args.framework in ("keras", "torch") and not args.model_path.strip():
        raise RuntimeError("--model-path is required for keras/torch frameworks.")
    if args.framework == "groq":
        if not args.api_key.strip():
            raise RuntimeError("--api-key is required for --framework groq.")
        if not args.api_model.strip():
            raise RuntimeError("--api-model is required for --framework groq.")

    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel not found: {excel_path}")

    output_excel = (
        Path(args.output_excel)
        if args.output_excel
        else excel_path.with_name(f"{excel_path.stem}_with_udise.xlsx")
    )

    model_config = load_model_config(args.model_config, args.model_path)

    config_framework = model_config.get("framework")
    if config_framework and str(config_framework) != args.framework:
        raise RuntimeError(
            f"Framework mismatch: CLI --framework={args.framework} but config says {config_framework}"
        )

    effective_digit_count = int(model_config.get("digit_count", args.digit_count))
    if args.digit_count != EXPECTED_DIGITS and args.digit_count != effective_digit_count:
        effective_digit_count = args.digit_count

    config_invert = model_config.get("invert", None)
    effective_invert = bool(config_invert) if args.invert is None and config_invert is not None else bool(args.invert)

    config_image_size = model_config.get("image_size", None)
    effective_image_size = int(config_image_size) if args.image_size is None and config_image_size is not None else (args.image_size or 28)
    if effective_image_size <= 0:
        raise RuntimeError("--image-size must be positive")

    config_morph = model_config.get("use_morphology_cleaning", None)
    effective_morphology_cleaning = (
        bool(config_morph)
        if args.use_morphology_cleaning is None and config_morph is not None
        else bool(args.use_morphology_cleaning)
    )

    runner = get_model_runner(
        framework=args.framework,
        model_path=args.model_path,
        api_key=args.api_key,
        api_model=args.api_model,
        api_endpoint=args.api_endpoint,
        timeout=args.timeout,
    )

    wb = load_workbook(excel_path)
    ws = resolve_sheet(wb, args.sheet)

    url_col = resolve_column(
        ws=ws,
        column_ref=args.url_column,
        auto_candidates=URL_HEADER_CANDIDATES,
        fallback_excel_column=DEFAULT_URL_COLUMN,
    )
    out_col = resolve_column(
        ws=ws,
        column_ref=args.output_column,
        auto_candidates=OUTPUT_HEADER_CANDIDATES,
        fallback_excel_column=DEFAULT_OUTPUT_COLUMN,
        allow_create=True,
    )

    if args.roi:
        roi = parse_manual_roi(args.roi)
    else:
        roi = None if args.pick_roi else load_roi(args.roi_file)

    first_debug_dumped = False
    processed = 0
    success = 0
    skipped = 0
    failures: List[Tuple[int, str, str]] = []

    first_row = max(2, args.start_row)
    max_sheet_row = ws.max_row
    last_row = (
        max_sheet_row
        if args.max_rows <= 0
        else min(max_sheet_row, first_row + args.max_rows - 1)
    )
    if first_row > last_row:
        raise RuntimeError("No rows to process with current --start-row/--max-rows.")

    total_candidates = last_row - first_row + 1
    run_started = time.time()
    autosave_counter = 0
    session = requests.Session()

    print(f"Sheet: {ws.title}")
    print(f"Rows: {first_row}..{last_row} ({total_candidates} candidate rows)")
    print(f"URL column index: {url_col}")
    print(f"Output column index: {out_col}")
    print(f"Digit count: {effective_digit_count}")
    print(f"Preprocess invert: {effective_invert}")
    print(f"Image size: {effective_image_size}")
    print(f"Morphology cleaning: {effective_morphology_cleaning}")

    try:
        for row in range(first_row, last_row + 1):
            existing = ws.cell(row=row, column=out_col).value
            if (existing is not None) and str(existing).strip() and not args.overwrite:
                skipped += 1
                continue

            source_val = ws.cell(row=row, column=url_col).value
            if not source_val:
                skipped += 1
                continue
            source = str(source_val).strip()
            if not source:
                skipped += 1
                continue

            processed += 1
            try:
                image_bgr = decode_image_from_source(
                    source=source,
                    timeout=args.timeout,
                    retries=args.retries,
                    retry_backoff=args.retry_backoff,
                    session=session,
                )

                if roi is None:
                    roi = choose_roi(image_bgr)
                    save_roi(args.roi_file, roi)

                udise_crop = crop_roi(image_bgr, roi)
                digit_cells = split_digits_equal(udise_crop, effective_digit_count)

                if args.debug_dir and not first_debug_dumped:
                    dump_debug(args.debug_dir, row, udise_crop, digit_cells)
                    first_debug_dumped = True

                batch = np.stack(
                    [
                        preprocess_digit(
                            cell,
                            image_size=effective_image_size,
                            invert=effective_invert,
                            use_morphology_cleaning=effective_morphology_cleaning,
                        )
                        for cell in digit_cells
                    ],
                    axis=0,
                )
                preds = runner.predict(batch)
                if len(preds) != effective_digit_count:
                    raise RuntimeError(
                        f"Model returned {len(preds)} digits, expected {effective_digit_count}."
                    )

                udise_code = "".join(str(int(p)) for p in preds)
                ws.cell(row=row, column=out_col).value = udise_code
                success += 1
                autosave_counter += 1

                if not args.quiet:
                    print(f"row {row}: {udise_code}")

                if args.save_every > 0 and autosave_counter >= args.save_every:
                    wb.save(output_excel)
                    autosave_counter = 0
                    if not args.quiet:
                        print(f"[autosave] saved: {output_excel}")
            except Exception as exc:
                failures.append((row, source, str(exc)))
                if not args.quiet:
                    print(f"row {row}: ERROR: {exc}")
    except KeyboardInterrupt:
        print("\nInterrupted by user. Saving partial output...")
    finally:
        session.close()

    wb.save(output_excel)
    write_failure_log(args.failure_log, failures)

    elapsed = time.time() - run_started
    print("\nDone.")
    print(f"Elapsed: {elapsed:.1f}s")
    print(f"Candidate rows: {total_candidates}")
    print(f"Processed rows: {processed}")
    print(f"Skipped rows: {skipped}")
    print(f"Success: {success}")
    print(f"Failed: {len(failures)}")
    print(f"Output Excel: {output_excel}")
    if failures:
        print(f"Failure log: {args.failure_log}")
    if roi is not None:
        print(f"ROI saved in: {args.roi_file}")

    if failures:
        print("\nFailed rows:")
        for row, _, err in failures[:20]:
            print(f" - row {row}: {err}")
        if len(failures) > 20:
            print(f" ... and {len(failures) - 20} more")


if __name__ == "__main__":
    main()
