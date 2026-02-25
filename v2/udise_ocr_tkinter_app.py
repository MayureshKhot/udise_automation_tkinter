#!/usr/bin/env python3
"""Minimal Tkinter UI for UDISE OCR calibration and prediction.

This v2 app intentionally keeps only essential functionality:
- OMR image preview
- ROI selection from image
- ROI movement with directional controls
- Model loading and prediction
"""

from __future__ import annotations

import tkinter as tk
from dataclasses import dataclass
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import List, Optional, Sequence, Tuple

import cv2
import numpy as np
import requests
from openpyxl import load_workbook
from PIL import Image, ImageTk

from udise_ocr_pipeline import (
    KerasRunner,
    TorchRunner,
    crop_roi,
    decode_image_from_source,
    get_model_runner,
    load_model_config,
    parse_manual_roi,
    resolve_column,
    resolve_sheet,
    split_digits_equal,
)


@dataclass
class PreprocessSettings:
    image_size: int = 28
    invert: bool = True
    blur_kernel: int = 3
    tight_crop: bool = True
    bbox_padding: int = 2


def ensure_odd(value: int, minimum: int) -> int:
    v = max(minimum, int(value))
    if v % 2 == 0:
        v += 1
    return v


def preprocess_digit_custom(digit_bgr: np.ndarray, s: PreprocessSettings) -> Tuple[np.ndarray, np.ndarray]:
    gray = cv2.cvtColor(digit_bgr, cv2.COLOR_BGR2GRAY)

    blur_k = ensure_odd(s.blur_kernel, 1)
    if blur_k > 1:
        gray = cv2.GaussianBlur(gray, (blur_k, blur_k), 0)

    _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    if s.invert:
        th = cv2.bitwise_not(th)

    proc = remove_grid_box_lines_dynamic(th, s.invert)
    if s.tight_crop:
        fg_mask = proc if s.invert else (255 - proc)
        nz = cv2.findNonZero(fg_mask)
        if nz is not None:
            x, y, w, h = cv2.boundingRect(nz)
            pad = max(0, int(s.bbox_padding))
            x0 = max(0, x - pad)
            y0 = max(0, y - pad)
            x1 = min(proc.shape[1], x + w + pad)
            y1 = min(proc.shape[0], y + h + pad)
            proc = proc[y0:y1, x0:x1]

    h, w = proc.shape[:2]
    image_size = max(8, int(s.image_size))
    scale = image_size / float(max(h, w))
    new_w = max(1, int(round(w * scale)))
    new_h = max(1, int(round(h * scale)))
    resized = cv2.resize(proc, (new_w, new_h), interpolation=cv2.INTER_AREA)

    bg = 0 if s.invert else 255
    canvas = np.full((image_size, image_size), bg, dtype=np.uint8)
    y0 = (image_size - new_h) // 2
    x0 = (image_size - new_w) // 2
    canvas[y0 : y0 + new_h, x0 : x0 + new_w] = resized

    arr = canvas.astype(np.float32) / 255.0
    arr = np.expand_dims(arr, axis=-1)
    return arr, canvas


def remove_grid_box_lines_dynamic(binary_img: np.ndarray, invert: bool) -> np.ndarray:
    """Remove straight grid-box lines while preserving handwritten strokes.

    Works per digit cell:
    - Extract long vertical/horizontal structures morphologically.
    - Remove only components that are tall/wide and near cell borders.
    - Trim residual border columns/rows that still look like ruling lines.
    """
    h, w = binary_img.shape[:2]
    if h < 8 or w < 8:
        return binary_img

    # Foreground mask in white: handwritten digit + grid lines.
    fg = binary_img.copy() if invert else (255 - binary_img)

    # Dynamic kernels based on cell geometry.
    v_kernel_h = max(5, int(round(h * 0.55)))
    h_kernel_w = max(5, int(round(w * 0.55)))
    v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, v_kernel_h))
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (h_kernel_w, 1))

    vertical = cv2.morphologyEx(fg, cv2.MORPH_OPEN, v_kernel)
    horizontal = cv2.morphologyEx(fg, cv2.MORPH_OPEN, h_kernel)
    line_mask = cv2.bitwise_or(vertical, horizontal)

    n_labels, labels, stats, _ = cv2.connectedComponentsWithStats(line_mask, connectivity=8)
    remove_mask = np.zeros_like(fg)
    edge_x = max(1, int(round(w * 0.22)))
    edge_y = max(1, int(round(h * 0.22)))
    max_line_w = max(2, int(round(w * 0.18)))
    max_line_h = max(2, int(round(h * 0.18)))

    for i in range(1, n_labels):
        x = int(stats[i, cv2.CC_STAT_LEFT])
        y = int(stats[i, cv2.CC_STAT_TOP])
        cw = int(stats[i, cv2.CC_STAT_WIDTH])
        ch = int(stats[i, cv2.CC_STAT_HEIGHT])
        area = int(stats[i, cv2.CC_STAT_AREA])

        if area <= 0:
            continue

        is_vertical_line = ch >= int(0.5 * h) and cw <= max_line_w
        is_horizontal_line = cw >= int(0.5 * w) and ch <= max_line_h

        near_left = x <= edge_x
        near_right = (x + cw) >= (w - edge_x)
        near_top = y <= edge_y
        near_bottom = (y + ch) >= (h - edge_y)

        if is_vertical_line and (near_left or near_right):
            remove_mask[labels == i] = 255
        elif is_horizontal_line and (near_top or near_bottom):
            remove_mask[labels == i] = 255

    fg_clean = cv2.bitwise_and(fg, cv2.bitwise_not(remove_mask))

    # Secondary dynamic trim for residual border rulings.
    fg_clean = _trim_border_rulings(fg_clean)

    # Safety fallback if too much foreground got removed.
    before = int(np.count_nonzero(fg))
    after = int(np.count_nonzero(fg_clean))
    if before > 0 and (after / float(before)) < 0.25:
        fg_clean = fg

    return fg_clean if invert else (255 - fg_clean)


def _trim_border_rulings(fg_white: np.ndarray) -> np.ndarray:
    """Trim border columns/rows that are almost full-height/width straight lines."""
    h, w = fg_white.shape[:2]
    if h < 8 or w < 8:
        return fg_white

    out = fg_white.copy()
    col_counts = np.count_nonzero(out, axis=0)
    row_counts = np.count_nonzero(out, axis=1)

    # A ruling line tends to occupy most of a border column/row.
    col_thr = int(round(h * 0.72))
    row_thr = int(round(w * 0.72))
    max_trim_cols = max(1, int(round(w * 0.2)))
    max_trim_rows = max(1, int(round(h * 0.2)))

    left = 0
    while left < min(max_trim_cols, w) and col_counts[left] >= col_thr:
        out[:, left] = 0
        left += 1

    right = w - 1
    rtrim = 0
    while right >= 0 and rtrim < max_trim_cols and col_counts[right] >= col_thr:
        out[:, right] = 0
        right -= 1
        rtrim += 1

    top = 0
    while top < min(max_trim_rows, h) and row_counts[top] >= row_thr:
        out[top, :] = 0
        top += 1

    bottom = h - 1
    btrim = 0
    while bottom >= 0 and btrim < max_trim_rows and row_counts[bottom] >= row_thr:
        out[bottom, :] = 0
        bottom -= 1
        btrim += 1

    return out


class UDISEMinimalApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("UDISE OCR - v2 Minimal")
        self.geometry("1320x860")
        self.minsize(1080, 740)

        self.runner = None
        self.current_image_bgr: Optional[np.ndarray] = None
        self.current_source: Optional[str] = None
        self.current_row: Optional[int] = None
        self.roi: Optional[Tuple[int, int, int, int]] = None
        self.row_sequence: List[int] = []
        self.row_cursor: int = -1
        self._run_all_active = False

        self.photo_refs = {}
        self.digit_photo_refs: List[ImageTk.PhotoImage] = []

        self._build_vars()
        self._build_ui()

    def _build_vars(self) -> None:
        self.excel_var = tk.StringVar(value="")
        self.output_excel_var = tk.StringVar(value="")
        self.sheet_var = tk.StringVar(value="")
        self.url_col_var = tk.StringVar(value="Original Image Path")
        self.out_col_var = tk.StringVar(value="UDISE Found")
        self.sample_row_var = tk.IntVar(value=2)
        self.start_row_var = tk.IntVar(value=2)
        self.overwrite_var = tk.BooleanVar(value=False)
        self.nav_var = tk.StringVar(value="Rows: 0")

        self.image_source_var = tk.StringVar(value="")
        self.framework_var = tk.StringVar(value="keras")
        self.model_path_var = tk.StringVar(value="")
        self.model_config_var = tk.StringVar(value="")

        self.digit_count_var = tk.IntVar(value=11)
        self.image_size_var = tk.IntVar(value=28)
        self.invert_var = tk.BooleanVar(value=True)
        self.shift_step_var = tk.IntVar(value=1)

        self.roi_text_var = tk.StringVar(value="")
        self.pred_var = tk.StringVar(value="Prediction: -")
        self.status_var = tk.StringVar(value="Ready")

    def _build_ui(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)
        self.rowconfigure(2, weight=1)

        top = ttk.Frame(self, padding=8)
        top.grid(row=0, column=0, sticky="ew")
        top.columnconfigure(1, weight=1)

        ttk.Label(top, text="Excel Input").grid(row=0, column=0, sticky="w")
        ttk.Entry(top, textvariable=self.excel_var).grid(row=0, column=1, sticky="ew", padx=(8, 4))
        ttk.Button(top, text="Browse", command=self.pick_excel).grid(row=0, column=2, sticky="ew")
        ttk.Button(top, text="Output", command=self.pick_output_excel).grid(row=0, column=3, sticky="ew", padx=(4, 0))

        ttk.Label(top, text="Image Path/URL").grid(row=1, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(top, textvariable=self.image_source_var).grid(row=1, column=1, sticky="ew", padx=(8, 4), pady=(6, 0))
        ttk.Button(top, text="Browse", command=self.pick_image).grid(row=1, column=2, sticky="ew", pady=(6, 0))
        ttk.Button(top, text="Load Image", command=self.load_image).grid(row=1, column=3, sticky="ew", padx=(4, 0), pady=(6, 0))

        ttk.Label(top, text="Model").grid(row=2, column=0, sticky="w", pady=(6, 0))
        ttk.Entry(top, textvariable=self.model_path_var).grid(row=2, column=1, sticky="ew", padx=(8, 4), pady=(6, 0))
        ttk.Button(top, text="Browse", command=self.pick_model).grid(row=2, column=2, sticky="ew", pady=(6, 0))
        ttk.Button(top, text="Load Model", command=self.load_model).grid(row=2, column=3, sticky="ew", padx=(4, 0), pady=(6, 0))

        options = ttk.Frame(top)
        options.grid(row=3, column=0, columnspan=4, sticky="ew", pady=(8, 0))
        for i in range(20):
            options.columnconfigure(i, weight=1)

        ttk.Label(options, text="Framework").grid(row=0, column=0, sticky="w")
        ttk.Combobox(options, textvariable=self.framework_var, state="readonly", values=["keras", "torch"], width=8).grid(
            row=0, column=1, sticky="w"
        )

        ttk.Label(options, text="Model Config").grid(row=0, column=2, sticky="e")
        ttk.Entry(options, textvariable=self.model_config_var).grid(row=0, column=3, columnspan=2, sticky="ew", padx=(4, 4))
        ttk.Button(options, text="Browse", command=self.pick_model_config).grid(row=0, column=5, sticky="ew")

        ttk.Label(options, text="Digits").grid(row=0, column=6, sticky="e")
        ttk.Spinbox(options, from_=1, to=20, textvariable=self.digit_count_var, width=6).grid(row=0, column=7, sticky="w")

        ttk.Label(options, text="Sheet").grid(row=0, column=8, sticky="e")
        ttk.Entry(options, textvariable=self.sheet_var, width=8).grid(row=0, column=9, sticky="w")

        ttk.Label(options, text="Src Col").grid(row=0, column=10, sticky="e")
        ttk.Entry(options, textvariable=self.url_col_var, width=12).grid(row=0, column=11, sticky="w")

        ttk.Label(options, text="Out Col").grid(row=0, column=12, sticky="e")
        ttk.Entry(options, textvariable=self.out_col_var, width=12).grid(row=0, column=13, sticky="w")

        ttk.Label(options, text="Start Row").grid(row=0, column=14, sticky="e")
        ttk.Spinbox(options, from_=2, to=999999, textvariable=self.start_row_var, width=8).grid(row=0, column=15, sticky="w")
        ttk.Checkbutton(options, text="Overwrite", variable=self.overwrite_var).grid(row=0, column=16, sticky="w")
        ttk.Button(options, text="Scan URLs", command=self.scan_excel_rows).grid(row=0, column=17, columnspan=3, sticky="ew")

        ttk.Label(options, text="Row").grid(row=1, column=0, sticky="e", pady=(6, 0))
        ttk.Spinbox(options, from_=2, to=999999, textvariable=self.sample_row_var, width=8).grid(row=1, column=1, sticky="w", pady=(6, 0))
        ttk.Button(options, text="Load Row Image", command=self.load_sample_row).grid(row=1, column=2, columnspan=2, sticky="ew", pady=(6, 0))
        ttk.Button(options, text="Previous", command=self.load_prev_row).grid(row=1, column=4, sticky="ew", pady=(6, 0))
        ttk.Button(options, text="Next", command=self.load_next_row).grid(row=1, column=5, sticky="ew", pady=(6, 0))
        ttk.Button(options, text="Predict + Write", command=self.predict).grid(row=1, column=6, columnspan=3, sticky="ew", pady=(6, 0))
        ttk.Button(options, text="Run All URLs", command=self.run_all_urls).grid(row=1, column=9, columnspan=3, sticky="ew", pady=(6, 0))

        ttk.Label(options, text="Shift Step").grid(row=1, column=12, sticky="e", pady=(6, 0))
        ttk.Spinbox(options, from_=1, to=100, textvariable=self.shift_step_var, width=6).grid(row=1, column=13, sticky="w", pady=(6, 0))
        ttk.Checkbutton(options, text="Invert", variable=self.invert_var).grid(row=1, column=14, sticky="w", pady=(6, 0))
        ttk.Label(options, textvariable=self.nav_var).grid(row=1, column=15, columnspan=5, sticky="e", pady=(6, 0))

        preview = ttk.LabelFrame(self, text="OMR Preview", padding=8)
        preview.grid(row=1, column=0, sticky="nsew", padx=8)
        preview.columnconfigure(0, weight=1)
        preview.columnconfigure(1, weight=1)
        preview.rowconfigure(0, weight=1)

        self.full_img_label = ttk.Label(preview, text="Load image to preview")
        self.full_img_label.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        self.roi_img_label = ttk.Label(preview, text="Select ROI")
        self.roi_img_label.grid(row=0, column=1, sticky="nsew")

        controls = ttk.Frame(self, padding=(8, 4, 8, 8))
        controls.grid(row=2, column=0, sticky="nsew")
        controls.columnconfigure(0, weight=0)
        controls.columnconfigure(1, weight=1)
        controls.rowconfigure(0, weight=1)

        move = ttk.LabelFrame(controls, text="ROI Move", padding=8)
        move.grid(row=0, column=0, sticky="ns", padx=(0, 8))

        ttk.Button(move, text="Select ROI", command=self.select_roi).grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 6))
        ttk.Button(move, text="Apply ROI", command=self.apply_roi_text).grid(row=1, column=0, columnspan=3, sticky="ew")
        ttk.Entry(move, textvariable=self.roi_text_var, width=18).grid(row=2, column=0, columnspan=3, sticky="ew", pady=(6, 6))

        ttk.Button(move, text="↑", width=4, command=lambda: self.shift_roi(0, -self._shift_step())).grid(row=3, column=1, pady=2)
        ttk.Button(move, text="←", width=4, command=lambda: self.shift_roi(-self._shift_step(), 0)).grid(row=4, column=0, padx=2)
        ttk.Button(move, text="→", width=4, command=lambda: self.shift_roi(self._shift_step(), 0)).grid(row=4, column=2, padx=2)
        ttk.Button(move, text="↓", width=4, command=lambda: self.shift_roi(0, self._shift_step())).grid(row=5, column=1, pady=2)

        digits = ttk.LabelFrame(controls, text="Preprocessed Digits", padding=8)
        digits.grid(row=0, column=1, sticky="nsew")
        digits.columnconfigure(0, weight=1)
        digits.rowconfigure(0, weight=1)

        self.digit_canvas = tk.Canvas(digits, height=190, highlightthickness=0)
        self.digit_canvas.grid(row=0, column=0, sticky="nsew")
        xscroll = ttk.Scrollbar(digits, orient="horizontal", command=self.digit_canvas.xview)
        xscroll.grid(row=1, column=0, sticky="ew")
        self.digit_canvas.configure(xscrollcommand=xscroll.set)
        self.digit_frame = ttk.Frame(self.digit_canvas)
        self.digit_canvas.create_window((0, 0), window=self.digit_frame, anchor="nw")
        self.digit_frame.bind("<Configure>", lambda _e: self.digit_canvas.configure(scrollregion=self.digit_canvas.bbox("all")))

        footer = ttk.Frame(self, padding=(8, 0, 8, 8))
        footer.grid(row=3, column=0, sticky="ew")
        footer.columnconfigure(0, weight=1)

        ttk.Label(footer, textvariable=self.pred_var, font=("TkDefaultFont", 11, "bold")).grid(row=0, column=0, sticky="w")
        ttk.Label(footer, textvariable=self.status_var).grid(row=0, column=1, sticky="e")

    def _settings(self) -> PreprocessSettings:
        return PreprocessSettings(
            image_size=max(8, int(self.image_size_var.get())),
            invert=bool(self.invert_var.get()),
            blur_kernel=3,
            tight_crop=True,
            bbox_padding=2,
        )

    def _shift_step(self) -> int:
        try:
            return max(1, int(self.shift_step_var.get()))
        except Exception:
            return 1

    def _to_photo(self, bgr_or_gray: np.ndarray, max_size=(620, 420)) -> ImageTk.PhotoImage:
        if bgr_or_gray.ndim == 2:
            rgb = cv2.cvtColor(bgr_or_gray, cv2.COLOR_GRAY2RGB)
        else:
            rgb = cv2.cvtColor(bgr_or_gray, cv2.COLOR_BGR2RGB)
        h, w = rgb.shape[:2]
        scale = min(max_size[0] / w, max_size[1] / h, 1.0)
        if scale < 1.0:
            rgb = cv2.resize(rgb, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_AREA)
        return ImageTk.PhotoImage(Image.fromarray(rgb))

    def _photo_for_widget(self, arr: np.ndarray, widget, fallback=(620, 420), pad=(16, 16)) -> ImageTk.PhotoImage:
        try:
            ww = max(1, int(widget.winfo_width()) - int(pad[0]))
            wh = max(1, int(widget.winfo_height()) - int(pad[1]))
            if ww <= 10 or wh <= 10:
                return self._to_photo(arr, max_size=fallback)
            return self._to_photo(arr, max_size=(ww, wh))
        except Exception:
            return self._to_photo(arr, max_size=fallback)

    def _select_roi_tk(self, image_bgr: np.ndarray) -> Optional[Tuple[int, int, int, int]]:
        src = image_bgr.copy()
        h, w = src.shape[:2]
        rgb = cv2.cvtColor(src, cv2.COLOR_BGR2RGB)

        top = tk.Toplevel(self)
        top.title("Select ROI")
        top.geometry("1200x820")
        top.transient(self)
        top.grab_set()

        header = ttk.Frame(top, padding=(8, 8, 8, 4))
        header.pack(fill="x")
        ttk.Label(header, text="Drag with left mouse to draw ROI. Confirm to apply.").pack(side="left")
        status_var = tk.StringVar(value="ROI: -")
        ttk.Label(header, textvariable=status_var).pack(side="right")

        body = ttk.Frame(top, padding=(8, 0, 8, 0))
        body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=1)
        body.rowconfigure(0, weight=1)

        canvas = tk.Canvas(body, background="#1e1e1e", highlightthickness=0, cursor="crosshair")
        hbar = ttk.Scrollbar(body, orient="horizontal", command=canvas.xview)
        vbar = ttk.Scrollbar(body, orient="vertical", command=canvas.yview)
        canvas.configure(xscrollcommand=hbar.set, yscrollcommand=vbar.set)

        canvas.grid(row=0, column=0, sticky="nsew")
        vbar.grid(row=0, column=1, sticky="ns")
        hbar.grid(row=1, column=0, sticky="ew")

        footer = ttk.Frame(top, padding=8)
        footer.pack(fill="x")

        state = {
            "photo": None,
            "start": None,
            "roi_src": None,
            "disp_w": 1,
            "disp_h": 1,
            "scale": 1.0,
            "fit_scale": 1.0,
            "zoom": 1.0,
            "initialized": False,
        }
        img_tag = "IMG"
        roi_tag = "ROI"
        result_holder: dict = {"roi": None}

        def _render() -> None:
            scale = float(state["scale"])
            disp_w = max(1, int(round(w * scale)))
            disp_h = max(1, int(round(h * scale)))
            interp = cv2.INTER_LINEAR if scale > 1.0 else cv2.INTER_AREA
            resized = cv2.resize(rgb, (disp_w, disp_h), interpolation=interp)
            photo = ImageTk.PhotoImage(Image.fromarray(resized))

            state["photo"] = photo
            state["disp_w"] = disp_w
            state["disp_h"] = disp_h

            canvas.delete(img_tag)
            canvas.create_image(0, 0, image=photo, anchor="nw", tags=img_tag)
            canvas.configure(scrollregion=(0, 0, disp_w, disp_h))
            _draw_roi()

        def _set_fit_scale() -> None:
            canvas_w = max(1, int(canvas.winfo_width()))
            canvas_h = max(1, int(canvas.winfo_height()))
            fit_scale = min(canvas_w / float(w), canvas_h / float(h), 1.0)
            state["fit_scale"] = max(1e-6, fit_scale)
            state["scale"] = float(state["fit_scale"]) * float(state["zoom"])

        def _canvas_to_src(cx: float, cy: float) -> Optional[Tuple[float, float]]:
            disp_w = float(state["disp_w"])
            disp_h = float(state["disp_h"])
            if cx < 0 or cy < 0 or cx > disp_w or cy > disp_h:
                return None
            rel_x = cx / max(1e-6, disp_w)
            rel_y = cy / max(1e-6, disp_h)
            return rel_x * w, rel_y * h

        def _clamp_roi(x0: float, y0: float, x1: float, y1: float) -> Optional[Tuple[int, int, int, int]]:
            ax0 = int(round(max(0, min(x0, x1))))
            ay0 = int(round(max(0, min(y0, y1))))
            ax1 = int(round(min(w, max(x0, x1))))
            ay1 = int(round(min(h, max(y0, y1))))
            rw = ax1 - ax0
            rh = ay1 - ay0
            if rw <= 0 or rh <= 0:
                return None
            return ax0, ay0, rw, rh

        def _draw_roi() -> None:
            canvas.delete(roi_tag)
            roi = state["roi_src"]
            if not roi:
                return
            x, y, rw, rh = roi
            scale = float(state["scale"])
            canvas.create_rectangle(
                int(round(x * scale)),
                int(round(y * scale)),
                int(round((x + rw) * scale)),
                int(round((y + rh) * scale)),
                outline="#00ff7f",
                width=2,
                tags=roi_tag,
            )

        def on_left_down(event):
            cx = canvas.canvasx(event.x)
            cy = canvas.canvasy(event.y)
            p = _canvas_to_src(float(cx), float(cy))
            if p is None:
                return
            state["start"] = p
            state["roi_src"] = None
            status_var.set("ROI: -")
            canvas.delete(roi_tag)

        def on_left_drag(event):
            if state["start"] is None:
                return
            cx = canvas.canvasx(event.x)
            cy = canvas.canvasy(event.y)
            p = _canvas_to_src(float(cx), float(cy))
            if p is None:
                return
            sx, sy = state["start"]
            ex, ey = p
            roi = _clamp_roi(sx, sy, ex, ey)
            state["roi_src"] = roi
            if roi:
                x, y, rw, rh = roi
                status_var.set(f"ROI: {x},{y},{rw},{rh}")
            else:
                status_var.set("ROI: -")
            _draw_roi()

        def on_left_up(_event):
            state["start"] = None

        def _zoom_to(new_zoom: float) -> None:
            old_scale = float(state["scale"])
            state["zoom"] = max(0.2, min(8.0, float(new_zoom)))
            _set_fit_scale()
            _render()

            # Keep viewport roughly centered at previous center.
            vx = canvas.canvasx(canvas.winfo_width() / 2.0)
            vy = canvas.canvasy(canvas.winfo_height() / 2.0)
            if old_scale > 1e-6:
                rx = vx / old_scale
                ry = vy / old_scale
                new_scale = float(state["scale"])
                nx = rx * new_scale
                ny = ry * new_scale
                total_w = max(1.0, float(state["disp_w"]))
                total_h = max(1.0, float(state["disp_h"]))
                canvas.xview_moveto(max(0.0, min(1.0, (nx - canvas.winfo_width() / 2.0) / total_w)))
                canvas.yview_moveto(max(0.0, min(1.0, (ny - canvas.winfo_height() / 2.0) / total_h)))

        def zoom_in():
            _zoom_to(float(state["zoom"]) * 1.15)

        def zoom_out():
            _zoom_to(float(state["zoom"]) / 1.15)

        def zoom_reset():
            state["zoom"] = 1.0
            _set_fit_scale()
            _render()
            canvas.xview_moveto(0.0)
            canvas.yview_moveto(0.0)

        def zoom_fit():
            zoom_reset()

        def on_mousewheel(event):
            delta = getattr(event, "delta", 0)
            if delta > 0:
                zoom_in()
            elif delta < 0:
                zoom_out()

        def on_mousewheel_linux(event):
            if getattr(event, "num", 0) == 4:
                zoom_in()
            elif getattr(event, "num", 0) == 5:
                zoom_out()

        def confirm():
            roi = state["roi_src"]
            if roi is None:
                messagebox.showerror("ROI", "Draw ROI first.")
                return
            result_holder["roi"] = roi
            top.destroy()

        def cancel():
            result_holder["roi"] = None
            top.destroy()

        canvas.bind("<ButtonPress-1>", on_left_down)
        canvas.bind("<B1-Motion>", on_left_drag)
        canvas.bind("<ButtonRelease-1>", on_left_up)
        canvas.bind("<MouseWheel>", on_mousewheel)
        canvas.bind("<Button-4>", on_mousewheel_linux)
        canvas.bind("<Button-5>", on_mousewheel_linux)

        def on_canvas_configure(_event):
            if not bool(state["initialized"]):
                state["initialized"] = True
                _set_fit_scale()
                _render()
            else:
                # Recompute fit scale only; keep current zoom ratio.
                old_fit = max(1e-6, float(state["fit_scale"]))
                _set_fit_scale()
                ratio = float(state["scale"]) / max(1e-6, old_fit * float(state["zoom"]))
                if abs(ratio - 1.0) > 1e-6:
                    _render()

        canvas.bind("<Configure>", on_canvas_configure)

        ttk.Button(footer, text="Zoom -", command=zoom_out).pack(side="left", padx=2)
        ttk.Button(footer, text="Zoom +", command=zoom_in).pack(side="left", padx=2)
        ttk.Button(footer, text="100%", command=zoom_reset).pack(side="left", padx=2)
        ttk.Button(footer, text="Fit", command=zoom_fit).pack(side="left", padx=2)
        ttk.Button(footer, text="Cancel", command=cancel).pack(side="right", padx=2)
        ttk.Button(footer, text="Confirm ROI", command=confirm).pack(side="right", padx=2)

        top.protocol("WM_DELETE_WINDOW", cancel)
        top.after(60, lambda: (_set_fit_scale(), _render()))
        self.wait_window(top)
        return result_holder["roi"]

    def pick_image(self) -> None:
        p = filedialog.askopenfilename(filetypes=[("Images", "*.png *.jpg *.jpeg *.bmp *.tif *.tiff"), ("All", "*.*")])
        if p:
            self.image_source_var.set(p)

    def pick_excel(self) -> None:
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not p:
            return
        self.excel_var.set(p)
        if not self.output_excel_var.get().strip():
            out = Path(p).with_name(f"{Path(p).stem}_with_udise.xlsx")
            self.output_excel_var.set(str(out))
        self.scan_excel_rows(load_first=False)

    def pick_output_excel(self) -> None:
        p = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if p:
            self.output_excel_var.set(p)

    def pick_model(self) -> None:
        p = filedialog.askopenfilename(filetypes=[("Model", "*.keras *.h5 *.pt *.pth"), ("All", "*.*")])
        if p:
            self.model_path_var.set(p)
            cfg_guess = Path(str(p) + ".json")
            if cfg_guess.exists() and not self.model_config_var.get().strip():
                self.model_config_var.set(str(cfg_guess))

    def pick_model_config(self) -> None:
        p = filedialog.askopenfilename(filetypes=[("JSON", "*.json")])
        if p:
            self.model_config_var.set(p)

    def load_image(self) -> None:
        source = self.image_source_var.get().strip()
        if not source:
            messagebox.showerror("Image", "Select image path first.")
            return
        session = requests.Session()
        try:
            image = decode_image_from_source(
                source=source,
                timeout=20,
                retries=2,
                retry_backoff=1.5,
                session=session,
            )
        except Exception:
            image = cv2.imread(source, cv2.IMREAD_COLOR)
        finally:
            session.close()
        if image is None:
            messagebox.showerror("Image", f"Failed to read image: {source}")
            return

        self.current_source = source
        self.current_row = None
        self.current_image_bgr = image
        self.status_var.set(f"Image loaded: {Path(source).name}")
        self.refresh_preview()

    def _open_sheet(self):
        excel_path = Path(self.excel_var.get().strip())
        if not excel_path.exists():
            raise RuntimeError(f"Excel not found: {excel_path}")
        wb = load_workbook(excel_path)
        ws = resolve_sheet(wb, self.sheet_var.get().strip() or None)
        return wb, ws

    def _resolve_cols(self, ws) -> Tuple[int, int]:
        url_ref = self.url_col_var.get().strip() or None
        out_ref = self.out_col_var.get().strip() or None
        url_col = resolve_column(ws, url_ref, ["original image path", "image path on server", "image url", "url"], "H")
        out_col = resolve_column(ws, out_ref, ["udise found", "udise", "udise code"], "E", allow_create=True)
        return url_col, out_col

    def load_sample_row(self) -> None:
        self._load_row_from_excel(int(self.sample_row_var.get()))

    def _load_row_from_excel(self, row: int) -> None:
        try:
            wb, ws = self._open_sheet()
            url_col, _ = self._resolve_cols(ws)
            source = ws.cell(row=row, column=url_col).value
            wb.close()
            if not source or not str(source).strip():
                raise RuntimeError(f"Row {row} has empty source in source column.")

            session = requests.Session()
            try:
                image = decode_image_from_source(
                    source=str(source).strip(),
                    timeout=20,
                    retries=2,
                    retry_backoff=1.5,
                    session=session,
                )
            finally:
                session.close()

            self.current_image_bgr = image
            self.current_source = str(source).strip()
            self.current_row = row
            self.image_source_var.set(self.current_source)
            self.status_var.set(f"Loaded row {row} image")
            self._sync_row_cursor(row)
            self.refresh_preview()
        except Exception as exc:
            messagebox.showerror("Load Row", str(exc))

    def scan_excel_rows(self, load_first: bool = True) -> None:
        try:
            rows = self._collect_rows_from_excel()
            self.row_sequence = rows
            if not rows:
                self.row_cursor = -1
                self.nav_var.set("Rows: 0")
                self.status_var.set("No URLs found from selected start row.")
                return

            if int(self.sample_row_var.get()) in rows:
                self.row_cursor = rows.index(int(self.sample_row_var.get()))
            else:
                self.row_cursor = 0
                self.sample_row_var.set(rows[0])
            self._update_nav_label()
            self.status_var.set(f"Scanned {len(rows)} URL rows.")
            if load_first:
                self._load_row_from_excel(int(self.sample_row_var.get()))
        except Exception as exc:
            messagebox.showerror("Scan URLs", str(exc))

    def _collect_rows_from_excel(self) -> List[int]:
        wb, ws = self._open_sheet()
        try:
            url_col, _ = self._resolve_cols(ws)
            start = max(2, int(self.start_row_var.get()))
            rows: List[int] = []
            for row in range(start, ws.max_row + 1):
                source = ws.cell(row=row, column=url_col).value
                if source and str(source).strip():
                    rows.append(row)
            return rows
        finally:
            wb.close()

    def _update_nav_label(self) -> None:
        total = len(self.row_sequence)
        if total == 0 or self.row_cursor < 0:
            self.nav_var.set("Rows: 0")
            return
        self.nav_var.set(f"Row {self.row_cursor + 1}/{total}")

    def _sync_row_cursor(self, row: int) -> None:
        if row in self.row_sequence:
            self.row_cursor = self.row_sequence.index(row)
        self.sample_row_var.set(row)
        self._update_nav_label()

    def load_next_row(self) -> None:
        if not self.row_sequence:
            self.scan_excel_rows(load_first=False)
        if not self.row_sequence:
            return
        if self.row_cursor < 0:
            self.row_cursor = 0
        elif self.row_cursor < len(self.row_sequence) - 1:
            self.row_cursor += 1
        row = self.row_sequence[self.row_cursor]
        self.sample_row_var.set(row)
        self._load_row_from_excel(row)

    def load_prev_row(self) -> None:
        if not self.row_sequence:
            self.scan_excel_rows(load_first=False)
        if not self.row_sequence:
            return
        if self.row_cursor < 0:
            self.row_cursor = 0
        elif self.row_cursor > 0:
            self.row_cursor -= 1
        row = self.row_sequence[self.row_cursor]
        self.sample_row_var.set(row)
        self._load_row_from_excel(row)

    def load_model(self) -> None:
        framework = self.framework_var.get().strip()
        model_path = self.model_path_var.get().strip()
        if not model_path:
            messagebox.showerror("Model", "Select model path first.")
            return

        try:
            cfg = load_model_config(self.model_config_var.get().strip() or None, model_path)
            if "image_size" in cfg:
                self.image_size_var.set(int(cfg["image_size"]))
            if "invert" in cfg:
                self.invert_var.set(bool(cfg["invert"]))
            if "digit_count" in cfg:
                self.digit_count_var.set(int(cfg["digit_count"]))

            self.runner = get_model_runner(framework=framework, model_path=model_path)
            self.status_var.set(f"Model loaded: {Path(model_path).name}")
            self.pred_var.set("Prediction: model loaded")
            self.refresh_preview()
        except Exception as exc:
            messagebox.showerror("Model Error", str(exc))

    def select_roi(self) -> None:
        if self.current_image_bgr is None:
            messagebox.showerror("ROI", "Load image first.")
            return

        roi = self._select_roi_tk(self.current_image_bgr)
        if roi is None:
            return

        x, y, w, h = roi
        self.roi = (int(x), int(y), int(w), int(h))
        self.roi_text_var.set(f"{x},{y},{w},{h}")
        self.status_var.set("ROI selected")
        self.refresh_preview()

    def apply_roi_text(self) -> None:
        txt = self.roi_text_var.get().strip()
        if not txt:
            messagebox.showerror("ROI", "Enter ROI as x,y,w,h")
            return
        try:
            roi = parse_manual_roi(txt)
            self.roi = (roi.x, roi.y, roi.w, roi.h)
            self.refresh_preview()
        except Exception as exc:
            messagebox.showerror("ROI", str(exc))

    def shift_roi(self, dx: int, dy: int) -> None:
        if self.roi is None:
            return
        x, y, w, h = self.roi
        self.roi = (x + int(dx), y + int(dy), w, h)
        self.roi_text_var.set(f"{self.roi[0]},{self.roi[1]},{self.roi[2]},{self.roi[3]}")
        self.refresh_preview()

    def _draw_overlay(self, image_bgr: np.ndarray, roi: Optional[Tuple[int, int, int, int]]) -> np.ndarray:
        out = image_bgr.copy()
        if roi is None:
            return out

        x, y, w, h = roi
        cv2.rectangle(out, (x, y), (x + w, y + h), (0, 255, 0), 2)

        try:
            roi_crop = crop_roi(out, parse_manual_roi(f"{x},{y},{w},{h}"))
            cells = split_digits_equal(roi_crop, int(self.digit_count_var.get()))
            cur_x = x
            for cell in cells:
                cw = max(1, int(cell.shape[1]))
                cv2.rectangle(out, (cur_x, y), (cur_x + cw, y + h), (255, 255, 0), 1)
                cur_x += cw
        except Exception:
            pass

        return out

    def _predict_with_confidence(self, batch: np.ndarray) -> Tuple[List[int], List[float]]:
        if self.runner is None:
            return [], []

        if hasattr(self.runner, "predict_with_confidence"):
            preds, confs = self.runner.predict_with_confidence(batch)
            return list(preds), list(confs)

        if isinstance(self.runner, KerasRunner):
            x = self.runner._prepare_input(batch)
            logits = np.asarray(self.runner.model.predict(x, verbose=0))
            return logits.argmax(axis=1).astype(int).tolist(), logits.max(axis=1).astype(float).tolist()

        if isinstance(self.runner, TorchRunner):
            x = np.transpose(batch, (0, 3, 1, 2)).astype(np.float32)
            tensor = self.runner.torch.from_numpy(x)
            with self.runner.torch.no_grad():
                out = self.runner.model(tensor)
            logits = out.detach().cpu().numpy()
            exp = np.exp(logits - logits.max(axis=1, keepdims=True))
            probs = exp / exp.sum(axis=1, keepdims=True)
            return probs.argmax(axis=1).astype(int).tolist(), probs.max(axis=1).astype(float).tolist()

        preds = self.runner.predict(batch)
        return list(preds), [1.0 for _ in preds]

    def predict(self) -> None:
        if self.current_image_bgr is None:
            messagebox.showerror("Predict", "Load image first.")
            return
        if self.roi is None:
            messagebox.showerror("Predict", "Select/apply ROI first.")
            return
        if self.runner is None:
            messagebox.showerror("Predict", "Load model first.")
            return

        try:
            code, previews, preds, confs = self._predict_code_for_image(self.current_image_bgr)
            self.pred_var.set(f"Prediction: {code}")
            self._render_digits(previews, preds, confs)
            self._write_prediction_to_excel(code)
            self.refresh_preview(update_digits=False)
        except Exception as exc:
            messagebox.showerror("Predict", str(exc))

    def _predict_code_for_image(
        self, image_bgr: np.ndarray
    ) -> Tuple[str, List[np.ndarray], List[int], List[float]]:
        if self.roi is None:
            raise RuntimeError("Select/apply ROI first.")
        x, y, w, h = self.roi
        roi_crop = crop_roi(image_bgr, parse_manual_roi(f"{x},{y},{w},{h}"))
        cells = split_digits_equal(roi_crop, int(self.digit_count_var.get()))
        settings = self._settings()

        batch_list: List[np.ndarray] = []
        previews: List[np.ndarray] = []
        for cell in cells:
            arr, prev = preprocess_digit_custom(cell, settings)
            batch_list.append(arr)
            previews.append(prev)

        if not batch_list:
            raise RuntimeError("No digit cells from ROI.")

        batch = np.stack(batch_list, axis=0)
        preds, confs = self._predict_with_confidence(batch)
        code = "".join(str(int(d)) for d in preds)
        return code, previews, preds, confs

    def _write_prediction_to_excel(self, code: str) -> None:
        excel_in = self.excel_var.get().strip()
        row = self.current_row
        if not excel_in or row is None:
            return

        output_excel = self.output_excel_var.get().strip()
        if not output_excel:
            output_excel = excel_in

        in_path = Path(excel_in)
        out_path = Path(output_excel)
        if not in_path.exists():
            return

        try:
            if out_path.exists():
                wb = load_workbook(out_path)
            else:
                wb = load_workbook(in_path)

            ws = resolve_sheet(wb, self.sheet_var.get().strip() or None)
            _, out_col = self._resolve_cols(ws)
            ws.cell(row=int(row), column=out_col).value = code
            wb.save(out_path)
            wb.close()
            self.status_var.set(f"Predicted value written to row {row} in {out_path.name}")
        except Exception as exc:
            raise RuntimeError(f"Prediction done, but Excel write failed: {exc}") from exc

    def run_all_urls(self) -> None:
        if self._run_all_active:
            return
        if self.runner is None:
            messagebox.showerror("Run All", "Load model first.")
            return
        if self.roi is None:
            messagebox.showerror("Run All", "Select/apply ROI first.")
            return

        self._run_all_active = True
        try:
            if not self.row_sequence:
                self.row_sequence = self._collect_rows_from_excel()
            if not self.row_sequence:
                raise RuntimeError("No URLs found. Check Excel/columns/start row and click Scan URLs.")

            excel_in = self.excel_var.get().strip()
            if not excel_in:
                raise RuntimeError("Select input Excel first.")
            output_excel = self.output_excel_var.get().strip() or excel_in
            in_path = Path(excel_in)
            out_path = Path(output_excel)

            if out_path.exists():
                wb = load_workbook(out_path)
            else:
                wb = load_workbook(in_path)

            ws = resolve_sheet(wb, self.sheet_var.get().strip() or None)
            url_col, out_col = self._resolve_cols(ws)
            overwrite = bool(self.overwrite_var.get())

            processed = 0
            success = 0
            skipped = 0
            failed = 0
            session = requests.Session()
            total = len(self.row_sequence)
            try:
                for idx, row in enumerate(self.row_sequence, start=1):
                    try:
                        source = ws.cell(row=row, column=url_col).value
                        if not source or not str(source).strip():
                            skipped += 1
                            continue
                        existing = ws.cell(row=row, column=out_col).value
                        if existing is not None and str(existing).strip() and not overwrite:
                            skipped += 1
                            continue

                        image = decode_image_from_source(
                            source=str(source).strip(),
                            timeout=20,
                            retries=2,
                            retry_backoff=1.5,
                            session=session,
                        )
                        code, _previews, _preds, _confs = self._predict_code_for_image(image)
                        ws.cell(row=row, column=out_col).value = code
                        processed += 1
                        success += 1
                    except Exception:
                        failed += 1
                    self.status_var.set(f"Run all: {idx}/{total} (row {row})")
                    self.update_idletasks()
                wb.save(out_path)
            finally:
                session.close()
                wb.close()

            messagebox.showinfo(
                "Run All Complete",
                (
                    f"Done.\nProcessed: {processed}\nSuccess: {success}\n"
                    f"Skipped: {skipped}\nFailed: {failed}\nOutput: {out_path}"
                ),
            )
        except Exception as exc:
            messagebox.showerror("Run All", str(exc))
        finally:
            self._run_all_active = False

    def _render_digits(self, previews: Sequence[np.ndarray], preds: Sequence[int], confs: Sequence[float]) -> None:
        for child in self.digit_frame.winfo_children():
            child.destroy()
        self.digit_photo_refs.clear()

        for i, d in enumerate(previews):
            p = self._to_photo(d, max_size=(86, 86))
            self.digit_photo_refs.append(p)
            lbl = ttk.Label(self.digit_frame, image=p)
            lbl.grid(row=0, column=i, padx=2, pady=2)

            txt = str(i)
            if i < len(preds):
                txt += f" | {preds[i]}"
            if i < len(confs):
                txt += f" ({confs[i]:.2f})"
            ttk.Label(self.digit_frame, text=txt).grid(row=1, column=i)

    def refresh_preview(self, update_digits: bool = True) -> None:
        if self.current_image_bgr is None:
            return

        overlay = self._draw_overlay(self.current_image_bgr, self.roi)
        full_photo = self._photo_for_widget(overlay, self.full_img_label, fallback=(740, 430))
        self.full_img_label.configure(image=full_photo, text="")
        self.photo_refs["full"] = full_photo

        if self.roi is not None:
            try:
                x, y, w, h = self.roi
                roi_preview = crop_roi(self.current_image_bgr, parse_manual_roi(f"{x},{y},{w},{h}"))
                roi_photo = self._photo_for_widget(roi_preview, self.roi_img_label, fallback=(740, 430))
                self.roi_img_label.configure(image=roi_photo, text="")
                self.photo_refs["roi"] = roi_photo
            except Exception:
                self.roi_img_label.configure(text="Invalid ROI", image="")
        else:
            self.roi_img_label.configure(text="Select ROI", image="")

        if update_digits and self.roi is None:
            for child in self.digit_frame.winfo_children():
                child.destroy()
            self.digit_photo_refs.clear()


def main() -> None:
    app = UDISEMinimalApp()
    app.mainloop()


if __name__ == "__main__":
    main()
